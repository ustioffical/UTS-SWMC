import json
import datetime
import requests
import xlwt
import pandas as pd
from django.forms.models import model_to_dict
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta, time
from django.utils import timezone
from django.utils.timezone import get_current_timezone, make_aware, is_naive, localdate, localtime
from django.views.decorators.http import require_http_methods, require_GET
from django.shortcuts import redirect
from django.core.serializers.json import DjangoJSONEncoder
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils.dateparse import parse_datetime

from django.db.models import Count, Sum, Value, Min, Max, F, Q, ExpressionWrapper, DurationField, FloatField, \
    IntegerField, Func, CharField, DateTimeField, Case, When, Subquery, OuterRef
from django.db.models.expressions import RawSQL
from django.db.models.functions import Cast, Coalesce, Concat, Now, Extract, TruncDate

from django.contrib.gis.db.models.aggregates import Collect, MakeLine
from django.contrib.gis.db.models.functions import Length

from django.contrib.gis.geos import GEOSGeometry

from urllib.parse import urlencode
from collections import Counter, defaultdict
from django.contrib import messages
from openpyxl.utils import get_column_letter
from datetime import timedelta
from itertools import groupby
from operator import attrgetter
from decimal import Decimal

from io import BytesIO
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST

from AppAdmin.utils import *
from AppVehicle.models import *

from AppSetting.models import TownBoundary
from VTMS.views import *
from AppRoute.views import *


# Create your views here.
def date_handler(obj):
    return obj.isoformat() if hasattr(obj, 'isoformat') else str(obj)


def AllVehicleMonitoringView(request):
    message = ""
    cursor = connections['default'].cursor()
    template_name = "VehicleMonitoring.html"
    dateTime = datetime.datetime.now()

    format_date = '%Y-%m-%d'
    today_date = dateTime.strftime(format_date)

    get_vehicle_status = request.GET.get('cmd_vehicle_status', 'NA')
    get_vehicle_type = request.GET.get('cmd_vehicle_type', 'NA')
    get_vehicle_code = request.GET.get('cmd_vehicle_list', 'NA')

    if request.method == "POST":
        get_vehicle_status = request.POST.get('cmd_vehicle_status')
        get_vehicle_type = request.POST.get('cmd_vehicle_type')
        get_vehicle_code = request.POST.get('cmd_vehicle_list')

    ### APPLY FILTER ON QUERY
    filters = {}
    filters['vehicle_code__status'] = "Active"

    filters_vehicle_data = {}
    filters_vehicle_data['status'] = "Active"
    if get_vehicle_status not in (None, "NA"):
        filters['g_status'] = get_vehicle_status
        filters_vehicle_data['live_monitor__g_status'] = get_vehicle_status

    if get_vehicle_type not in (None, "NA"):
        filters_vehicle_data['vehicle_type'] = get_vehicle_type

    ### VEHICLE STATUS LIST
    vehicle_g_status = list(
        VehicleLiveMonitor.objects
        .filter(vehicle_code__status="Active")  # 👈 join filter on VehicleData
        .values('g_status')
        .annotate(count=Count('id'))
    )

    ### VEHICLE TYPE LIST WITH VEHICLE STATUS
    vehicle_type_list = (
        VehicleLiveMonitor.objects
        .filter(**filters)
        .values(vehicle_type=F('vehicle_code__vehicle_type'))
        .annotate(count=Count('vehicle_code__vehicle_type'))
        .order_by('vehicle_code__vehicle_type')
    )

    ### VEHICLE DATA DETAIL WITH VEHICLE STATUS AND VEHICLE TYPE
    vehicle_data_list = list(
        VehicleData.objects.filter(**filters_vehicle_data)
        .values('vehicle_code', 'vehicle_type', 'register_no', 'chasis_no')
        .annotate(
            pitb_code=Coalesce('pitb_code', Value('Waiting'), output_field=CharField())
        )
        .order_by('register_no')
        .distinct()
    )

    query_feature = f"""
    WITH veh_live AS (
        SELECT
            ST_X(geom) AS x,
            ST_Y(geom) AS y,
            chasis_no,
            register_no,
            vehicle_code_id,
            COALESCE(pitb_code, 'Waiting') AS pitb_code,
            vehicle_type,
            g_status,
            direction,
            speed,
            COALESCE(device_status, 'NA') AS device_status,
            COALESCE(ignition_status, 'NA') AS ignition_status,
            COALESCE(geo_location, 'NA') AS geo_location,
            vendor_date_time,
            COALESCE(duration, 'NA') AS duration
        FROM
            tbl_vehicle_live_monitor AS vlm
        LEFT OUTER JOIN
            tbl_vehicle_data AS veh
        ON
            vlm.vehicle_code_id = veh.vehicle_code WHERE veh.status = 'Active'
    )
    SELECT
        x,
        y,
        chasis_no,
        register_no,
        veh_live.vehicle_code_id,
        pitb_code,
        vehicle_type,
        g_status,
        direction,
        speed,
        device_status,
        ignition_status,
        geo_location,
        vendor_date_time,
        COALESCE(distance, 0) AS distance,
        COALESCE(working_hours, 0) AS working_hours
    FROM
        veh_live
    LEFT OUTER JOIN (
        SELECT
            vehicle_code_id,
            COALESCE(distance, 0) AS distance,
            COALESCE(working_hours, 0) AS working_hours
        FROM
            tbl_vehicle_schedule_gprs_api
        WHERE
            veh_sch_date::date = '{today_date}'
    ) AS veh_sche
    ON
        veh_live.vehicle_code_id = veh_sche.vehicle_code_id
    WHERE
        1=1
    """

    if get_vehicle_status != "NA":
        query_feature += f" AND g_status = '{get_vehicle_status}'"

    if get_vehicle_type != "NA":
        query_feature += f" AND vehicle_type = '{get_vehicle_type}'"

    if get_vehicle_code != "NA":
        query_feature += f" AND (veh_live.vehicle_code_id = '{get_vehicle_code}')"

    query_feature += """
    ORDER BY
        distance DESC,
        vendor_date_time DESC,
        duration;
    """

    cursor.execute(query_feature)
    vehicle_live_lists = DictinctFetchAll(cursor)

    ### GET TOTAL VEHICLE DIFFERENCE FROM API BY VEHICLE DATA
    api_vehicle_diff = 0

    town_boundary = TownBoundary.objects.all().order_by('town_name')

    ### ASSET TYPE SUMMARY WITH COUNT
    query_asset_type = "SELECT 'Container' as name, count(*) AS value FROM tbl_container_data UNION ALL SELECT 'Drum' as name, count(*) AS value FROM tbl_drum_data UNION ALL SELECT 'Collection-Site' as name, count(*) AS value FROM tbl_collection_site UNION ALL SELECT 'Dumping-Site' as name, count(*) AS value FROM tbl_dumping_site UNION ALL SELECT 'Landuse' as name, count(*) AS value FROM tbl_landuse_boundary UNION ALL SELECT 'Admin' as name, count(*) AS value FROM tbl_administrative_boundary; ";
    cursor.execute(query_asset_type)
    asset_type_summary = DictinctFetchAll(cursor)

    page_title = "Create Network"
    params = {
        'vehicle_g_status': vehicle_g_status,
        'vehicle_type_list': vehicle_type_list,
        'vehicle_data_list': vehicle_data_list,
        'selected_vehicle_status': get_vehicle_status,
        'selected_vehicle_type': get_vehicle_type,
        'selected_vehicle_code': get_vehicle_code,
        'vehicle_live_lists': vehicle_live_lists,
        'town_boundary': town_boundary,
        'api_vehicle_diff': api_vehicle_diff,
        'asset_type_summary': asset_type_summary,
        'page_title': page_title,
        'message': message,
    }

    return render(request, template_name, params)


def AllVehicleSyncView(request):
    message = ""
    cursor = connections['default'].cursor()
    template_name = "VehicleSync.html"
    dateTime = datetime.datetime.now()

    if request.method == "POST":
        get_action_type = request.POST['action_type']
        if get_action_type == "vehicle-sync":

            vehicle_codes = VehicleData.objects.all().values('id',
                                                             'vehicle_code')  # or use another field instead of 'id'

            # Build initial dictionary with default status "No"
            prev_vehicle_data = {
                f"[{v['vehicle_code']}]": {
                    "id": v["id"],
                    "vehicle_code": v["vehicle_code"],
                    "status": "No"
                }
                for v in vehicle_codes
            }

            ### Base URL
            ### GET RESPONSE FROM VEHICLE API FUNCTION
            response_data = ResponseVehicleApi_By_Vendor_Function()

            ### API RESPONSE CHECK ###
            if len(response_data) > 0:  ### RECORD EXIST ###

                vehicle_response_api = response_data['MainData']
                for i in range(len(vehicle_response_api)):

                    # CREATE AND UPDATE VEHICLE (START)
                    ch_system_vehicle_id = vehicle_response_api[i]['VehicleId']
                    ch_vendor_vehicle_id = vehicle_response_api[i]['VehID']
                    api_terminal_id = vehicle_response_api[i]['TerminalNo']

                    ### FETCH VEHICLE DATA
                    VehicleStatus = VehicleData.objects.filter(vehicle_code=ch_system_vehicle_id)
                    if len(VehicleStatus) > 0:  # UPDATE

                        ### UPDATE PREVIOUS VEHICLE DATA OBJECT
                        # Update status for selected vehicle
                        key = f"[{ch_system_vehicle_id}]"
                        if key in prev_vehicle_data:
                            prev_vehicle_data[key]["status"] = "Yes"

                        ### UPDATE VEHICLE DATA IN DB (START)
                        VehicleData.objects.filter(vehicle_code=ch_system_vehicle_id).update(
                            register_no=vehicle_response_api[i]['CarReg'],
                            make=vehicle_response_api[i]['Make'],
                            engine_no=vehicle_response_api[i]['EngNo'],
                            chasis_no=vehicle_response_api[i]['ChaNo'],
                            color=vehicle_response_api[i]['Color'],
                            model=vehicle_response_api[i]['Model'],
                            cc=vehicle_response_api[i]['CC'],
                            fuel_type=vehicle_response_api[i]['FuelType'],
                            vehicle_type=vehicle_response_api[i]['VehicleType'],
                            status=vehicle_response_api[i]['VehicleStatus'],

                            mileage_cur_value=vehicle_response_api[i]['mileage_cur_value'],
                            date_installed=datetime.datetime.strptime(vehicle_response_api[i]['DateInstalled'],
                                                                      "%b %d, %Y"),
                            installation_date=vehicle_response_api[i]['InstallationDate'],
                            year=vehicle_response_api[i]['MYear'],
                            engin_temp=vehicle_response_api[i]['engin_temp'],
                            engine_hours=vehicle_response_api[i]['engine_hours'],
                            fuel_level=vehicle_response_api[i]['fuel_level'],
                            fuel_consumed=vehicle_response_api[i]['fuel_consumed'],
                            ext_bat_voltage=vehicle_response_api[i]['ext_bat_voltage'],
                            int_bat_voltage=vehicle_response_api[i]['int_bat_voltage'],
                            veh_status_chg_sec=vehicle_response_api[i]['VehStatusChgSec'],

                            updated_at=dateTime,
                            updated_by="admin"
                        )
                        ### UPDATE VEHICLE (END)

                        ### STEP-2 INSTANCE AND UPDATE TELECOM DATA
                        telecomdata = TelecomData.objects.filter(sim_no=vehicle_response_api[i]['SimNo']).exists()

                        if not telecomdata:
                            InsTelecomData = TelecomData(
                                sim_no=vehicle_response_api[i]['SimNo'],
                                gsm_co=vehicle_response_api[i]['GSMCo'],
                                connected='Yes',
                                created_at=dateTime,
                                created_by='admin'
                            )
                            InsTelecomData.save()
                        else:
                            TelecomData.objects.filter(sim_no=vehicle_response_api[i]['SimNo']).update(
                                gsm_co=vehicle_response_api[i]['GSMCo'],
                                updated_by="admin",
                                updated_at=dateTime
                            )
                        ### STEP-2 INSTANCE AND UPDATE TELECOM DATA (END)

                        ### STEP-3 INSTANCE AND UPDATE TRACKER DATA
                        trackerData = TrackerData.objects.filter(
                            terminal_code=api_terminal_id).exists()
                        if not trackerData:
                            InsTrackerData = TrackerData(
                                terminal_code=api_terminal_id,
                                tracker_company_code=vehicle_response_api[i]['ServerDetail'],
                                minutes_diff=vehicle_response_api[i]['MinutesDiff'],
                                ins_co=vehicle_response_api[i]['InsCo'],
                                region=vehicle_response_api[i]['Region'],
                                group_title=vehicle_response_api[i]['GroupTitle'],
                                gps_satelite=vehicle_response_api[i]['gps_satelite'],
                                gsm_signal=vehicle_response_api[i]['gsm_signal'],
                                sale_type_name=vehicle_response_api[i]['SaleTypeName'],
                                terminal_type="installed",
                                created_by="admin"
                            )
                            InsTrackerData.save()
                        else:
                            TrackerData.objects.filter(terminal_code=api_terminal_id).update(
                                tracker_company_code=vehicle_response_api[i]['ServerDetail'],
                                minutes_diff=vehicle_response_api[i]['MinutesDiff'],
                                ins_co=vehicle_response_api[i]['InsCo'],
                                region=vehicle_response_api[i]['Region'],
                                group_title=vehicle_response_api[i]['GroupTitle'],
                                gps_satelite=vehicle_response_api[i]['gps_satelite'],
                                gsm_signal=vehicle_response_api[i]['gsm_signal'],
                                sale_type_name=vehicle_response_api[i]['SaleTypeName'],
                                terminal_type="installed",
                                updated_by="admin",
                                updated_at=dateTime
                            )
                        ### STEP-3 INSTANCE AND UPDATE TRACKER DATA (END)

                        # Terminal Change and Sim Shange logic if exist and generate log
                        # Check for existing tracker

                        ### STEP-4 INSTANCE AND UPDATE(BLOCK) TRACKER VEHICLE DATA
                        ex_tracker_vehicle = TrackerVehicleData.objects.filter(
                            vehicle_code=ch_system_vehicle_id, status='Active').first()

                        if ex_tracker_vehicle:
                            connected_terminal = ex_tracker_vehicle.terminal_code.terminal_code
                            connected_sim = TrackerTelecomData.objects.filter(
                                terminal=ex_tracker_vehicle).last().sim.sim_no if TrackerTelecomData.objects.filter(
                                terminal=ex_tracker_vehicle).exists() else None

                            new_terminal = api_terminal_id
                            new_sim = vehicle_response_api[i]['SimNo']

                            # Check if terminal or sim changed (True / False)
                            terminal_changed = connected_terminal and connected_terminal != new_terminal
                            sim_changed = connected_sim and connected_sim != new_sim

                            # If either changed, log it
                            if terminal_changed:  # IF TRUE THEN CHANGE
                                # Block old terminal tracker status
                                if ex_tracker_vehicle:
                                    ex_tracker_vehicle.status = "Blocked"
                                    ex_tracker_vehicle.terminal_code.terminal_type = "advanced"
                                    ex_tracker_vehicle.save()
                                tracker = TrackerData.objects.filter(terminal_code=connected_terminal).first()
                                if tracker:
                                    tracker.terminal_type = ""
                                    tracker.save()

                            ### TELECOM SIM STATUS CHANGE
                            if sim_changed:
                                sim = TelecomData.objects.filter(sim_no=connected_sim).first()
                                if sim:
                                    sim.connected = "No"
                                    sim.save()

                            if terminal_changed or sim_changed:
                                # Block old TerminalTelecomData records
                                TrackerTelecomData.objects.filter(
                                    terminal=connected_terminal,
                                    sim__sim_no=connected_sim,
                                    status="Active"
                                ).update(status="Blocked")

                        try:
                            trackerVehicleData = TrackerVehicleData.objects.filter(
                                terminal_code=api_terminal_id).exists()
                            if not trackerVehicleData:
                                TrackerVehicleData.objects.create(
                                    vehicle_code_id=ch_system_vehicle_id,
                                    terminal_code_id=api_terminal_id,
                                    status="Active",
                                    created_at=dateTime,
                                    created_by='admin'
                                )
                            else:
                                TrackerVehicleData.objects.filter(terminal_code=api_terminal_id).update(
                                    vehicle_code_id=ch_system_vehicle_id,
                                    terminal_code_id=api_terminal_id,
                                    status="Active",
                                    updated_at=dateTime,
                                    updated_by='admin'
                                )

                        except Exception as e:
                            print(f"Error creating TrackerVehicleData: {str(e)}")

                        # Get the TrackerData and TelecomData instances
                        tracker_instance = TrackerVehicleData.objects.filter(
                            terminal_code=api_terminal_id).first()
                        sim_instance = TelecomData.objects.filter(sim_no=vehicle_response_api[i]['SimNo']).first()

                        if tracker_instance and sim_instance:
                            obj, trackertelecomdatacreated = TrackerTelecomData.objects.update_or_create(
                                terminal=tracker_instance,
                                sim=sim_instance,
                                defaults={
                                    'assigned_at': dateTime,
                                    'assigned_by': 'admin',
                                    'status': 'Active',
                                }
                            )
                            if not trackertelecomdatacreated:
                                obj.changed_by = "admin"
                                obj.changed_at = dateTime
                                obj.save()
                        else:
                            print(
                                f"Tracker or Sim not found for Terminal: {api_terminal_id}, Sim: {vehicle_response_api[i]['SimNo']}")

                        ### IF VEHICLE EXIST END

                        # VEHICLE LIVE MONITORING UPDATE (START)
                        auto_gprs_live_code = AutoGenerateCodeForModel(VehicleLiveMonitor, "veh_live_mont_code",
                                                                       "GPSL-")
                        get_feature_coordinate = "POINT(" + str(vehicle_response_api[i]['Long']) + " " + str(
                            vehicle_response_api[i][
                                'Lat']) + ")"

                        ck_vendor_date_time = str(vehicle_response_api[i]['GPSTime'])
                        format_str = "%Y-%m-%d %H:%M:%S"
                        try:
                            if '.' in ck_vendor_date_time:
                                ck_vendor_date_time = ck_vendor_date_time.split('.')[0]  # Remove the decimal part

                            _ck_vendor_date_time = datetime.datetime.fromisoformat(ck_vendor_date_time)
                            formatted_vendor_date_time = _ck_vendor_date_time.strftime(format_str)  # Format as a string
                            # print("Valid ISO format:")
                        except ValueError:
                            # Parse into datetime object
                            formatted_vendor_date_time = datetime.datetime.strptime(ck_vendor_date_time, format_str)

                        VehicleLiveMonitor.objects.update_or_create(
                            vehicle_code_id=ch_system_vehicle_id,
                            defaults={
                                'veh_live_mont_code': auto_gprs_live_code,
                                'geom': get_feature_coordinate,
                                'latitude': vehicle_response_api[i]['Lat'],
                                'longitude': vehicle_response_api[i]['Long'],
                                'g_status': vehicle_response_api[i]['GStatus'],
                                'speed': vehicle_response_api[i]['Speed'],
                                'device_status': vehicle_response_api[i]['dev_status'],
                                'direction': vehicle_response_api[i]['Direction'],
                                'ignition_status': vehicle_response_api[i]['IgnStatus'],
                                'geo_location': vehicle_response_api[i]['Location'],
                                'vendor_date_time': formatted_vendor_date_time,
                                'duration': vehicle_response_api[i]['Duration'],
                                'updated_at': dateTime,
                                'updated_by': "admin",
                                'ext_bat_voltage': vehicle_response_api[i]['ext_bat_voltage'],
                                'int_bat_voltage': vehicle_response_api[i]['int_bat_voltage'],
                                'gsm_signal': vehicle_response_api[i]['gsm_signal'],
                            }
                        )
                        # VEHICLE LIVE MONITORING UPDATE (END)

                    else:  # CREATE NEW VEHICLE DATA

                        ### INSERT NEW VEHICLE DATA IN DB (START)
                        InstVehicleData = VehicleData(
                            vehicle_code=ch_system_vehicle_id,
                            vendor_code=ch_vendor_vehicle_id,
                            vehicle_type=vehicle_response_api[i]['VehicleType'],
                            register_no=vehicle_response_api[i]['CarReg'],
                            make=vehicle_response_api[i]['Make'],
                            engine_no=vehicle_response_api[i]['EngNo'],
                            chasis_no=vehicle_response_api[i]['ChaNo'],
                            color=vehicle_response_api[i]['Color'],
                            model=vehicle_response_api[i]['Model'],
                            cc=vehicle_response_api[i]['CC'],
                            fuel_type=vehicle_response_api[i]['FuelType'],
                            status=vehicle_response_api[i]['VehicleStatus'],

                            mileage_cur_value=vehicle_response_api[i]['mileage_cur_value'],
                            date_installed=datetime.datetime.strptime(vehicle_response_api[i]['DateInstalled'],
                                                                      "%b %d, %Y"),
                            installation_date=vehicle_response_api[i]['InstallationDate'],
                            year=vehicle_response_api[i]['MYear'],
                            engin_temp=vehicle_response_api[i]['engin_temp'],
                            engine_hours=vehicle_response_api[i]['engine_hours'],
                            fuel_level=vehicle_response_api[i]['fuel_level'],
                            fuel_consumed=vehicle_response_api[i]['fuel_consumed'],
                            ext_bat_voltage=vehicle_response_api[i]['ext_bat_voltage'],
                            int_bat_voltage=vehicle_response_api[i]['int_bat_voltage'],
                            veh_status_chg_sec=vehicle_response_api[i]['VehStatusChgSec'],

                            created_at=dateTime,
                            # created_by=request.session['username']
                            created_by="admin"
                        )
                        InstVehicleData.save()
                        # CREATE VEHICLE (END)

                        InsTrackerData = TrackerData(
                            terminal_code=api_terminal_id,
                            tracker_company_code=vehicle_response_api[i]['ServerDetail'],
                            minutes_diff=vehicle_response_api[i]['MinutesDiff'],
                            ins_co=vehicle_response_api[i]['InsCo'],
                            region=vehicle_response_api[i]['Region'],
                            group_title=vehicle_response_api[i]['GroupTitle'],
                            gps_satelite=vehicle_response_api[i]['gps_satelite'],
                            gsm_signal=vehicle_response_api[i]['gsm_signal'],
                            sale_type_name=vehicle_response_api[i]['SaleTypeName'],
                            terminal_type="Installed",
                            created_by="admin"
                        )
                        InsTrackerData.save()

                        # VEHICLE LIVE MONITORING (START)
                        auto_gprs_live_code = AutoGenerateCodeForModel(VehicleLiveMonitor, "veh_live_mont_code",
                                                                       "GPSL-")
                        get_feature_coordinate = "POINT(" + str(vehicle_response_api[i]['Long']) + " " + str(
                            vehicle_response_api[i][
                                'Lat']) + ")"

                        ck_vendor_date_time = str(vehicle_response_api[i]['GPSTime'])
                        format_str = "%Y-%m-%d %H:%M:%S"
                        try:
                            if '.' in ck_vendor_date_time:
                                ck_vendor_date_time = ck_vendor_date_time.split('.')[0]  # Remove the decimal part

                            _ck_vendor_date_time = datetime.datetime.fromisoformat(ck_vendor_date_time)
                            formatted_vendor_date_time = _ck_vendor_date_time.strftime(format_str)  # Format as a string
                            # print("Valid ISO format:")
                        except ValueError:
                            # Parse into datetime object
                            formatted_vendor_date_time = datetime.datetime.strptime(ck_vendor_date_time, format_str)

                        InstVehicleLive = VehicleLiveMonitor(
                            veh_live_mont_code=auto_gprs_live_code,
                            vehicle_code_id=ch_system_vehicle_id,
                            geom=get_feature_coordinate,
                            latitude=vehicle_response_api[i]['Lat'],
                            longitude=vehicle_response_api[i]['Long'],
                            g_status=vehicle_response_api[i]['GStatus'],
                            speed=vehicle_response_api[i]['Speed'],
                            device_status=vehicle_response_api[i]['dev_status'],
                            direction=vehicle_response_api[i]['Direction'],
                            ignition_status=vehicle_response_api[i]['IgnStatus'],
                            geo_location=vehicle_response_api[i]['Location'],
                            vendor_date_time=formatted_vendor_date_time,
                            duration=vehicle_response_api[i]['Duration'],
                            created_at=dateTime,
                            created_by="admin"
                        )
                        InstVehicleLive.save()
                        # VEHICLE LIVE MONITORING (END)
                    ### ELSE END (CREATE)

                ### LOOP END
            ### GREATER THEN 0

        ## CUSTOMER INFORMATION (START) (END)
        ## DRIVER INFORMATION (START) (END)

        ## VEHICLE BLOCK WHICH IS NOT EXIST IN VENDOR API DATA
        for key, value in prev_vehicle_data.items():
            if value["status"] == "No":
                vehicle = VehicleData.objects.filter(vehicle_code=str(value['vehicle_code'])).first()
                if vehicle and vehicle.status != "Block":
                    vehicle.status = "Block"
                    vehicle.save()
        ### REQUEST METHOD END

    ### ASSET TYPE SUMMARY WITH COUNT
    query_asset_type = "SELECT 'Container' as name, count(*) AS value FROM tbl_container_data UNION ALL SELECT 'Drum' as name, count(*) AS value FROM tbl_drum_data UNION ALL SELECT 'Collection-Site' as name, count(*) AS value FROM tbl_collection_site UNION ALL SELECT 'Dumping-Site' as name, count(*) AS value FROM tbl_dumping_site UNION ALL SELECT 'Landuse' as name, count(*) AS value FROM tbl_landuse_boundary UNION ALL SELECT 'Admin' as name, count(*) AS value FROM tbl_administrative_boundary; ";
    cursor.execute(query_asset_type)
    asset_type_summary = DictinctFetchAll(cursor)

    vehicle_list = VehicleData.objects.all().order_by('vehicle_type')

    page_title = "Vehicle Management"
    params = {
        'vehicle_list': vehicle_list,
        'asset_type_summary': asset_type_summary,
        'page_title': page_title,
        'message': message,
    }

    return render(request, template_name, params)


def AllVehicleManagementView(request):
    message = ""

    template_name = "VehicleManagement.html"
    current_date = localdate()
    current_time = localtime()

    # Get filter parameters
    vehicle_type = request.POST.get("cmd_vehicle_type")
    device_status_param = request.POST.get("cmd_device_status", "all")
    vehicle_status = request.POST.get("cmd_vehicle_status")
    vehicle_code = request.POST.get("cmd_vehicle_list")

    chart_vehicles = VehicleLiveMonitor.objects.select_related('vehicle_code').filter(
        vehicle_code__status="Active").all()

    # Initialize filters
    filter_vehicles = chart_vehicles
    filters = Q()

    # query_feature += f" AND (veh.vehicle_code LIKE '%{search_term}%') "
    ### CONVERT RAW SWL INTO THIS CODE
    # Apply vehicle type filter
    if vehicle_type and vehicle_type != "NA":
        filters &= Q(vehicle_code__vehicle_type=vehicle_type)

    # Apply vehicle code filter
    if vehicle_code and vehicle_code != "NA":
        filters &= Q(vehicle_code__vehicle_code=vehicle_code)

    # Apply device status filter
    if device_status_param != "all":
        filters &= Q(vendor_date_time__contains=current_date.strftime('%Y-%m-%d'))

    # Apply vehicle status filter DAYWISE AND STATUS WISE (MOVIE, IDLE AND I DAY ETC)
    if vehicle_status and vehicle_status != "NA":
        if "Day" in vehicle_status:  # No-response days filter
            days = int(vehicle_status.split()[0])
            target_date = current_date - timedelta(days=days)
            filters &= Q(vendor_date_time__contains=target_date.strftime('%Y-%m-%d'))
        else:  # G-status filter
            filters &= Q(g_status=vehicle_status)
    # Apply all filters
    vehicles = filter_vehicles.filter(filters)

    # Handle date filters
    start_dt = current_date
    end_dt = current_date
    start_dt_report = end_dt - timedelta(days=1)

    chart_data = process_chart_data(chart_vehicles, current_date, current_time)

    ### FUNCTION TO PROCESS VEHICLE MANAGEMENT TABLE IN TEMPLATE
    current_vehicles = ProcessTableData_VehicleManagement_Function(vehicles, current_date, current_time)

    # --- Filter Data ---
    vehicle_type_list = VehicleData.objects.filter(status="Active").values('vehicle_type').annotate(
        count=Count('id')).order_by('vehicle_type')

    working_count = 0
    no_response_count = 0

    for vehicle in current_vehicles['current_vehicles']:
        if vehicle['device_status'] == "Working":
            working_count += 1
        elif vehicle['device_status'] == "No Response":
            no_response_count += 1

    print("Working:", working_count)
    print("No Response:", no_response_count)

    device_status = [
        ("all", "All"),
        ("working", f"Working - {working_count}"),
        ("no_response", f"No Response - {no_response_count}")
    ]
    working_status_options = list(chart_data['current_working'].items())  # ['Parked', 'Idle']

    no_response_options = list(chart_data['no_responsed'].items())  # ['1 Day', '2 Day']

    # live_monitor__
    g_statuses = VehicleLiveMonitor.objects.filter(vehicle_code__status="Active").values_list('g_status',
                                                                                              flat=True).distinct()

    ### DB field is really an integer
    chart_vehicle_ids = chart_vehicles.values_list('vehicle_code_id', flat=True)
    # Cast to int:
    chart_vehicle_ids = [int(x) for x in chart_vehicle_ids if x is not None]

    vehicle_codes = list(VehicleData.objects.filter(
        vehicle_code__in=chart_vehicle_ids,
        status="Active"
    ).values('vehicle_code', 'vehicle_type', 'register_no', 'pitb_code').order_by('vehicle_type'))

    page_title = "Vehicle Management"
    params = {
        'page_title': page_title,
        'current_date': current_date,
        'message': message,
        "chart_data": json.dumps(chart_data),
        "all_vehicles_today_count": chart_data['current_count'],
        "vendor_date_groups_count": chart_data['no_response_count'],
        "total_vehicles_count": chart_data['total_count'],
        "current_vehicles": current_vehicles['current_vehicles'],
        "vehicle_type_list": vehicle_type_list,
        "device_status": device_status,
        "working_status_options": working_status_options,
        "no_response_options": no_response_options,
        "g_statuses": g_statuses,
        "vehicle_codes": vehicle_codes,
        "request": request,
        "start_date": start_dt,
        "end_date": end_dt,
        "start_date_report": start_dt_report,
        # Selected filter values for maintaining state
        "selected_device_status": device_status_param,
        "selected_status": vehicle_status,
        "selected_vehicle_type": vehicle_type,
        "selected_vehicle_code": vehicle_code,
    }

    return render(request, template_name, params)


def NoResponseVehicleView(request):
    current_date = localdate()
    vehicles = VehicleLiveMonitor.objects.select_related('vehicle_code').filter(vehicle_code__status="Active").order_by(
        '-vendor_date_time')

    all_vehicles = []  # This will contain all vehicles with their no-response issues for the table
    issue_categories = {
        'battery_issue': [],
        'battery_drained': [],
        'battery_disconnected': [],
        'telecom_issue': [],
        'sim_issue': [],
        'internet_issue': []
    }

    for v in vehicles:
        terminal = TrackerVehicleData.objects.filter(vehicle_code_id=v.vehicle_code).first()
        terminal_id = terminal.terminal_code if terminal else "N/A"
        # Get the sim number
        sim_no = None
        if terminal:
            tracker_telecom = TrackerTelecomData.objects.filter(terminal=terminal, status='Active').order_by(
                '-assigned_at').first()
            if tracker_telecom and tracker_telecom.sim:
                sim_no = tracker_telecom.sim.sim_no
        print(v.vendor_date_time)
        vdt = parse_vendor_date(v.vendor_date_time)

        if vdt.date() != current_date:

            nr_days = (current_date - vdt.date()).days if vdt.date() < current_date else 0
            vehicle_data = {
                'vehicle': v,
                'nr_days': nr_days,
                'vendor_date_time': vdt,
                'terminal_id': terminal_id,
                'sim_no': sim_no,
                'reason': '',
                'subreason': '',
                'value': None
            }

            # Check battery issues
            if 1 <= v.ext_bat_voltage < 9:
                vehicle_data['reason'] = 'Battery'
                vehicle_data['subreason'] = 'Battery Drained'
                vehicle_data['value'] = v.ext_bat_voltage
                issue_categories['battery_drained'].append(vehicle_data)
                issue_categories['battery_issue'].append(vehicle_data)

            elif v.ext_bat_voltage == 0:
                vehicle_data['reason'] = 'Battery'
                vehicle_data['subreason'] = 'Battery Disconnected'
                vehicle_data['value'] = v.ext_bat_voltage
                issue_categories['battery_disconnected'].append(vehicle_data)
                issue_categories['battery_issue'].append(vehicle_data)

            # Check telecom/signal issues
            else:
                vehicle_data['reason'] = 'Telecom'
                vehicle_data['subreason'] = 'Telecom'
                vehicle_data['value'] = v.gsm_signal
                issue_categories['telecom_issue'].append(vehicle_data)

            if vehicle_data['reason']:  # Only add vehicles with no-response issues
                all_vehicles.append(vehicle_data)

    # Check if Excel download is requested
    if request.POST.get('export') == 'excel':
        # Prepare data for Excel
        data = []
        for vehicle in all_vehicles:
            data.append({
                'Vehicle Code': vehicle['vehicle'].vehicle_code,
                'Terminal No': vehicle['terminal_id'],
                'PITB Code': vehicle['vehicle'].vehicle_code.pitb_code,
                'Battery Voltage': vehicle['value'],
                'Sim No': vehicle['sim_no'],
                'Reason': vehicle['reason'],
                'Sub-Reason': vehicle['subreason'],
                'Last Update': vehicle['vendor_date_time'].strftime('%Y-%m-%d %H:%M:%S'),
                'Days Not Responding': vehicle['nr_days']
            })

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, sheet_name='Terminals', index=False)
        writer.close()
        output.seek(0)

        # Prepare response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=NoResponseVehicles.xlsx'
        return response

    # Get filter from request
    filter_type = request.POST.get('filter', None)

    # Determine which vehicles to display in Table after applying filters of Advanced/Installed Terrminals
    if filter_type and filter_type in issue_categories:
        display_vehicles = issue_categories[filter_type]
    else:
        display_vehicles = all_vehicles

    context = {
        'all_vehicles': display_vehicles,
        'issue_categories': issue_categories,
        'total_count': len(all_vehicles),
        'counts': {
            'battery_issue': len(issue_categories['battery_issue']),
            'battery_drained': len(issue_categories['battery_drained']),
            'battery_disconnected': len(issue_categories['battery_disconnected']),
            'sim_issue': len(issue_categories['sim_issue']),
            'internet_issue': len(issue_categories['internet_issue']),
            'telecom_issue': len(issue_categories['telecom_issue']),
        }
    }

    return render(request, "VehicleNoResponse.html", context)


def SingleVehicleCompleteDetailView(request, vehicle_code):
    template_name = "Vehicle/SingleVehicleCompleteDetails.html"

    # Get filter parameters from request
    g_status = request.GET.get('g_status')
    # start_date = request.GET.get('start_date')
    # end_date = request.GET.get('end_date')
    # export = request.GET.get('export')
    #
    # filters = {}
    # dropdown_filters = {}
    #
    # if vehicle_code:
    #     filters['vehicle_code__vehicle_code'] = vehicle_code
    #     dropdown_filters['vehicle_code__vehicle_code'] = vehicle_code
    # if start_date and end_date:
    #     filters['vendor_date_time__range'] = (start_date, end_date)
    #     dropdown_filters['vendor_date_time__range'] = (start_date, end_date)
    #
    # queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by(
    #     'vehicle_code__vehicle_code', 'vendor_date_time'
    # )
    # dropdown_queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**dropdown_filters).order_by(
    #     'vehicle_code__vehicle_code', 'vendor_date_time'
    # )
    #
    # dropdown_status = vehicle_status_duration(dropdown_queryset)
    #
    # # Initialize data structures
    # status_periods = []
    # working_periods = []
    # no_response_periods = []
    # all_vehicles = []
    #
    # # Trackers for current periods
    # current_status = None
    # current_status_start = None
    # current_working_start = None
    # current_no_response_start = None
    #
    # # Totals
    # total_filtered_duration = timedelta()
    # total_working_duration = timedelta()
    # total_no_response_duration = timedelta()
    #
    # previous_record = None
    # consecutive_records = []
    #
    # for record in queryset:
    #     v_code = record.vehicle_code.vehicle_code
    #     status = record.g_status
    #     device_status = record.device_status
    #     current_time = record.vendor_date_time
    #     lat = record.latitude
    #     lon = record.longitude
    #     pitb_code = record.vehicle_code.pitb_code
    #     reg_no = record.vehicle_code.register_no
    #     vendor_date_time = record.vendor_date_time
    #     speed = record.speed
    #     mileage = record.mileage
    #
    #     # Add record to consecutive records list
    #     consecutive_records.append({
    #         'pitb_code': pitb_code,
    #         'reg_no': reg_no,
    #         'vendor_date_time': vendor_date_time,
    #         'speed': speed,
    #         'mileage': mileage,
    #         'g_status': status,
    #         'device_status': device_status
    #     })
    #
    #     # Check if status changed
    #     if status != current_status:
    #         if current_status is not None and current_status_start is not None:
    #             # Save the previous status period
    #             duration = current_time - current_status_start
    #             total_hours = duration.total_seconds() / 3600
    #             if not g_status or current_status.lower() == g_status.lower():
    #                 total_filtered_duration += duration
    #                 status_periods.append({
    #                     'status': current_status,
    #                     'time_in': current_status_start.strftime("%Y-%m-%d - %H:%M"),
    #                     'time_out': current_time.strftime("%Y-%m-%d - %H:%M"),
    #                     'gap': f"{total_hours:.2f}h",
    #                     'vehicle_code': v_code,
    #                 })
    #
    #                 # Add consecutive records data to all_vehicles
    #                 if consecutive_records:
    #                     last_record = consecutive_records[-1]
    #                     total_speed = sum(r['speed'] for r in consecutive_records if r['speed'] is not None)
    #                     total_mileage = sum(r['mileage'] for r in consecutive_records if r['mileage'] is not None)
    #
    #                     # Calculate working hours for this period
    #                     working_hours = timedelta()
    #                     for r in consecutive_records:
    #                         engine_status = r['device_status'].strip().split(',')[0] if r['device_status'] else None
    #                         if engine_status == "ACC On":
    #                             working_hours += timedelta(minutes=1)  # Assuming 1-minute intervals
    #
    #                     all_vehicles.append({
    #                         'pitb_code': last_record['pitb_code'],
    #                         'reg_no': last_record['reg_no'],
    #                         'vendor_date_time': last_record['vendor_date_time'],
    #                         'total_speed': total_speed,
    #                         'total_mileage': total_mileage,
    #                         'working_hour': f"{working_hours.total_seconds() / 3600:.2f}h",
    #                         'g_status': current_status
    #                     })
    #                     consecutive_records = []  # Reset for next status
    #
    #         # Start new status period
    #         current_status = status
    #         current_status_start = current_time
    #
    #     # Check working status (ACC On/Off)
    #     engine_status = device_status.strip().split(',')[0] if device_status else None
    #     is_working = engine_status == "ACC On"
    #
    #     if is_working and current_working_start is None:
    #         current_working_start = current_time
    #     elif not is_working and current_working_start is not None:
    #         duration = current_time - current_working_start
    #         total_hours = duration.total_seconds() / 3600
    #         total_working_duration += duration
    #         working_periods.append({
    #             'status': 'working',
    #             'time_in': current_working_start.strftime("%Y-%m-%d - %H:%M"),
    #             'time_out': current_time.strftime("%Y-%m-%d - %H:%M"),
    #             'gap': f"{total_hours:.2f}h",
    #             'vehicle_code': v_code,
    #         })
    #         current_working_start = None
    #
    #     # Check no response status
    #     has_location = lat is not None and lon is not None and lat != '' and lon != ''
    #     is_no_response = not has_location
    #
    #     if is_no_response and current_no_response_start is None:
    #         current_no_response_start = current_time
    #     elif not is_no_response and current_no_response_start is not None:
    #         duration = current_time - current_no_response_start
    #         total_hours = duration.total_seconds() / 3600
    #         total_no_response_duration += duration
    #         no_response_periods.append({
    #             'status': 'no_response',
    #             'time_in': current_no_response_start.strftime("%Y-%m-%d - %H:%M"),
    #             'time_out': current_time.strftime("%Y-%m-%d - %H:%M"),
    #             'gap': f"{total_hours:.2f}h",
    #             'vehicle_code': v_code,
    #         })
    #         current_no_response_start = None
    #
    #     previous_record = record
    #
    # # Handle any ongoing periods at the end of the data
    # if current_status is not None and current_status_start is not None and previous_record:
    #     duration = previous_record.vendor_date_time - current_status_start
    #     total_hours = duration.total_seconds() / 3600
    #     if not g_status or current_status.lower() == g_status.lower():
    #         total_filtered_duration += duration
    #         status_periods.append({
    #             'status': current_status,
    #             'time_in': current_status_start.strftime("%Y-%m-%d - %H:%M"),
    #             'time_out': previous_record.vendor_date_time.strftime("%Y-%m-%d - %H:%M"),
    #             'gap': f"{total_hours:.2f}h",
    #             'vehicle_code': v_code,
    #         })
    #
    #     # Add final consecutive records data
    #     if consecutive_records:
    #         last_record = consecutive_records[-1]
    #         total_speed = sum(r['speed'] for r in consecutive_records if r['speed'] is not None)
    #         total_mileage = sum(r['mileage'] for r in consecutive_records if r['mileage'] is not None)
    #
    #         # Calculate working hours for this period
    #         working_hours = timedelta()
    #         for r in consecutive_records:
    #             engine_status = r['device_status'].strip().split(',')[0] if r['device_status'] else None
    #             if engine_status == "ACC On":
    #                 working_hours += timedelta(minutes=1)  # Assuming 1-minute intervals
    #
    #         all_vehicles.append({
    #             'pitb_code': last_record['pitb_code'],
    #             'reg_no': last_record['reg_no'],
    #             'vendor_date_time': last_record['vendor_date_time'],
    #             'total_speed': total_speed,
    #             'total_mileage': total_mileage,
    #             'working_hour': f"{working_hours.total_seconds() / 3600:.2f}h",
    #             'g_status': current_status
    #         })
    #
    # if current_working_start is not None and previous_record:
    #     duration = previous_record.vendor_date_time - current_working_start
    #     total_hours = duration.total_seconds() / 3600
    #     total_working_duration += duration
    #     working_periods.append({
    #         'status': 'working',
    #         'time_in': current_working_start.strftime("%Y-%m-%d - %H:%M"),
    #         'time_out': previous_record.vendor_date_time.strftime("%Y-%m-%d - %H:%M"),
    #         'gap': f"{total_hours:.2f}h",
    #         'vehicle_code': v_code,
    #     })
    #
    # if current_no_response_start is not None and previous_record:
    #     duration = previous_record.vendor_date_time - current_no_response_start
    #     total_hours = duration.total_seconds() / 3600
    #     total_no_response_duration += duration
    #     no_response_periods.append({
    #         'status': 'no_response',
    #         'time_in': current_no_response_start.strftime("%Y-%m-%d - %H:%M"),
    #         'time_out': current_time.strftime("%Y-%m-%d - %H:%M"),
    #         'gap': f"{total_hours:.2f}h",
    #         'vehicle_code': v_code,
    #     })
    #
    # # Handle Excel export
    # if export == 'excel_detail':
    #     # Create Excel writer
    #     output = BytesIO()
    #     with pd.ExcelWriter(output, engine='openpyxl') as writer:
    #         # First table data
    #         summary_data = []
    #         for row in dropdown_status:
    #             summary_data.append({
    #                 'Vehicle Code': row.get('vehicle_code', ''),
    #                 'PITB Code': row.get('pitb_code', 'N/A'),
    #                 'Reg. No': row.get('reg_no', 'N/A'),
    #                 'Working Hour': row.get('working', '--'),
    #                 'Moving': row.get('moving', '--'),
    #                 'Parked': row.get('parked', '--'),
    #                 'Idle': row.get('idle', '--'),
    #                 'Offline': row.get('offline', '--'),
    #                 'No Response': row.get('no_response', '--')
    #             })
    #
    #         # Second table data
    #         detail_data = []
    #         for data in all_vehicles:
    #             detail_data.append({
    #                 'Pitb Code': data['pitb_code'],
    #                 'Reg. No': data['reg_no'],
    #                 'Status': data['g_status'],
    #                 'Time': data['vendor_date_time'].strftime('%Y-%m-%d %H:%M:%S'),
    #                 'Speed': data['total_speed'],
    #                 'Mileage': data['total_mileage'],
    #                 'Working Hours': data['working_hour']
    #             })
    #
    #         # Create DataFrames
    #         df_summary = pd.DataFrame(summary_data)
    #         df_detail = pd.DataFrame(detail_data)
    #
    #         # Write to Excel
    #         df_summary.to_excel(writer, sheet_name='Vehicle Report', index=False, startrow=0)
    #         df_detail.to_excel(writer, sheet_name='Vehicle Report', index=False, startrow=len(df_summary) + 3)
    #
    #         # Auto-adjust column widths
    #         worksheet = writer.sheets['Vehicle Report']
    #         for idx, col in enumerate(df_summary.columns):
    #             max_length = max(
    #                 df_summary[col].astype(str).apply(len).max(),
    #                 len(col)
    #             )
    #             worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
    #
    #     # Prepare response
    #     output.seek(0)
    #     response = HttpResponse(
    #         output.getvalue(),
    #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #     )
    #     response[
    #         'Content-Disposition'] = f'attachment; filename="vehicle_report_{vehicle_code}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    #     return response
    #
    # g_statuses = VehicleLiveMonitor.objects.values_list('g_status', flat=True).distinct()
    #
    # context = {
    #     'request': request,
    #     'dropdown_status': dropdown_status,
    #     'status_duration': status_periods,
    #     'working_duration': working_periods,
    #     'no_response_duration': no_response_periods,
    #     'g_statuses': g_statuses,
    #     'selected_g_status': g_status,
    #     'start_date': start_date,
    #     'end_date': end_date,
    #     'vehicle_code': vehicle_code,
    #     'total_duration': f"{total_filtered_duration.total_seconds() / 3600:.2f}h",
    #     'total_working_duration': f"{total_working_duration.total_seconds() / 3600:.2f}h",
    #     'total_no_response_duration': f"{total_no_response_duration.total_seconds() / 3600:.2f}h",
    #     'all_vehicles': all_vehicles,
    # }
    return render(request)


def process_chart_data(vehicles, current_date, current_time):
    g_status_count = Counter()
    no_responsed_vehicles = defaultdict(list)

    current_count = 0
    no_response_count = 0

    for v in vehicles:
        vdt = parse_vendor_date(v.vendor_date_time)
        if not vdt:
            continue

        if current_date == vdt.date():
            g_status_count[v.g_status] += 1
            current_count += 1
        else:
            delta_days = (current_date - vdt.date()).days
            if delta_days > 0:
                key = f"{delta_days} Day{'s' if delta_days != 1 else ''}"
                no_responsed_vehicles[key].append((v, delta_days))
                no_response_count += 1

    return {
        "current_working": dict(g_status_count),
        "no_responsed": {k: len(v) for k, v in no_responsed_vehicles.items()},
        "current_count": current_count,
        "no_response_count": no_response_count,
        "total_count": current_count + no_response_count
    }


### FUNCTION TO PROCESS VEHICLE MANAGEMENT TABLE IN TEMPLATE
def ProcessTableData_VehicleManagement_Function(vehicles, current_date, current_time):
    all_vehicles_today = []

    for v in vehicles:
        vdt = parse_vendor_date(v.vendor_date_time)
        if not vdt:
            continue

        delay_time = None
        if vdt:
            time_diff = current_time - vdt
            delay_time = time_diff.total_seconds() / 3600

        is_current = vdt.date() == current_date
        nr_days = (current_date - vdt.date()).days if vdt.date() < current_date else 0

        device_status = "Working"
        if nr_days > 0:
            device_status = "No Response"

        all_vehicles_today.append({
            "vehicle": v,
            "vendor_date_time": vdt,
            "nr_days": nr_days,
            "is_current": is_current,
            "delay_time": delay_time,
            "device_status": device_status
        })

    current_vehicles = sorted(
        all_vehicles_today,
        key=lambda x: (not x['is_current'], x['nr_days'])
    )

    return {
        'current_vehicles': current_vehicles,
        'all_vehicles_today_count': sum(1 for v in all_vehicles_today if v['is_current'])
    }


def parse_vendor_date(vendor_date_time):
    if not vendor_date_time:
        return None

    try:
        if isinstance(vendor_date_time, str):
            for fmt in ["%Y-%m-%d %H:%M:%S", "%B %d, %Y at %H:%M", "%Y-%m-%d %H:%M"]:
                try:
                    vdt = datetime.datetime.strptime(vendor_date_time, fmt)
                    break
                except ValueError:
                    continue
            else:
                vdt = parse_datetime(vendor_date_time)
        else:
            vdt = vendor_date_time

        if vdt and is_naive(vdt):
            vdt = make_aware(vdt, get_current_timezone())

        return vdt.replace(second=0, microsecond=0) if vdt else None

    except Exception as e:
        print(f"Error parsing vendor_date_time: {e}")
        return None


# def VehicleRouteReportMohidView(request):
#     message = ""
#     cursor = connections['default'].cursor()
#     template_name = "VehicleRouteReport11.html"
#
#     # Set default values
#     current_data_time = datetime.datetime.now()
#     current_date = current_data_time.date()
#     format_str_date = "%Y-%m-%d"
#     work_date_str = current_date.strftime(format_str_date)
#
#     # Default values for filters
#     filtered_vehicle_type = None
#     filtered_vehicle_id = None
#     from_datetime = None
#     to_datetime = None
#
#     # Process filters if submitted
#     if request.method == "POST":
#         filtered_vehicle_type = request.POST.get('cmd_vehicle_type')
#         filtered_vehicle_id = request.POST.get('cmd_vehicle_list')
#         from_datetime = request.POST.get('from_datetime')
#         to_datetime = request.POST.get('to_datetime')
#
#         # Convert to proper formats for filtering
#         if from_datetime:
#             from_datetime = parse_datetime(from_datetime)
#         if to_datetime:
#             to_datetime = parse_datetime(to_datetime)
#
#         # Use provided date instead of current date if filters are set
#         if from_datetime:
#             filter_date = from_datetime.date()
#         else:
#             filter_date = current_date
#     else:
#         filter_date = current_date
#
#     #### GENERATE GRID FOR VEHICLE DATA TABLE
#     # Base query
#     vehicle_query = VehicleData.objects.all().order_by('vehicle_type')
#
#     # Apply type filter if provided
#     if filtered_vehicle_type and filtered_vehicle_type != "NA":
#         vehicle_query = vehicle_query.filter(vehicle_type=filtered_vehicle_type)
#
#     # Apply vehicle ID filter if provided
#     if filtered_vehicle_id and filtered_vehicle_id != "NA":
#         vehicle_query = vehicle_query.filter(vehicle_code=filtered_vehicle_id)
#
#     # Get the filtered vehicle list
#     vehicle_type_list = vehicle_query
#
#     # Process each vehicle to calculate distance
#     vehicle_data_obj = []
#     for data in vehicle_type_list:
#         vehicle_code = data.vehicle_code
#
#         # Query points, applying date filter based on form inputs
#         points_query = TrackerRawData.objects.filter(vehicle_code_id=vehicle_code)
#
#         # Apply date filters if provided
#         if from_datetime and to_datetime:
#             points_query = points_query.filter(
#                 vendor_date_time__gte=from_datetime,
#                 vendor_date_time__lte=to_datetime
#             )
#         else:
#             # Default to current date if no date filters
#             points_query = points_query.filter(vendor_date_time__date=filter_date)
#
#         points = points_query.order_by('vendor_date_time').values('geom')
#
#         # Calculate distance
#         total_distance = 0.0
#         previous_point = None
#
#         for row in points:
#             point = GEOSGeometry(row['geom'])
#             if previous_point:
#                 dist = geodesic(
#                     (previous_point.y, previous_point.x),
#                     (point.y, point.x)
#                 ).meters
#                 total_distance += dist
#             previous_point = point
#
#         # Create vehicle data dictionary
#         vehicle_dict = dict()
#         vehicle_dict["vehicle_code"] = data.vehicle_code
#         vehicle_dict["register_no"] = data.register_no
#         vehicle_dict["vehicle_type"] = data.vehicle_type
#         vehicle_dict["line_length_m"] = total_distance
#         vehicle_dict["line_length_km"] = f"{(float(total_distance) / 1000):.2f}"
#
#         vehicle_data_obj.append(vehicle_dict)
#
#     # Sort by distance
#     vehicle_data_obj_sort = sorted(vehicle_data_obj, key=lambda x: float(x['line_length_km']), reverse=True)
#
#     # Calculate statistics for the filtered data
#     vehicle_distance_greater_than_20km = [vehicle for vehicle in vehicle_data_obj
#                                           if float(vehicle['line_length_km']) >= 20.00]
#     vehicle_distance_greater_than_20km_count = len(vehicle_distance_greater_than_20km)
#
#     vehicle_distance_less_than_20km = [vehicle for vehicle in vehicle_data_obj
#                                        if float(vehicle['line_length_km']) < 20.00]
#     vehicle_distance_less_than_20km_count = len(vehicle_distance_less_than_20km)
#
#     vehicle_distance_less_than_5km = [vehicle for vehicle in vehicle_data_obj
#                                       if float(vehicle['line_length_km']) <= 5.00]
#     vehicle_distance_less_than_5km_count = len(vehicle_distance_less_than_5km)
#
#     # Get vehicle IDs and count for dropdowns
#     all_vehicle_types = VehicleData.objects.values('vehicle_type').annotate(count=Count('id')).order_by('vehicle_type')
#     all_vehicle_ids = VehicleData.objects.values_list('vehicle_code', flat=True)
#     vehicle_count = all_vehicle_ids.count()
#
#     # Get other necessary data for the template
#     query_asset_type = "SELECT 'Container' as name, count(*) AS value FROM tbl_container_data UNION ALL SELECT 'Drum' as name, count(*) AS value FROM tbl_drum_data UNION ALL SELECT 'Collection-Site' as name, count(*) AS value FROM tbl_collection_site UNION ALL SELECT 'Dumping-Site' as name, count(*) AS value FROM tbl_dumping_site UNION ALL SELECT 'Landuse' as name, count(*) AS value FROM tbl_landuse_boundary UNION ALL SELECT 'Admin' as name, count(*) AS value FROM tbl_administrative_boundary; "
#     cursor.execute(query_asset_type)
#     asset_type_summary = DictinctFetchAll(cursor)
#
#     query_feature = "SELECT ST_X(geom) as x, ST_Y(geom) as y, chasis_no, register_no, vehicle_type, g_status, direction, speed, COALESCE(device_status, 'NA') AS device_status, COALESCE(ignition_status, 'NA') AS ignition_status, COALESCE(geo_location, 'NA') AS geo_location, vendor_date_time, COALESCE(duration, 'NA') AS duration FROM tbl_vehicle_live_monitor AS vlm LEFT OUTER JOIN tbl_vehicle_data AS veh ON vlm.vehicle_code_id = veh.vehicle_code; "
#     cursor.execute(query_feature)
#     vehicle_live_lists = DictinctFetchAll(cursor)
#
#     context = {
#         'vehicle_type_list': all_vehicle_types,
#         'vehicle_data_obj': vehicle_data_obj_sort,
#         'asset_type_summary': asset_type_summary,
#         'vehicle_live_lists': vehicle_live_lists,
#         'vehicle_ids': all_vehicle_ids,
#         'vehicle_count': vehicle_count,
#         'vehicle_distance_less_than_20km_count': vehicle_distance_less_than_20km_count,
#         'vehicle_distance_greater_than_20km_count': vehicle_distance_greater_than_20km_count,
#         'vehicle_distance_less_than_5km_count': vehicle_distance_less_than_5km_count,
#         # Add filter values to context for re-populating the form
#         'filtered_vehicle_type': filtered_vehicle_type,
#         'filtered_vehicle_id': filtered_vehicle_id,
#         'from_datetime': from_datetime.isoformat() if from_datetime else None,
#         'to_datetime': to_datetime.isoformat() if to_datetime else None,
#     }
#     return render(request, template_name, context)


@require_GET
def get_vehicle_vtms_status(request):
    vehicle_code = request.GET.get('vehicle_code')
    try:
        vehicle = VehicleData.objects.get(vehicle_code=vehicle_code)
        return JsonResponse({'success': True, 'vtms_status': vehicle.vtms_status})
    except VehicleData.DoesNotExist:
        return JsonResponse({'success': False, 'vtms_status': ''})


@require_POST
def update_vehicle_vtms_status(request):
    vehicle_code = request.POST.get('vehicle_code')
    new_status = request.POST.get('vtms_status')
    allowed = [c[0] for c in VehicleData.VTMS_STATUS_CHOICES]
    if new_status not in allowed:
        return JsonResponse({'success': False, 'message': 'Invalid status'}, status=400)
    try:
        vehicle = VehicleData.objects.get(vehicle_code=vehicle_code)
        vehicle.vtms_status = new_status
        vehicle.updated_at = timezone.now()
        vehicle.updated_by = request.user.username if request.user.is_authenticated else 'system'
        vehicle.save(update_fields=['vtms_status', 'updated_at', 'updated_by'])
        return JsonResponse({'success': True, 'vtms_status': vehicle.vtms_status})
    except VehicleData.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Vehicle not found'}, status=404)


def SyncVehicleTrackerGPRS_RawDataView(request):
    message = ""
    template_name = "Tracker/VehicleTrackerGPRSRaw.html"
    dateTime = datetime.datetime.now()
    today_date = dateTime.strftime('%Y-%m-%d')
    yesterday_date = (dateTime - timedelta(days=1)).strftime('%Y-%m-%d')

    # Initialize filter variables with defaults
    get_vehicle_type = "NA"
    get_vehicle_list = ""
    get_gprs_status = ""
    get_start_date = yesterday_date
    get_end_date = yesterday_date

    # Process filter parameters from request
    if request.method == "POST":
        params = request.POST

        get_vehicle_type = params.get('cmd_vehicle_type', "NA")
        get_vehicle_list = params.get('cmd_vehicle_list', "")
        get_gprs_status = params.get('cmd_gprs_status', "")
        get_start_date = params.get('start_date', yesterday_date)
        get_end_date = params.get('end_date', yesterday_date)

    # Parse dates once for all operations
    try:
        start_date = datetime.datetime.strptime(get_start_date, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(get_end_date, "%Y-%m-%d").date()
    except ValueError:
        # Handle invalid date format
        start_date = datetime.datetime.strptime(yesterday_date, "%Y-%m-%d").date()
        end_date = start_date

    # Here Applying Sir logic
    ### STEP-1 GENERATE WORKING SCHEDULA WITH VEHICLE SCHEDULE IN SYSTEM FUNCTION
    single_date = start_date
    while single_date <= end_date:
        logger.info(f"Processing {single_date}")
        GenerateWorkingWithVehicleSchedule_Function(single_date)
        logger.info(f"Finished processing {single_date}")
        single_date += timedelta(days=1)

    # Always filter by date first to reduce initial result set
    base_query = VehicleScheduleGPRSApi.objects.select_related('vehicle_code').filter(
        veh_sch_date__gte=start_date,
        veh_sch_date__lte=end_date
    )

    # Count summary queries - use the filtered base to improve performance
    total_vehicle_count = VehicleScheduleGPRSApi.objects.count()  # Keep the total count of all vehicles

    # Only filter the pending/completed counts using the date filter
    date_filtered_query = VehicleScheduleGPRSApi.objects.filter(
        veh_sch_date__gte=start_date,
        veh_sch_date__lte=end_date
    )
    pending_vehicle_count = date_filtered_query.filter(process_status='Pending').count()
    completed_vehicle_count = date_filtered_query.filter(process_status='Completed').count()

    # Apply additional filters to the base query
    if get_vehicle_type != "NA":
        base_query = base_query.filter(vehicle_code__vehicle_type=get_vehicle_type)

    if get_vehicle_list:
        base_query = base_query.filter(vehicle_code__pitb_code=get_vehicle_list)

    if get_gprs_status:
        base_query = base_query.filter(process_status=get_gprs_status)

    # Count by Vehicle Type - only for the current filtered set
    count_by_vehicle_type = date_filtered_query.values('vehicle_code__vehicle_type').annotate(count=Count('id'))

    # Apply annotations and get final data
    data = base_query.annotate(
        vehicle_type=F('vehicle_code__vehicle_type'),
        pitb_code=F('vehicle_code__pitb_code'),
        difference=ExpressionWrapper(F('retrieve_record') - F('vendor_record'), output_field=IntegerField()),
    ).order_by('-veh_sch_date', 'vehicle_code__vehicle_type')

    # Filter dropdown data efficiently
    if get_vehicle_type != "NA":
        vehicle_pitb_code = list(
            VehicleData.objects.filter(vehicle_type=get_vehicle_type)
            .values('pitb_code').distinct()
            .exclude(pitb_code__isnull=True)
        )
    else:
        vehicle_pitb_code = list(
            VehicleData.objects.values('pitb_code').distinct()
            .exclude(pitb_code__isnull=True)
        )

    response_status = list(VehicleScheduleGPRSApi.objects.values('process_status').distinct())
    vehicle_type_list = VehicleData.objects.values('vehicle_type').annotate(count=Count('id')).order_by('vehicle_type')
    vehicle_code = VehicleData.objects.values('vehicle_code')

    # Sync Function logic
    # Add this to handle the AJAX request for sync function
    if request.method == "POST" and 'sync_function' in request.POST:
        get_vehicle_id = request.POST.get('vehicle_id')
        get_selected_date = request.POST.get('selected_date')
        # get_time_from = request.POST.get('time_from', '00:00:00')
        get_time_from = "00:00:01"
        # get_time_to = request.POST.get('time_to', '23:59')

        # Get current date and time
        current_date = dateTime.date()
        # Convert selected date to date object for comparison
        selected_date = datetime.datetime.strptime(get_selected_date, '%Y-%m-%d').date()

        # If selected date is current date, use current time
        # Otherwise use end of day (23:59:59)
        if selected_date == current_date:
            get_time_to = dateTime.strftime('%H:%M:%S')
        else:
            get_time_to = request.POST.get('time_to', '23:59:59')  # Changed default to include seconds
        try:
            # Run the sync function with parameters from the table row
            result = SyncTrackerGPRSVehicleData_By_Vendor_Function(
                request,
                get_vehicle_id,
                get_selected_date,
                get_time_from,
                get_time_to
            )

            # Check result and return appropriate response
            if result == "RECORD FOUND":
                return JsonResponse({
                    'status': 'success',
                    'message': 'Data retrieved and processed successfully.'
                })
            else:
                return JsonResponse({
                    'status': 'warning',
                    'message': 'No new records found or processing incomplete.'
                })
        except Exception as e:
            logger.error(f"Error in sync function: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f'Error: {str(e)}'
            })
    params = {
        'vehicle_type_list': vehicle_type_list,
        'page_title': "Vehicle Management",
        'message': message,
        'data': data,
        'vehicle_regis': vehicle_code,
        'total_vehicle_count': total_vehicle_count,
        'pending_vehicle_count': pending_vehicle_count,
        'completed_vehicle_count': completed_vehicle_count,
        'count_by_vehicle_type': count_by_vehicle_type,
        'vehicle_pitb_code': vehicle_pitb_code,
        'response_status': response_status,
        'yesterday_date': yesterday_date,
        # Pass the filter values back to the template
        'selected_vehicle_type': get_vehicle_type,
        'selected_vehicle_list': get_vehicle_list,
        'selected_gprs_status': get_gprs_status,
        'start_date': get_start_date,
        'end_date': get_end_date,
        'today_date': today_date
    }

    return render(request, template_name, params)


def FetchVehicleFeatureDataView(request):
    cursor = connections['default'].cursor()

    query_feature = "SELECT ST_X(geom) as x, ST_Y(geom) as y, register_no FROM tbl_vehicle_live_monitor AS vlm LEFT OUTER JOIN tbl_vehicle_data AS veh ON vlm.vehicle_code_id = veh.vehicle_code; ";
    cursor.execute(query_feature)
    feature_lists = DictinctFetchAll(cursor)

    message = "Success"
    params = {
        'message': message,
        'feature_lists': feature_lists,
    }

    return HttpResponse(json.dumps(params, default=date_handler))


def FetchVehicleTripHistoryDataView(request):
    cursor = connections['default'].cursor()

    format_str = "%Y-%m-%d %H:%M:%S"
    current_data_time = datetime.datetime.now()
    datetime_to_inst = current_data_time

    # GetTrackerSchedule = TrackerSchedule.objects.all().order_by('-created_at')
    # GetTrackerSchedule_Length = len(GetTrackerSchedule)
    # if GetTrackerSchedule_Length == 0:
    #
    #     # Get the start of the current day (midnight)
    #     datetime_from_inst = datetime.combine(current_data_time.date(), datetime.time.min)
    #
    # else:
    #     ### FETCH TRACKER SCHEDULE FROM AND TO DATE
    #     latest_tracker_schedule = TrackerSchedule.objects.latest('id')
    #     datetime_from_5_hour = latest_tracker_schedule.check_out
    #
    #     datetime_from_5_hour_str = (datetime_from_5_hour + datetime.timedelta(hours=5)).strftime(format_str)
    #     datetime_from_inst = datetime.strptime(datetime_from_5_hour_str, format_str)
    #
    # ### INSERT RECORD IN TRACKER SCHEDULE
    # time_difference = datetime_to_inst - datetime_from_inst
    # total_hours = time_difference.total_seconds() / 3600
    #
    # InstTrackerSchedule = TrackerSchedule(
    #     check_in=datetime_from_inst,
    #     check_out=datetime_to_inst,
    #     time_duration=total_hours,
    #     process_status="Pending",
    #     created_at=current_data_time,
    #     created_by="admin"
    # )
    # InstTrackerSchedule.save()
    # InstTrackerSchedule_id = InstTrackerSchedule.id
    # ### INSERT RECORD IN TRACKER SCHEDULE
    #
    # ### FETCH TRACKER SCHEDULE FROM AND TO DATE
    # datetime_from = (datetime_from_inst + datetime.timedelta(hours=0)).strftime(format_str)
    # datetime_to = (datetime_to_inst + datetime.timedelta(hours=0)).strftime(format_str)

    # query_track_schedule = "SELECT TO_CHAR(check_in, 'YYYY-MM-DD HH24:MI:SS') AS check_in, TO_CHAR(check_out, 'YYYY-MM-DD HH24:MI:SS') AS check_out FROM tbl_tracker_schedule WHERE id = '" + str(
    #     InstTrackerSchedule_id) + "' "

    # ### FETCH AND SAVE VEHICLE TRACKER GPRS DATA
    # VehicleTrackerGPRS_Status = VehicleTrackerGPRSRecord_Function("", "", "")
    # VehicleTrackerGPRS_Status = VehicleTrackerGPRSRecord_Function("", datetime_from, datetime_to)
    #
    # if VehicleTrackerGPRS_Status:
    #     ### ALL GPRS POINT SAVE IN DATABASE THEN CLOSE TRACKER SCHEDULE PROCESS
    #     UpdateTrackerSchedule = TrackerSchedule.objects.get(id=InstTrackerSchedule_id)
    #     UpdateTrackerSchedule.process_status = "Completed"
    #     UpdateTrackerSchedule.save()
    # ### FETCH AND SAVE VEHICLE TRACKER GPRS DATA

    ### GET GEO-FENCE BASED CONTAINER GPRS LOCATION
    ContainerGeoFence_TrackerRawData_Function()
    ### GET GEO-FENCE BASED CONTAINER GPRS LOCATION

    query_feature = "SELECT ST_X(geom) as x, ST_Y(geom) as y, register_no FROM tbl_vehicle_live_monitor AS vlm LEFT OUTER JOIN tbl_vehicle_data AS veh ON vlm.vehicle_code_id = veh.vehicle_code; ";
    cursor.execute(query_feature)
    feature_lists = DictinctFetchAll(cursor)

    message = "Success"
    params = {
        'message': message,
        'feature_lists': feature_lists,
    }

    return HttpResponse(json.dumps(params, default=date_handler))


### AJAX
def FetchVehicleLiveMonitoringByGPRSView(request):
    cursor = connections['default'].cursor()

    get_vehicle_status = request.POST['vehicle_status']
    get_vehicle_type = request.POST['vehicle_type']
    get_vehicle_code = request.POST['vehicle_code']

    dateTime = datetime.datetime.now()
    selected_date = dateTime.date()

    ### STEP-1 GENERATE WORKING SCHEDULA WITH VEHICLE SCHEDULE IN SYSTEM FUNCTION
    GenerateWorkingWithVehicleSchedule_Function(selected_date)

    ### STEP-2 GET RESPONSE FROM VEHICLE API FUNCTION
    response_data = ResponseVehicleApi_By_Vendor_Function()
    if len(response_data) > 0:  ### RECORD EXIST ###

        vehicle_response_api = response_data['MainData']
        for i in range(len(vehicle_response_api)):
            # CREATE AND UPDATE VEHICLE (START)
            ch_system_vehicle_id = vehicle_response_api[i]['VehicleId']

            VehicleStatus = VehicleData.objects.filter(vehicle_code=ch_system_vehicle_id)
            if len(VehicleStatus) > 0:  # UPDATE

                ck_vendor_date_time = str(vehicle_response_api[i]['GPSTime'])
                ck_system_date_time = vehicle_response_api[i]['RecTime']
                try:
                    format_str = "%Y-%m-%d %H:%M:%S"
                    if '.' in ck_vendor_date_time:
                        ck_vendor_date_time = ck_vendor_date_time.split('.')[0]  # Remove the decimal part

                    _ck_vendor_date_time = datetime.datetime.fromisoformat(ck_vendor_date_time)
                    formatted = _ck_vendor_date_time.strftime(format_str)  # Format as a string
                    # print("Valid ISO format:")
                except ValueError:
                    print(f"_ck_vendor_date_time - {ck_vendor_date_time}")
                    # formatted = ck_vendor_date_time.strftime(format_str)

                ### IF VEHICLE API DATE IS CURRENT DATE
                if _ck_vendor_date_time.date() == selected_date:

                    vt_latitude = vehicle_response_api[i]['Lat']
                    vt_longitude = vehicle_response_api[i]['Long']
                    get_feature_coordinate = "POINT(" + str(vt_longitude) + " " + str(vt_latitude) + ")"

                    ### GENERATE WORKING SCHEDULA
                    qs = VehicleLiveMonitor.objects.filter(vehicle_code_id=ch_system_vehicle_id)
                    if not qs.exists():
                        auto_gprs_live_code = AutoGenerateCodeForModel(VehicleLiveMonitor, "veh_live_mont_code",
                                                                       "GPSL-")
                        VehicleLiveMonitor.objects.create(
                            veh_live_mont_code=auto_gprs_live_code,
                            vehicle_code_id=ch_system_vehicle_id,
                            geom=get_feature_coordinate,
                            latitude=vt_latitude,
                            longitude=vt_longitude,
                            g_status=vehicle_response_api[i]['GStatus'],
                            speed=vehicle_response_api[i]['Speed'],
                            device_status=vehicle_response_api[i]['dev_status'],
                            direction=vehicle_response_api[i]['Direction'],
                            ignition_status=vehicle_response_api[i]['IgnStatus'],
                            geo_location=vehicle_response_api[i]['Location'],
                            vendor_date_time=formatted,
                            duration=vehicle_response_api[i]['Duration'],
                            created_at=dateTime,
                            created_by="admin"
                        )
                    else:
                        ### UPDATE LIVE MONITORING LOCATION START
                        VehicleLiveMonitor.objects.filter(vehicle_code_id=ch_system_vehicle_id).update(
                            geom=get_feature_coordinate,
                            latitude=vt_latitude,
                            longitude=vt_longitude,
                            g_status=vehicle_response_api[i]['GStatus'],
                            speed=vehicle_response_api[i]['Speed'],
                            device_status=vehicle_response_api[i]['dev_status'],
                            direction=vehicle_response_api[i]['Direction'],
                            ignition_status=vehicle_response_api[i]['IgnStatus'],
                            geo_location=vehicle_response_api[i]['Location'],
                            vendor_date_time=formatted,
                            duration=vehicle_response_api[i]['Duration'],
                            updated_at=dateTime,
                            updated_by="admin"
                        )
                        ### UPDATE LIVE MONITORING LOCATION START

            ### IF END (UPDATE)
        ### LOOP END
    ### STEP-2

    ### GREATER THEN 0
    query_feature = "SELECT ST_X(geom) as x, ST_Y(geom) as y, chasis_no, register_no, COALESCE(pitb_code, 'Waiting') AS pitb_code, vehicle_type, g_status, direction, speed FROM tbl_vehicle_live_monitor AS vlm LEFT OUTER JOIN tbl_vehicle_data AS veh ON vlm.vehicle_code_id = veh.vehicle_code WHERE 1=1 "

    if get_vehicle_status != 'NA':
        query_feature += f" AND g_status = '{get_vehicle_status}'"

    if get_vehicle_type != "NA":
        query_feature += f" AND veh.vehicle_type = '{get_vehicle_type}' "

    if get_vehicle_code != "NA":
        query_feature += f" AND (veh.vehicle_code = '{get_vehicle_code}') "

    query_feature += ";"
    cursor.execute(query_feature)
    feature_lists = DictinctFetchAll(cursor)

    vehicle_g_status = list(VehicleLiveMonitor.objects.values('g_status').annotate(count=Count('id')))

    message = "Success"
    params = {
        'message': message,
        'feature_lists': feature_lists,
        'vehicle_g_status': vehicle_g_status,
    }

    return HttpResponse(json.dumps(params, default=date_handler))


def FetchSingleVehicleRouteDataView(request):
    get_vehicle_id = request.POST['vehicle_id']
    get_selected_date = request.POST['selected_date']
    get_time_from = request.POST['time_from']
    get_time_to = request.POST['time_to']

    format_str_date = "%Y-%m-%d"
    current_data_time = datetime.datetime.now()
    today_date = current_data_time.date()

    dt_selected_date = datetime.datetime.strptime(get_selected_date, format_str_date).date()

    ### STEP-1 GENERATE WORKING SCHEDULA WITH VEHICLE SCHEDULE IN SYSTEM FUNCTION
    GenerateWorkingWithVehicleSchedule_Function(dt_selected_date)

    start_of_day_time_str = get_time_from + ':00'  # '00:00:01'
    end_of_day_time_str = get_time_to + ':00'  # '23:59:00'

    if dt_selected_date != today_date:
        ### SYNC TRACKER GPRS VEHICLE DATA BY VENDOR SYSTEM
        SyncTrackerGPRSVehicleData_By_Vendor_Function(request,
                                                      get_vehicle_id,
                                                      get_selected_date,
                                                      start_of_day_time_str,
                                                      end_of_day_time_str)
    else:

        current_time_str = datetime.datetime.now().strftime('%H:%M')
        if current_time_str >= get_time_to:
            end_of_day_time_str = get_time_to + ':00'  # '00:01:00'
        else:
            end_of_day_time_str = current_time_str + ':00'  # '23:59:00'

        # ### SYNC TRACKER GPRS VEHICLE DATA BY VENDOR SYSTEM
        # SyncTrackerGPRSVehicleData_By_Vendor_Function(request,
        #                                               get_vehicle_id,
        #                                               get_selected_date,
        #                                               start_of_day_time_str,
        #                                               end_of_day_time_str)

    ### COMBINE DATE AND TIME INTO STRING
    django_datetime_from = datetime.datetime.strptime(f"{get_selected_date} {get_time_from}", "%Y-%m-%d %H:%M")
    django_datetime_to = datetime.datetime.strptime(f"{get_selected_date} {get_time_to}", "%Y-%m-%d %H:%M")
    # Check if request is from a server IP
    server_request = Retrieve_IP_Address(request)
    if server_request == "Local Development":
        django_datetime_from = datetime.datetime.strptime(f"{get_selected_date} {get_time_from}",
                                                          "%Y-%m-%d %H:%M") + timedelta(hours=5)

    tracker_raw_gprs_lists = list(
        TrackerRawData.objects.filter(
            vehicle_code_id=get_vehicle_id,
            vendor_date_time__gte=django_datetime_from,
            vendor_date_time__lte=django_datetime_to
        )
        .annotate(
            x=RawSQL("ST_X(geom)", []),
            y=RawSQL("ST_Y(geom)", [])
        )
        .values(
            "x", "y", "vehicle_code_id", "g_status", "system_date_time", "vendor_date_time",
            "device_status", "max_speed", "speed", "vehicle_status", "direction",
            "distance", "mileage"
        )
        .order_by("id")
    )

    ### VEHICLE TRACK CALCULATE LENGTH FROM GPRS DATA
    VehicleTracks = TrackerRawData.objects.filter(
        vehicle_code_id=get_vehicle_id,
        vendor_date_time__gte=django_datetime_from,
        vendor_date_time__lte=django_datetime_to
    ).values('vehicle_code_id').annotate(
        line=MakeLine('geom'),  # Create a line from grouped points
        length=Length(MakeLine('geom'))  # Calculate line length
    )

    # Convert data into JSON response
    VehicleTracks_Length = [
        {
            "vehicle_code_id": track["vehicle_code_id"],
            "length_meters": track["length"].m,  # Convert to meters
            "line_geojson": track["line"].geojson  # Convert to GeoJSON
        }
        for track in VehicleTracks
    ]

    message = "Success"
    params = {
        'message': message,
        'tracker_raw_gprs_lists': tracker_raw_gprs_lists,
        'vehicle_tracks_length': VehicleTracks_Length,
    }

    return HttpResponse(json.dumps(params, default=date_handler))


def FetchSingleVehicleTripHistoryDataView(request):
    get_vehicle_id = request.POST['vehicle_id']
    get_start_date = request.POST['start_date']
    get_end_date = request.POST['end_date']

    format_str_date = "%Y-%m-%d"
    format_str_date_time = "%Y-%m-%d %H:%M:%S"
    current_data_time = datetime.datetime.now()

    ### IF DATE NOT EXIST
    dt_start_date = datetime.strptime(get_start_date, format_str_date_time).date()
    dt_end_date = datetime.strptime(get_end_date, format_str_date_time).date()

    # ### FETCH NO OF DATE WHICH IS NOT EXIST
    # range_date_list = []
    # between_date_list = []
    # current = dt_start_date
    # while current <= dt_end_date:
    #     ### IF VEHICLE SCHEDULE DATE EXIST OR NOT
    #     vehicle_schedule_record = VehicleScheduleGPRSApi.objects.filter(veh_sch_date=current,
    #                                                                     vehicle_code_id=get_vehicle_id)
    #     if len(vehicle_schedule_record) == 0:  ## IF SELECTED DATE EXIST
    #         between_date_list.append(current)
    #
    #     range_date_list.append(current)
    #     current += timedelta(days=1)
    # # print(date_list)
    # # string_dates = [d.strftime("%Y-%m-%d") for d in date_list]

    # if len(between_date_list) > 0:
    #     for i in range(len(between_date_list)):
    #         converted_date_inst = between_date_list[i]
    #
    #         ### IF CURRENT DATE EXIST OR NOT
    #         current_date_records = WorkScheduleGPRSApi.objects.filter(work_date=converted_date_inst)
    #         if len(current_date_records) == 0:  ## IF CURRENT DATE EXIST
    #             auto_gprs_code = AutoGenerateCodeForModel(WorkScheduleGPRSApi, "code", "WS-")
    #             ### GPRS - WORKING SCHEDULE API (START)
    #             InstWorkSche = WorkScheduleGPRSApi(
    #                 code=auto_gprs_code,
    #                 work_date=converted_date_inst,
    #                 run_count=0,
    #                 process_status="Pending",
    #                 description="Sync-Current",
    #                 created_at=current_data_time,
    #                 created_by="admin"
    #             )
    #             InstWorkSche.save()
    #             ### GPRS - WORKING SCHEDULE API (END)
    #
    #         ### GENERATE VEHICLE SCHEDULA IN SYSTEM FUNCTION
    #         GenerateVehicleSchedule_Function("", converted_date_inst)
    #     ### LOOP (END)
    # ### CHECK SCHEDULE ADDED OR NOT (END)

    ### CHECK VEHICLE VENDOR RECORD EQUAL TO DUMP RECORD WITH STATUS
    fetch_vehicle_sche_record = VehicleScheduleGPRSApi.objects.get(veh_sch_date=dt_start_date,
                                                                   vehicle_code_id=get_vehicle_id)
    ck_process_status = fetch_vehicle_sche_record.process_status

    if ck_process_status == "Pending":  ### PROCESS STATUS "PENDING" ###
        ### CHECK VEHICLE GPRS LIST (START)

        start_of_day_time_str = '00:00:00'
        end_of_day_time_str = '23:59:59'
        ### DATE CONVERT INTO STRING FORMAT
        work_date_str = dt_start_date.strftime(format_str_date)

        ### COMBINE DATE AND TIME INTO STRING
        datetime_from = work_date_str + " " + start_of_day_time_str  ## "2025-02-27 00:00:00"
        datetime_to = work_date_str + " " + end_of_day_time_str  ## "2025-02-27 23:59:59"

        ### GET VENDOR SERVER RESPONSE FROM TRACKER GPRS RAW DATA API FUNCTION
        response_gprs_api = ResponseTrackerGPRSRawApi_By_Vendor_Function(get_vehicle_id, datetime_from, datetime_to)

        ### API RESPONSE CHECK ###
        if len(response_gprs_api) > 0:  ### RECORD EXIST ###
            total_vendor_record = len(response_gprs_api['Table'])

            vehicle_gprs_response_api = response_gprs_api['Table']
            vehicle_gprs_response_api_len = len(vehicle_gprs_response_api)
            for g in range(vehicle_gprs_response_api_len):
                ck_vehicle_code = vehicle_gprs_response_api[g]['VehicleID']
                tr_vendor_date_time = vehicle_gprs_response_api[g]['GpsTime']
                tr_system_date_time = vehicle_gprs_response_api[g]['RecTime']
                ### TRACKER RAW DATA (START)

                GetTrackerRawData = TrackerRawData.objects.filter(vehicle_code_id=ck_vehicle_code,
                                                                  system_date_time=tr_system_date_time,
                                                                  vendor_date_time=tr_vendor_date_time)
                if len(GetTrackerRawData) == 0:
                    auto_gprs_code = AutoGenerateCodeForModel(TrackerRawData, "gprs_raw_code", "GPRS-")

                    vt_latitude = vehicle_gprs_response_api[g]['Lat']
                    vt_longitude = vehicle_gprs_response_api[g]['Long']
                    get_feature_coordinate = "POINT(" + str(vt_longitude) + " " + str(vt_latitude) + ")"

                    InstTrackerRawData = TrackerRawData(
                        gprs_raw_code=auto_gprs_code,
                        vehicle_code_id=ck_vehicle_code,
                        terminal_no=vehicle_gprs_response_api[g]['TerminalNo'],
                        geom=get_feature_coordinate,
                        latitude=vt_latitude,
                        longitude=vt_longitude,
                        g_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        vehicle_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        device_status=vehicle_gprs_response_api[g]['Dev_status'],
                        vendor_date_time=vehicle_gprs_response_api[g]['GpsTime'],
                        system_date_time=vehicle_gprs_response_api[g]['RecTime'],
                        speed=vehicle_gprs_response_api[g]['Speed'],
                        distance=vehicle_gprs_response_api[g]['Distance'],
                        direction=vehicle_gprs_response_api[g]['Direction'],
                        mileage=vehicle_gprs_response_api[g]['Mileage_Val'],
                        created_at=current_data_time,
                        created_by="admin"
                    )
                    InstTrackerRawData.save()
                    # TRACKER RAW DATA (END)
                # RECORD FOUND (END)
            ### LOOP END

            tracker_gprs_length = TrackerRawData.objects.filter(
                vehicle_code_id=ck_vehicle_code,
                vendor_date_time__date=dt_start_date
            ).aggregate(
                total_count=Count('id')
            )['total_count']

            ### UPDATE VEHICLE SCHEDULE GPRS API RECORD
            UpdateVehiclScheduleGPRSApi = VehicleScheduleGPRSApi.objects.get(
                veh_sch_date=dt_start_date,
                vehicle_code_id=get_vehicle_id
            )
            UpdateVehiclScheduleGPRSApi.retrieve_record = tracker_gprs_length

            if tracker_gprs_length >= total_vendor_record:
                UpdateVehiclScheduleGPRSApi.process_status = "Completed"

            UpdateVehiclScheduleGPRSApi.save()
            ### UPDATE VEHICLE SCHEDULE GPRS API RECORD

        ### API RESPONSE LOOP ###

    ### RECORD EXIST ###

    ### CHECK VEHICLE GPRS LIST (END)

    django_datetime_from = parse_datetime(get_start_date)
    django_datetime_to = parse_datetime(get_end_date)

    tracker_raw_gprs_lists = list(TrackerRawData.objects.filter(
        vehicle_code_id=get_vehicle_id,
        vendor_date_time__gte=django_datetime_from,
        vendor_date_time__lte=django_datetime_to
    ).annotate(
        x=RawSQL("ST_X(geom)", []),
        y=RawSQL("ST_Y(geom)", [])
    ).values("x", "y", "vehicle_code_id", "g_status", "system_date_time", "vendor_date_time", "device_status",
             "max_speed", "speed", "vehicle_status", "direction", "distance", "mileage"
             ).order_by('id'))

    ### VEHICLE TRACK CALCULATE LENGTH FROM GPRS DATA
    VehicleTracks = TrackerRawData.objects.filter(
        vehicle_code_id=get_vehicle_id,
        vendor_date_time__gte=django_datetime_from,
        vendor_date_time__lte=django_datetime_to
    ).values('vehicle_code_id').annotate(
        line=MakeLine('geom'),  # Create a line from grouped points
        length=Length(MakeLine('geom'))  # Calculate line length
    )

    # Convert data into JSON response
    VehicleTracks_Length = [
        {
            "vehicle_code_id": track["vehicle_code_id"],
            "length_meters": track["length"].m,  # Convert to meters
            "line_geojson": track["line"].geojson  # Convert to GeoJSON
        }
        for track in VehicleTracks
    ]

    message = "Success"
    params = {
        'message': message,
        'tracker_raw_gprs_lists': tracker_raw_gprs_lists,
        'vehicle_tracks_length': VehicleTracks_Length,
    }

    return HttpResponse(json.dumps(params, default=date_handler))


def VehicleListWithTypeCodeView(request):
    get_vehicle_type = request.POST['vehicle_type']
    get_vehicle_status = request.POST['vehicle_status']

    if get_vehicle_status == 'None':
        vehicle_list = list(VehicleData.objects.filter(vehicle_type=get_vehicle_type).values('vehicle_code',
                                                                                             'vehicle_type',
                                                                                             'register_no',
                                                                                             'pitb_code').order_by(
            'vehicle_type'))
    else:
        vehicle_list = list(VehicleLiveMonitor.objects.select_related('vehicle_code').filter(
            g_status=get_vehicle_status,
            vehicle_code__vehicle_type=get_vehicle_type
        ).values(
            'vehicle_code__vehicle_code',
            'vehicle_code__register_no',
            'vehicle_code__pitb_code',
            'vehicle_code__vehicle_type'
        ))

        # Rename keys
        vehicle_list = [
            {
                'vehicle_code': item['vehicle_code__vehicle_code'],
                'register_no': item['vehicle_code__register_no'],
                'pitb_code': item['vehicle_code__pitb_code'],
                'vehicle_type': item['vehicle_code__vehicle_type'],
            }
            for item in vehicle_list
        ]

    # Replace None with "waiting"
    for vehicle in vehicle_list:
        if vehicle['pitb_code'] is None:
            vehicle['pitb_code'] = "waiting"

    return HttpResponse(json.dumps({'cmd_list': vehicle_list}, default=date_handler))


def VehicleDataListWithTypeCode_Function(set_filter):
    vehicle_difference = ""

    return vehicle_difference

    # get_vehicle_type = request.POST['vehicle_type']
    # get_vehicle_status = request.POST['vehicle_status']
    #
    # if get_vehicle_status == 'None':
    #     vehicle_list = list(VehicleData.objects.filter(vehicle_type=get_vehicle_type).values('vehicle_code',
    #                                                                                          'vehicle_type',
    #                                                                                          'register_no',
    #                                                                                          'pitb_code').order_by(
    #         'vehicle_type'))
    # else:
    #     vehicle_list = list(VehicleLiveMonitor.objects.select_related('vehicle_code').filter(
    #         g_status=get_vehicle_status,
    #         vehicle_code__vehicle_type=get_vehicle_type
    #     ).values(
    #         'vehicle_code__vehicle_code',
    #         'vehicle_code__register_no',
    #         'vehicle_code__pitb_code',
    #         'vehicle_code__vehicle_type'
    #     ))
    #
    #     # Rename keys
    #     vehicle_list = [
    #         {
    #             'vehicle_code': item['vehicle_code__vehicle_code'],
    #             'register_no': item['vehicle_code__register_no'],
    #             'pitb_code': item['vehicle_code__pitb_code'],
    #             'vehicle_type': item['vehicle_code__vehicle_type'],
    #         }
    #         for item in vehicle_list
    #     ]
    #
    # # Replace None with "waiting"
    # for vehicle in vehicle_list:
    #     if vehicle['pitb_code'] is None:
    #         vehicle['pitb_code'] = "waiting"
    #
    # return HttpResponse(json.dumps({'cmd_list': vehicle_list}, default=date_handler))


### AJAX END

### FUNCTION START

# GET GEO-FENCE BASED CONTAINER GPRS LOCATION
def ContainerGeoFence_TrackerRawData_Function():
    cursor = connections['default'].cursor()

    format_date = '%Y-%m-%d'
    current_data_time = datetime.datetime.now()
    today_date = current_data_time.strftime(format_date)

    ### RETRIEVE GPRS LOCATION POINT BASED ON GEO-FENCE INTERSECTION
    query_geo_fence_container = "WITH resu AS(WITH geo_f AS( SELECT geofence_code, feature_code, container_code, size AS container_size, gf.geom AS gf_geom FROM tbl_asset_geo_fence AS gf INNER JOIN tbl_container_data AS cont ON gf.feature_code = cont.container_code WHERE asset_code_id = 'AT-7' ) SELECT gprs_raw_code, container_code, container_size, vehicle_code_id, vendor_date_time, gf_geom, loc.geom AS gprs_geom FROM tbl_tracker_raw_data AS loc, geo_f WHERE vendor_date_time::date = '2025-02-27' AND st_intersects(geo_f.gf_geom, loc.geom) ) SELECT gprs_raw_code, container_code, vendor_date_time, gprs_geom FROM resu ORDER BY SPLIT_PART(gprs_raw_code, '-', 2)::INTEGER; "
    cursor.execute(query_geo_fence_container)
    geo_fence_container_lists = DictinctFetchAll(cursor)

    geo_fence_container_length = len(geo_fence_container_lists)
    if geo_fence_container_length > 0:
        for g in range(geo_fence_container_length):
            ck_gprs_raw_code = geo_fence_container_lists[g]['gprs_raw_code']

            ExistContainerGPRS = ContainerTrackerGPRS.objects.filter(gprs_raw_code_id=ck_gprs_raw_code)
            if len(ExistContainerGPRS) == 0:
                InstContainerTrackerGPRS = ContainerTrackerGPRS(
                    gprs_raw_code_id=ck_gprs_raw_code,
                    container_code_id=geo_fence_container_lists[g]['container_code'],
                    created_at=current_data_time,
                    created_by="admin"
                )
                InstContainerTrackerGPRS.save()
    ### RETRIEVE GPRS LOCATION POINT BASED ON GEO-FENCE INTERSECTION

    ### CALCULATE THE STAY TIME NEAR TO CONTAINER (START)
    query_stay_time_container = "WITH resu AS ( WITH geo_f AS ( SELECT vehicle_code_id, container_code_id, vendor_date_time FROM tbl_container_tracker_gprs AS ctr INNER JOIN tbl_tracker_raw_data AS tr ON ctr.gprs_raw_code_id = tr.gprs_raw_code WHERE vendor_date_time::date = '2025-02-27') SELECT vehicle_code_id, container_code_id, vendor_date_time, MIN(vendor_date_Time) OVER (PARTITION BY vehicle_code_id, container_code_id, TO_CHAR(vendor_date_time,'YYYY-MM-DD HH24' )) min_date, MAX(vendor_date_Time) OVER (PARTITION BY vehicle_code_id, container_code_id, TO_CHAR(VENDOR_DATE_TIME,'YYYY-MM-DD HH24' )) max_date FROM geo_f )  SELECT DISTINCT vehicle_code_id, container_code_id, max_date, min_date, TO_CHAR((max_date - min_date), 'HH24:MI:SS') as duration, TO_CHAR(EXTRACT(EPOCH FROM (TO_CHAR((max_date - min_date), 'HH24:MI:SS'))::time) / 60, '999.99') AS total_minutes FROM resu; "
    cursor.execute(query_stay_time_container)
    stay_time_container_lists = DictinctFetchAll(cursor)

    stay_time_container_length = len(stay_time_container_lists)
    if stay_time_container_length > 0:
        for s in range(stay_time_container_length):

            ### GET PARAMETER (STAY TIME)
            st_vehicle_code = stay_time_container_lists[s]['vehicle_code_id']
            st_container_code = stay_time_container_lists[s]['container_code_id']
            st_min_date = stay_time_container_lists[s]['min_date']
            st_max_date = stay_time_container_lists[s]['max_date']
            st_net_spent = stay_time_container_lists[s]['duration']
            st_total_minutes = float(stay_time_container_lists[s]['total_minutes'].strip())

            if st_total_minutes > 0:  ### NET SPENT TIME GREATER THEN ZERO

                ### ASSIGN CONTAINER PROCESS TYPE
                assign_process_type = "CPT-2"
                if st_total_minutes > 0 and st_total_minutes < 1:
                    assign_process_type = "CPT-3"
                elif st_total_minutes > 1:  ### VISITED
                    assign_process_type = "CPT-1"
                ### ASSIGN CONTAINER PROCESS TYPE

                ### UPDATE CONTAINER PROCESS TYPE
                # RetriveContainProcess = ContainerProcess.objects.filter(created_at__date=today_date)
                FetchContainerProcess = ContainerProcess.objects.filter(created_at__date="2025-02-27",
                                                                        container_code_id=st_container_code)
                if len(FetchContainerProcess) > 0:
                    FeatureExist = "No"
                    Ex_Created_AT = datetime.datetime.now()
                    for cp in range(len(FetchContainerProcess)):

                        verif_check_in = FetchContainerProcess[cp].check_in
                        verif_check_out = FetchContainerProcess[cp].check_out

                        if verif_check_in == st_min_date and verif_check_out == st_max_date:
                            pass
                        else:
                            ck_cont_proc_code = FetchContainerProcess[cp].container_process_code
                            ck_cont_proc_type = FetchContainerProcess[cp].cont_proc_type_code_id
                            Ex_Created_AT = FetchContainerProcess[cp].created_at

                            ### OTHER THEN VISITED CONTAINER
                            if ck_cont_proc_type == "CPT-1":
                                FeatureExist = "Yes"
                            else:
                                UpdateProcessType = ContainerProcess.objects.get(
                                    container_process_code=ck_cont_proc_code)
                                UpdateProcessType.vehicle_code_id = st_vehicle_code
                                UpdateProcessType.check_in = st_min_date
                                UpdateProcessType.check_out = st_max_date
                                UpdateProcessType.net_spent = st_net_spent
                                UpdateProcessType.updated_at = datetime.datetime.now()
                                UpdateProcessType.updated_by = "admin"
                                UpdateProcessType.cont_proc_type_code_id = assign_process_type
                                UpdateProcessType.save()

                    ### LOOP END

                    #### NEW CONTAINER PROCESS INFORMATION ADDED
                    if FeatureExist == "Yes":
                        if assign_process_type == "CPT-1":
                            auto_process_code = AutoGenerateCodeForModel(ContainerProcess, "container_process_code",
                                                                         "CP-")
                            # INSERT CONTAINER PROCESS DATA (START)
                            InstContainerProcess = ContainerProcess(
                                container_process_code=auto_process_code,
                                container_code_id=st_container_code,
                                vehicle_code_id=st_vehicle_code,
                                cont_proc_type_code_id=assign_process_type,
                                check_in=st_min_date,
                                check_out=st_max_date,
                                net_time_spent=st_net_spent,
                                created_at=Ex_Created_AT,
                                created_by="admin"
                            )
                            InstContainerProcess.save()
                    #### NEW CONTAINER PROCESS INFORMATION ADDED

                ### CONTAINER EXIST
            ### NET SPENT TIME GREATER THEN ZERO

        ### CALCULATE THE STAY TIME NEAR TO CONTAINER (END)

    return True


### GET TOTAL VEHICLE DIFFERENCE FROM API BY VEHICLE DATA
def VehicleDifference_Api_By_VehicleData_Function():
    total_vehicle_records = VehicleData.objects.count()

    api_vehicle_records = 0
    ### GET RESPONSE FROM VEHICLE API FUNCTION
    response_data = ResponseVehicleApi_By_Vendor_Function()
    if len(response_data) > 0:  ### RECORD EXIST ###
        vehicle_response_api = response_data['MainData']
        api_vehicle_records = len(vehicle_response_api)

    vehicle_difference = api_vehicle_records - total_vehicle_records

    return vehicle_difference


### GENERATE WORKING SCHEDULA WITH VEHICLE SCHEDULE IN SYSTEM FUNCTION
def GenerateWorkingWithVehicleSchedule_Function(selected_date):
    response_message = ""
    current_data_time = datetime.datetime.now()

    ### GENERATE VEHICLE SCHEDULA IN SYSTEM FUNCTION len(vehicle_list)
    vehicle_list = list(VehicleData.objects.filter(status="Active").order_by('vehicle_type'))
    ### IF CURRENT DATE EXIST OR NOT
    veh_sche_date_records = VehicleScheduleGPRSApi.objects.filter(veh_sch_date=selected_date)

    if len(vehicle_list) == len(veh_sche_date_records):
        response_message = "Equal Record"
        return response_message
    ### IF VEHICLE AND SCHEDULE VEHICLE EQUAL

    ### GENERATE WORKING SCHEDULA
    qs = WorkScheduleGPRSApi.objects.filter(work_date=selected_date)
    if not qs.exists():
        generated_code = AutoGenerateCodeForModel(WorkScheduleGPRSApi, "code", "WS-")
        new_record = WorkScheduleGPRSApi.objects.create(
            code=generated_code,
            work_date=selected_date,
            run_count=0,
            process_status="Pending",
            description="Sync-Current",
            created_at=current_data_time,
            created_by="admin"
        )
        ws_code = new_record.code
    else:
        existing = qs.first()
        ws_code = existing.code

    ### INSERTED ALL VEHICLE IN SCHEDULA MODEL
    for v in range(len(vehicle_list)):

        set_vehicle_code = vehicle_list[v].vehicle_code
        ### IF CURRENT DATE EXIST OR NOT
        current_date_records = VehicleScheduleGPRSApi.objects.filter(veh_sch_date=selected_date,
                                                                     vehicle_code_id=set_vehicle_code)
        if len(current_date_records) == 0:  ## IF CURRENT DATE EXIST
            response_message = "Created Record"
            gprs_code_number = VehicleScheduleGPRSApi.objects.count() + 1
            auto_vs_code = f"VAI-{gprs_code_number}"

            ### GPRS - VEHICLE SCHEDULE API (START)
            InstVehSche = VehicleScheduleGPRSApi(
                veh_api_code=auto_vs_code,
                vehicle_code_id=set_vehicle_code,
                veh_sch_date=selected_date,
                wsa_code_id=ws_code,
                retrieve_record=0,
                vendor_record=0,
                process_status="Pending",
                created_at=current_data_time,
                created_by="admin"
            )
            InstVehSche.save()
    #         ### GPRS - VEHICLE SCHEDULE API (END)
    #     ### NO RECORD FOUND CONDITION (END)
    # ### LOOP CONDITION (END)

    return response_message


### GENERATE WORKING SCHEDULA IN SYSTEM FUNCTION
def GenerateWorkingSchedule_Function():
    response_message = False

    # format_str_date_time = "%Y-%m-%d %H:%M:%S"
    format_str_date = "%Y-%m-%d"
    current_data_time = datetime.datetime.now()
    current_date = current_data_time.date()
    start_of_day_time_str = '00:00:00'
    end_of_day_time_str = '23:59:59'

    ###  GENERATE CODE IN WORKING SCHEDULE  ###
    WorkSchedule = WorkScheduleGPRSApi.objects.all()
    if len(WorkSchedule) == 0:
        ## IF SYSTEM FIRST TIME RUN THEN ##

        response_message = True
        InstWorkSche_id = 0
        loop_count = 0
        loop_empty = 0
        loop_running = True
        while loop_running:

            if loop_count == 0:  ## APPLY ON CURRENT DATE
                work_date_str = current_date.strftime(format_str_date)

            else:  ## APPLY ON PREVIOUS DATE TILL FIRST DEVICE INSTALL DATE
                # work_date_str =
                converted_date = datetime.strptime(work_date_str, "%Y-%m-%d")
                work_date_str = (converted_date - timedelta(days=1)).strftime(format_str_date)
            ###### ELSE END

            ### COMBINE DATE AND TIME INTO STRING
            datetime_from = work_date_str + " " + start_of_day_time_str  ## "2025-02-27 00:00:00"
            datetime_to = work_date_str + " " + end_of_day_time_str  ## "2025-02-27 23:59:59"

            ### GET VENDOR SERVER RESPONSE FROM TRACKER GPRS RAW DATA API FUNCTION
            response_gprs_api = ResponseTrackerGPRSRawApi_By_Vendor_Function(0, datetime_from, datetime_to)

            ### API RESPONSE CHECK ###
            if len(response_gprs_api) > 0:  ### RECORD EXIST ###
                converted_date_inst = datetime.strptime(work_date_str, '%Y-%m-%d').date()
                auto_gprs_code = AutoGenerateCodeForModel(WorkScheduleGPRSApi, "code",
                                                          "WS-")
                ### GPRS - WORKING SCHEDULE API (START)
                InstWorkSche = WorkScheduleGPRSApi(
                    code=auto_gprs_code,
                    work_date=converted_date_inst,
                    run_count=0,
                    process_status="Pending",
                    created_at=current_data_time,
                    created_by="admin"
                )
                InstWorkSche.save()
                InstWorkSche_id = InstWorkSche.id  # This is the latest ID after save
                ### GPRS - WORKING SCHEDULE API (END)
            else:  ## GPRS - NOT RECORD FOUND THEN
                loop_empty += 1

            loop_count += 1
            if loop_empty > 3:
                if InstWorkSche_id > 0:
                    loop_running = False
                    UpdateWorkSchedule = WorkScheduleGPRSApi.objects.filter(id=InstWorkSche_id).update(
                        description="Sync-Finished",
                        updated_at=current_data_time,
                        updated_by="admin"
                    )
        ### LOOP STOP
    ## IF SYSTEM FIRST TIME RUN THEN END ##
    else:  #

        WorkScheduleEnd = WorkScheduleGPRSApi.objects.filter(description="Sync-Finished")

        if len(WorkScheduleEnd) == 0:  # UPDATE START APPLICATION REMAINING DATE
            pass
            # WorkSchedule_MinDate = WorkScheduleGPRSApi.objects.aggregate(min_date=Min('work_date'))
            #
            # converted_date = WorkSchedule_MinDate['min_date']
            # work_date_str = (converted_date - timedelta(days=1)).strftime(format_str_date)

        else:
            ### IF CURRENT DATE EXIST OR NOT
            current_date_records = WorkScheduleGPRSApi.objects.filter(work_date=current_date)
            if len(current_date_records) > 0:  ## IF CURRENT DATE EXIST
                pass
            else:  ## IF CURRENT DATE NOT EXIST

                WorkSchedule_MaxDate = WorkScheduleGPRSApi.objects.aggregate(max_date=Max('work_date'))
                last_max_date = WorkSchedule_MaxDate['max_date']

                date_list = []
                current = last_max_date + timedelta(days=1)
                while current <= current_date:
                    date_list.append(current)
                    current += timedelta(days=1)
                print(date_list)
                # string_dates = [d.strftime("%Y-%m-%d") for d in date_list]
                # print(string_dates)

                for i in range(len(date_list)):
                    response_message = True
                    affected_new_date = date_list[i]

                    ### DATE CONVERT INTO STRING FORMAT
                    work_date_str = affected_new_date.strftime(format_str_date)

                    ### COMBINE DATE AND TIME INTO STRING
                    datetime_from = work_date_str + " " + start_of_day_time_str
                    datetime_to = work_date_str + " " + end_of_day_time_str

                    ### GET RESPONSE GEO VEHICLE SUMMARY API FUNCTION
                    response_geo_veh_api = ResponseGeoVehicleSummaryApi_By_Vendor_Function(0,
                                                                                           datetime_from,
                                                                                           datetime_to)
                    ### API RESPONSE CHECK ###
                    if len(response_geo_veh_api) > 0:  ### RECORD EXIST ###

                        converted_date_inst = datetime.strptime(work_date_str, '%Y-%m-%d').date()
                        auto_gprs_code = AutoGenerateCodeForModel(WorkScheduleGPRSApi, "code",
                                                                  "WS-")
                        ### GPRS - WORKING SCHEDULE API (START)
                        InstWorkSche = WorkScheduleGPRSApi(
                            code=auto_gprs_code,
                            work_date=converted_date_inst,
                            run_count=0,
                            process_status="Pending",
                            description="Sync-Current",
                            created_at=current_data_time,
                            created_by="admin"
                        )
                        InstWorkSche.save()
                        ### GPRS - WORKING SCHEDULE API (END)

                        ### GENERATE VEHICLE SCHEDULA IN SYSTEM FUNCTION
                        GenerateVehicleSchedule_Function("", converted_date_inst)
                        # GenerateVehicleSchedule_Function(response_geo_veh_api, converted_date_inst)

                    ### API RECORD EXIST (END) ###

                ### MISSING DATE LIST LOOP (END)

            ## IF CURRENT DATE NOT EXIST (END)

        ## Sync-Finished CONDITION (END)

    return response_message


### CALCULATE VEHICLE TRIP FROM AND TO SIDE FUNCTION
def CalculateVehicleTripFrom_To_Side_Function(vehicle_code, selected_date):
    message = False

    return message


### GENERATE VEHICLE SCHEDULA IN SYSTEM FUNCTION
def GenerateVehicleSchedule_Function(response_json_api, selected_date):
    response_message = False
    current_data_time = datetime.datetime.now()

    ### Filter and Get List of Code Values
    ws_code = list(WorkScheduleGPRSApi.objects.filter(work_date=selected_date).values_list('code', flat=True))[0]

    vehicle_list = []
    if response_json_api == "":
        vehicle_list = list(VehicleData.objects.filter(status="Active").order_by('vehicle_type'))
    else:
        vehicle_list = response_json_api
        ### API RESPONSE CHECK ###

    for v in range(len(vehicle_list)):

        if response_json_api == "":
            set_vehicle_code = vehicle_list[v].vehicle_code
        else:  ### API RESPONSE CHECK ###
            set_vehicle_code = vehicle_list[v]['VehicleID']

        ### IF CURRENT DATE EXIST OR NOT
        current_date_records = VehicleScheduleGPRSApi.objects.filter(veh_sch_date=selected_date,
                                                                     vehicle_code_id=set_vehicle_code)
        if len(current_date_records) == 0:  ## IF CURRENT DATE EXIST
            response_message = True
            gprs_code_number = VehicleScheduleGPRSApi.objects.count() + 1
            auto_vs_code = f"VAI-{gprs_code_number}"
            ### GPRS - VEHICLE SCHEDULE API (START)
            InstVehSche = VehicleScheduleGPRSApi(
                veh_api_code=auto_vs_code,
                vehicle_code_id=set_vehicle_code,
                veh_sch_date=selected_date,
                wsa_code_id=ws_code,
                retrieve_record=0,
                vendor_record=0,
                process_status="Pending",
                created_at=current_data_time,
                created_by="admin"
            )
            InstVehSche.save()
            ### GPRS - VEHICLE SCHEDULE API (END)
        ### NO RECORD FOUND CONDITION (END)
    ### LOOP CONDITION (END)

    return response_message


### SYNC TRACKER GPRS VEHICLE DATA BY VENDOR SYSTEM
def SyncTrackerGPRSVehicleData_By_Vendor_Function(request, vehicle_code: str, selected_date: str, from_time: str,
                                                  to_time: str):
    message = ""
    get_vehicle_id = vehicle_code
    get_selected_date = selected_date
    get_from_time = from_time
    get_to_time = to_time

    ### STEP-1 GENERATE WORKING SCHEDULA WITH VEHICLE SCHEDULE IN SYSTEM FUNCTION
    ck_process_status = ""
    Check_Vehicle_Schedule = VehicleScheduleGPRSApi.objects.filter(
        vehicle_code_id=get_vehicle_id,
        veh_sch_date=selected_date
    ).first()
    if not Check_Vehicle_Schedule:
        GenerateWorkingWithVehicleSchedule_Function(selected_date)
    else:
        ck_process_status = Check_Vehicle_Schedule.process_status
        ck_vehicle_type = Check_Vehicle_Schedule.vehicle_code.vehicle_type

    format_str_date = "%Y-%m-%d"
    format_str = "%Y-%m-%d %H:%M:%S"
    current_data_time = datetime.datetime.now()

    today_date = current_data_time.date()
    str_today_date = today_date.strftime(format_str_date)

    ### IF DATE NOT EXIST
    dt_start_date = datetime.datetime.strptime(get_selected_date, format_str_date).date()

    if ck_process_status == "Pending":  ### PROCESS STATUS "PENDING" ###

        ### COMBINE DATE AND TIME INTO STRING
        datetime_from = get_selected_date + " " + get_from_time  ## "2025-02-27 00:00:00"
        datetime_to = get_selected_date + " " + get_to_time  ## "2025-02-27 23:59:59"

        ### GET VENDOR SERVER RESPONSE FROM TRACKER GPRS RAW DATA API FUNCTION
        response_gprs_api = ResponseTrackerGPRSRawApi_By_Vendor_Function(get_vehicle_id, datetime_from, datetime_to)

        ### API RESPONSE CHECK ###
        if len(response_gprs_api) > 0:  ### RECORD EXIST ###
            # tracker_gprs_length = 0

            vehicle_gprs_response_api = response_gprs_api['Table']
            total_vendor_record = len(vehicle_gprs_response_api)

            for g in range(total_vendor_record):
                ck_vehicle_code = vehicle_gprs_response_api[g]['VehicleID']
                ck_vendor_date_time = vehicle_gprs_response_api[g]['GpsTime']

                try:
                    if '.' in ck_vendor_date_time:
                        ck_vendor_date_time = ck_vendor_date_time.split('.')[0]  # Remove the decimal part

                    _ck_vendor_date_time = datetime.datetime.fromisoformat(ck_vendor_date_time)

                    # Check if request is from a server IP
                    server_request = Retrieve_IP_Address(request)
                    if server_request == "Local Development":
                        # Add 5 hours
                        formatted_vendor_date_time = (_ck_vendor_date_time + timedelta(hours=5)).strftime(format_str)
                    else:
                        formatted_vendor_date_time = _ck_vendor_date_time.strftime(format_str)  # Format as a string

                except Exception as e:
                    print(f"_ck_vendor_date_time - {ck_vendor_date_time}")

                ### TRACKER RAW DATA (START)
                GetTrackerRawData = TrackerRawData.objects.filter(
                    vehicle_code_id=ck_vehicle_code,
                    vendor_date_time=formatted_vendor_date_time
                )

                if not GetTrackerRawData.exists():
                    vt_latitude = vehicle_gprs_response_api[g]['Lat']
                    vt_longitude = vehicle_gprs_response_api[g]['Long']
                    get_feature_coordinate = "POINT(" + str(vt_longitude) + " " + str(vt_latitude) + ")"

                    InstTrackerRawData = TrackerRawData(
                        # gprs_raw_code=auto_gprs_code,
                        vehicle_code_id=ck_vehicle_code,
                        terminal_no=vehicle_gprs_response_api[g]['TerminalNo'],
                        geom=get_feature_coordinate,
                        latitude=vt_latitude,
                        longitude=vt_longitude,
                        g_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        vehicle_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        device_status=vehicle_gprs_response_api[g]['Dev_status'],
                        vendor_date_time=formatted_vendor_date_time,
                        system_date_time=vehicle_gprs_response_api[g]['RecTime'],
                        speed=vehicle_gprs_response_api[g]['Speed'],
                        distance=vehicle_gprs_response_api[g]['Distance'],
                        direction=vehicle_gprs_response_api[g]['Direction'],
                        mileage=vehicle_gprs_response_api[g]['Mileage_Val'],
                        location=vehicle_gprs_response_api[g]['Location'],
                        acc_status=vehicle_gprs_response_api[g]['Acc_Status'],
                        rpm=vehicle_gprs_response_api[g]['Rpm'],
                        engine_temp=vehicle_gprs_response_api[g]['Engin_temp'],
                        engine_hours=vehicle_gprs_response_api[g]['Engine_Hours'],
                        fuel_level=vehicle_gprs_response_api[g]['FuelLevel'],
                        fuel_consumed=vehicle_gprs_response_api[g]['FuelConsumed'],
                        gps_satelite=vehicle_gprs_response_api[g]['Gps_satelite'],
                        gsm_signal=vehicle_gprs_response_api[g]['Gsm_signal'],
                        ext_bat_voltage=vehicle_gprs_response_api[g]['Ext_bat_voltage'],
                        int_bat_voltage=vehicle_gprs_response_api[g]['Int_bat_voltage'],
                        created_at=current_data_time,
                        created_by="admin"
                    )
                    InstTrackerRawData.save()
                else:
                    # Update existing record with the specified fields
                    tracker_raw_data = GetTrackerRawData.first()
                    tracker_raw_data.location = vehicle_gprs_response_api[g]['Location']
                    tracker_raw_data.gis_geo_status = vehicle_gprs_response_api[g]['Veh_Status']
                    tracker_raw_data.acc_status = vehicle_gprs_response_api[g]['Acc_Status']
                    tracker_raw_data.rpm = vehicle_gprs_response_api[g]['Rpm']
                    tracker_raw_data.engine_temp = vehicle_gprs_response_api[g]['Engin_temp']
                    tracker_raw_data.engine_hours = vehicle_gprs_response_api[g]['Engine_Hours']
                    tracker_raw_data.fuel_level = vehicle_gprs_response_api[g]['FuelLevel']
                    tracker_raw_data.fuel_consumed = vehicle_gprs_response_api[g]['FuelConsumed']
                    tracker_raw_data.gps_satelite = vehicle_gprs_response_api[g]['Gps_satelite']
                    tracker_raw_data.gsm_signal = vehicle_gprs_response_api[g]['Gsm_signal']
                    tracker_raw_data.ext_bat_voltage = vehicle_gprs_response_api[g]['Ext_bat_voltage']
                    tracker_raw_data.int_bat_voltage = vehicle_gprs_response_api[g]['Int_bat_voltage']
                    tracker_raw_data.updated_at = current_data_time
                    tracker_raw_data.updated_by = "admin"
                    tracker_raw_data.save()
                    # TRACKER RAW DATA (END)
                # RECORD FOUND (END)
                # tracker_gprs_length = tracker_gprs_length + 1
            ### LOOP END

            # Set up filters for today's data
            tracker_filters = {
                'vendor_date_time__date': dt_start_date,
                'vehicle_code_id': get_vehicle_id
            }

            # Get tracker records for this vehicle today
            vehicle_tracker_records = TrackerRawData.objects.filter(**tracker_filters).order_by(
                'vendor_date_time')

            tracker_gprs_length = vehicle_tracker_records.aggregate(total_count=Count('id'))['total_count']

            """
            Calculate total distance (in km) for a vehicle on a given date.
            Skips records without geometry.
            """
            round_distance_km = DistanceFinder(vehicle_tracker_records)

            """
            Calculate total working (in hour) for a vehicle on a given date.
            """
            fun_working_hours = CalculateSingleVehicleWorkingHour_Function(vehicle_tracker_records)

            up_ignition_status = "No"
            vehicle_type_threshold = VehicleThreshold.objects.filter(
                vehicle_type=ck_vehicle_type
            ).values('distance', 'working_hours', 'min_distance')
            if vehicle_type_threshold.exists():
                # .values() returns a dict, so use index [0] to get the first record
                veh_sche_distance = vehicle_type_threshold[0]['distance']
                veh_sche_min_distance = vehicle_type_threshold[0]['min_distance']
                veh_sche_working_hours = vehicle_type_threshold[0]['working_hours']
                if veh_sche_min_distance is None:
                    veh_sche_min_distance = 0
                if veh_sche_working_hours is None:
                    veh_sche_working_hours = 0

                if float(round_distance_km) > float(veh_sche_min_distance):
                    up_ignition_status = "Yes"
                if float(fun_working_hours) > float(veh_sche_working_hours):
                    up_ignition_status = "Yes"

            if get_selected_date != str_today_date:

                ### UPDATE VEHICLE SCHEDULE GPRS API RECORD
                try:
                    UpdateVehicleScheduleGPRSApi = VehicleScheduleGPRSApi.objects.get(
                        veh_sch_date=dt_start_date,
                        vehicle_code_id=get_vehicle_id
                    )

                    UpdateVehicleScheduleGPRSApi.distance = float(round_distance_km)
                    UpdateVehicleScheduleGPRSApi.working_hours = float(fun_working_hours)
                    UpdateVehicleScheduleGPRSApi.retrieve_record = tracker_gprs_length
                    UpdateVehicleScheduleGPRSApi.ignition_status = up_ignition_status

                    if total_vendor_record <= tracker_gprs_length:
                        UpdateVehicleScheduleGPRSApi.process_status = "Completed"

                    UpdateVehicleScheduleGPRSApi.save()

                except VehicleScheduleGPRSApi.DoesNotExist:
                    print(f"Vehicle schedule record not found for vehicle_id {get_vehicle_id} on {dt_start_date}")
                ### UPDATE VEHICLE SCHEDULE GPRS API RECORD

            message = "RECORD FOUND"

        ### API RESPONSE LOOP ###

    ### RECORD EXIST ###
    return message


# This is the function which was converted from view to function
def FetchSingleVehicleTripHistoryData_Function(get_vehicle_id, get_start_date, get_end_date):
    format_str_date = "%Y-%m-%d"
    format_str_date_time = "%Y-%m-%d %H:%M:%S"
    current_data_time = datetime.datetime.now()

    ### STEP LIST
    # STEP-1 SET DATE IN DJANGO FORMAT

    # STEP-1 SET DATE IN DJANGO FORMAT
    ### IF DATE NOT EXIST
    if isinstance(get_start_date, datetime):
        dt_start_date = get_start_date.date()
    else:
        try:
            # Check if it's in HTML5 format
            if 'T' in get_start_date:
                if len(get_start_date.split(':')) == 2:  # Missing seconds
                    get_start_date = get_start_date + ":00"  # Add seconds
                get_start_date = get_start_date.replace('T', ' ')  # Replace T with space
            dt_start_date = datetime.strptime(get_start_date, format_str_date_time).date()
        except ValueError:
            # If that fails, try to parse just the date part
            dt_start_date = datetime.strptime(get_start_date.split('T')[0], format_str_date).date()

    if isinstance(get_end_date, datetime):
        dt_end_date = get_end_date.date()
    else:
        try:
            # Check if it's in HTML5 format
            if 'T' in get_end_date:
                if len(get_end_date.split(':')) == 2:  # Missing seconds
                    get_end_date = get_end_date + ":00"  # Add seconds
                get_end_date = get_end_date.replace('T', ' ')  # Replace T with space
            dt_end_date = datetime.strptime(get_end_date, format_str_date_time).date()
        except ValueError:
            # If that fails, try to parse just the date part
            dt_end_date = datetime.strptime(get_end_date.split('T')[0], format_str_date).date()

    ### STEP-2 FETCH NO OF DATE WHICH IS NOT EXIST
    range_date_list = []
    between_date_list = []
    current = dt_start_date
    while current <= dt_end_date:
        ### IF VEHICLE SCHEDULE DATE EXIST OR NOT
        vehicle_schedule_record = VehicleScheduleGPRSApi.objects.filter(veh_sch_date=current,
                                                                        vehicle_code_id=get_vehicle_id)
        if len(vehicle_schedule_record) == 0:  ## IF SELECTED DATE EXIST
            between_date_list.append(current)

        range_date_list.append(current)
        current += datetime.timedelta(days=1)
    # print(date_list)
    # string_dates = [d.strftime("%Y-%m-%d") for d in date_list]

    if len(between_date_list) > 0:
        for i in range(len(between_date_list)):
            converted_date_inst = between_date_list[i]

            ### IF CURRENT DATE EXIST OR NOT
            current_date_records = WorkScheduleGPRSApi.objects.filter(work_date=converted_date_inst)
            if len(current_date_records) == 0:  ## IF CURRENT DATE EXIST
                auto_gprs_code = AutoGenerateCodeForModel(WorkScheduleGPRSApi, "code", "WS-")
                ### GPRS - WORKING SCHEDULE API (START)
                InstWorkSche = WorkScheduleGPRSApi(
                    code=auto_gprs_code,
                    work_date=converted_date_inst,
                    run_count=0,
                    process_status="Pending",
                    description="Sync-Current",
                    created_at=current_data_time,
                    created_by="admin"
                )
                InstWorkSche.save()
                ### GPRS - WORKING SCHEDULE API (END)

            ### GENERATE VEHICLE SCHEDULA IN SYSTEM FUNCTION
            GenerateVehicleSchedule_Function("", converted_date_inst)
        ### LOOP (END)
    ### CHECK SCHEDULE ADDED OR NOT (END)
    ### STEP-2 FETCH NO OF DATE WHICH IS NOT EXIST (END)

    ### CHECK VEHICLE VENDOR RECORD EQUAL TO DUMP RECORD WITH STATUS
    fetch_vehicle_sche_record = VehicleScheduleGPRSApi.objects.get(veh_sch_date=dt_start_date,
                                                                   vehicle_code_id=get_vehicle_id)
    ck_process_status = fetch_vehicle_sche_record.process_status

    if ck_process_status == "Pending":  ### PROCESS STATUS "PENDING" ###
        ### CHECK VEHICLE GPRS LIST (START)

        start_of_day_time_str = '00:00:00'
        end_of_day_time_str = '23:59:59'
        ### DATE CONVERT INTO STRING FORMAT
        work_date_str = dt_start_date.strftime(format_str_date)

        ### COMBINE DATE AND TIME INTO STRING
        datetime_from = work_date_str + " " + start_of_day_time_str  ## "2025-02-27 00:00:00"
        datetime_to = work_date_str + " " + end_of_day_time_str  ## "2025-02-27 23:59:59"

        ### GET VENDOR SERVER RESPONSE FROM TRACKER GPRS RAW DATA API FUNCTION
        response_gprs_api = ResponseTrackerGPRSRawApi_By_Vendor_Function(get_vehicle_id, datetime_from, datetime_to)
        logger.info("Vendor Response JSON: %s", json.dumps(response_gprs_api))
        ### API RESPONSE CHECK ###
        if len(response_gprs_api) > 0:  ### RECORD EXIST ###
            total_vendor_record = len(response_gprs_api['Table'])

            vehicle_gprs_response_api = response_gprs_api['Table']
            vehicle_gprs_response_api_len = len(vehicle_gprs_response_api)
            for g in range(vehicle_gprs_response_api_len):
                ck_vehicle_code = vehicle_gprs_response_api[g]['VehicleID']
                tr_vendor_date_time = vehicle_gprs_response_api[g]['GpsTime']
                ### TRACKER RAW DATA (START)

                GetTrackerRawData = TrackerRawData.objects.filter(vehicle_code_id=ck_vehicle_code,
                                                                  vendor_date_time=tr_vendor_date_time)
                if len(GetTrackerRawData) == 0:
                    auto_gprs_code = AutoGenerateCodeForModel(TrackerRawData, "gprs_raw_code", "GPRS-")

                    vt_latitude = vehicle_gprs_response_api[g]['Lat']
                    vt_longitude = vehicle_gprs_response_api[g]['Long']
                    get_feature_coordinate = "POINT(" + str(vt_longitude) + " " + str(vt_latitude) + ")"

                    InstTrackerRawData = TrackerRawData(
                        gprs_raw_code=auto_gprs_code,
                        vehicle_code_id=ck_vehicle_code,
                        terminal_no=vehicle_gprs_response_api[g]['TerminalNo'],
                        geom=get_feature_coordinate,
                        latitude=vt_latitude,
                        longitude=vt_longitude,
                        g_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        vehicle_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        device_status=vehicle_gprs_response_api[g]['Dev_status'],
                        vendor_date_time=vehicle_gprs_response_api[g]['GpsTime'],
                        system_date_time=vehicle_gprs_response_api[g]['RecTime'],
                        speed=vehicle_gprs_response_api[g]['Speed'],
                        distance=vehicle_gprs_response_api[g]['Distance'],
                        direction=vehicle_gprs_response_api[g]['Direction'],
                        mileage=vehicle_gprs_response_api[g]['Mileage_Val'],
                        created_at=current_data_time,
                        created_by="admin"
                    )
                    InstTrackerRawData.save()
                    # TRACKER RAW DATA (END)
                # RECORD FOUND (END)
            ### LOOP END

            tracker_gprs_length = TrackerRawData.objects.filter(
                vehicle_code_id=ck_vehicle_code,
                vendor_date_time__date=dt_start_date
            ).aggregate(
                total_count=Count('id')
            )['total_count']

            ### UPDATE VEHICLE SCHEDULE GPRS API RECORD
            UpdateVehiclScheduleGPRSApi = VehicleScheduleGPRSApi.objects.get(
                veh_sch_date=dt_start_date,
                vehicle_code_id=get_vehicle_id
            )
            UpdateVehiclScheduleGPRSApi.retrieve_record = tracker_gprs_length

            if tracker_gprs_length >= total_vendor_record:
                UpdateVehiclScheduleGPRSApi.process_status = "Completed"

            UpdateVehiclScheduleGPRSApi.save()
            ### UPDATE VEHICLE SCHEDULE GPRS API RECORD

        ### API RESPONSE LOOP ###

    ### RECORD EXIST ###

    ### CHECK VEHICLE GPRS LIST (END)

    django_datetime_from = parse_datetime(get_start_date)
    django_datetime_to = parse_datetime(get_end_date)

    tracker_raw_gprs_lists = list(TrackerRawData.objects.filter(
        vehicle_code_id=get_vehicle_id,
        vendor_date_time__gte=django_datetime_from,
        vendor_date_time__lte=django_datetime_to
    ).annotate(
        x=RawSQL("ST_X(geom)", []),
        y=RawSQL("ST_Y(geom)", [])
    ).values("x", "y", "vehicle_code_id", "g_status", "system_date_time", "vendor_date_time", "device_status",
             "max_speed", "speed", "vehicle_status", "direction", "distance", "mileage"
             ).order_by('id'))

    ### VEHICLE TRACK CALCULATE LENGTH FROM GPRS DATA
    VehicleTracks = TrackerRawData.objects.filter(
        vehicle_code_id=get_vehicle_id,
        vendor_date_time__gte=django_datetime_from,
        vendor_date_time__lte=django_datetime_to
    ).values('vehicle_code_id').annotate(
        line=MakeLine('geom'),  # Create a line from grouped points
        length=Length(MakeLine('geom'))  # Calculate line length
    )

    # Convert data into JSON response
    VehicleTracks_Length = [
        {
            "vehicle_code_id": track["vehicle_code_id"],
            "length_meters": track["length"].m,  # Convert to meters
            "line_geojson": track["line"].geojson  # Convert to GeoJSON
        }
        for track in VehicleTracks
    ]

    message = "Success"
    params = {
        'message': message,
        'tracker_raw_gprs_lists': tracker_raw_gprs_lists,
        'vehicle_tracks_length': VehicleTracks_Length,
    }
    logger.info("FetchSingleVehicleTripHistoryData has been executed.")
    logger.info("-------------------------")
    return HttpResponse(json.dumps(params, default=date_handler))


###### API - FUNCTION
### GET RESPONSE FROM VEHICLE API FUNCTION
def ResponseVehicleApi_By_Vendor_Function():
    ### Base URL
    base_url = "https://labs3.unitedtracker.com/api/UTS360/GetUserAllVehiclesDetailwithLocationbyUserCredentials"

    ### Define parameters ###
    params = {
        "UserName": "AKHTAR6792",
        "Password": "AKHTAR@6792",
        "VehicleID": 0,
        "isExcLocation": False,  # Current time
        "isExcTeltonikaLoc": False  # 30 seconds before current time
    }

    ### Generate URL with dynamic dates
    url = f"{base_url}?{urlencode(params)}"
    headers = {
        'API_KEY': '7b6f169d544e4eda4b2b263e6bffe50d',
        'Authorization': 'Basic UElUQjo4NTMwNWFlMjg1ZjVhNzk1OWY4OWE4YWY5Y2FhNWY1Nw=='
    }

    response_status_code = 0
    response_data = []
    response_message = ""
    response_status = ""
    try:
        response = requests.get(url, headers=headers)
        response_status_code = response.status_code
        if response.status_code == 200:
            response_data = response.json()  # Get the response data from the external API
            response_status = "Success"
            response_message = "Data fetched successfully"
        else:
            response_status = "Error"
            response_message = "Data fetched successfully"
    except:
        response_status = "Url Error"
        response_message = "Url Error"

    return response_data


### GET RESPONSE GEO VEHICLE SUMMARY API FUNCTION
def ResponseGeoVehicleSummaryApi_By_Vendor_Function(vehicle_code, from_date, to_date):
    ### Base URL
    base_url = "https://labs3.unitedtracker.com//api/Trackers/GetVehicleWorkingScheduleByUserName"

    ### Define parameters ###
    params = {
        "UserName": "AKHTAR6792",
        "Password": "AKHTAR@6792",
        "VehicleID": vehicle_code,
        "fromDate": from_date,  # 2025-04-15 14:00:02
        "toDate": to_date  # toDate=2025-04-15 14:10:02
    }

    ### Generate URL with dynamic dates
    url = f"{base_url}?{urlencode(params)}"
    headers = {
        'API_KEY': '7b6f169d544e4eda4b2b263e6bffe50d',
        'Authorization': 'Basic UElUQjo4NTMwNWFlMjg1ZjVhNzk1OWY4OWE4YWY5Y2FhNWY1Nw=='
    }

    response_data = []
    response_message = ""
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response_data = response.json()  # Get the response data from the external API
            response_message = "Data fetched successfully"
        else:
            response_message = "Data Not Fetch"
    except:
        response_message = "Url Error"

    return response_data


### GET VENDOR SERVER RESPONSE FROM TRACKER GPRS RAW DATA API FUNCTION
def ResponseTrackerGPRSRawApi_By_Vendor_Function(vehicle_code, from_date, to_date):
    ### Base URL
    base_url = "https://labs3.unitedtracker.com/api/Trackers/GetVehicleHistoryByUserName"

    ### Define parameters ###
    params = {
        "UserName": "AKHTAR6792",
        "Password": "AKHTAR@6792",
        "VehicleID": vehicle_code,
        "fromDate": from_date,  # 2025-04-15 14:00:02
        "toDate": to_date  # toDate=2025-04-15 14:10:02
    }

    ### Generate URL with dynamic dates
    url = f"{base_url}?{urlencode(params)}"
    headers = {
        'API_KEY': '7b6f169d544e4eda4b2b263e6bffe50d',
        'Authorization': 'Basic UElUQjo4NTMwNWFlMjg1ZjVhNzk1OWY4OWE4YWY5Y2FhNWY1Nw=='
    }

    response_data = []
    response_message = ""
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response_data = response.json()  # Get the response data from the external API
            response_message = "Data fetched successfully"
        else:
            response_message = "Data Not Fetch"
    except:
        response_message = "Url Error"

    return response_data


### DELETED FUNCTION
def VehicleTrackerGPRSRecord_GISAPI_Function(vehicle_code, from_date, to_date):
    get_vehicle_code = vehicle_code
    if get_vehicle_code == "":
        get_vehicle_code = 0

    dateTime = datetime.datetime.now()

    # Base URL
    base_url = "http://202.142.158.118/WebTrackAPI/api/Trackers/GetVehicleHistoryByUserName"
    # Define parameters
    params = {
        "UserName": "WEB.GIS",
        "Password": "GIS@UTS1",
        # "VehicleID": get_vehicle_code,
        # "fromDate": from_date,  # Current time
        # "toDate": to_date  # 30 seconds before current time
        "VehicleID": "102795",
        "fromDate": "2025-02-27 00:00:02",  # Current time
        "toDate": "2025-02-27 16:00:02"  # 30 seconds before current time
    }
    # Generate URL with dynamic dates
    url = f"{base_url}?{urlencode(params)}"
    headers = {
        'API_KEY': '7b6f169d544e4eda4b2b263e6bffe50d',
        'Authorization': 'Basic UElUQjo4NTMwNWFlMjg1ZjVhNzk1OWY4OWE4YWY5Y2FhNWY1Nw=='
    }
    response_data = []
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response_data = response.json()  # Get the response data from the external API
            message = "Success"
        else:
            message = "Error"
    except:
        message = "Url Error"

    ### API RESPONSE CHECK ###
    if message == "Success":
        if len(response_data) > 0:  ### RECORD EXIST ###

            gprs_vehicle_obj = []

            vehicle_gprs_response_api = response_data['Table']
            vehicle_gprs_response_api_len = len(vehicle_gprs_response_api)
            for g in range(vehicle_gprs_response_api_len):
                ck_vehicle_code = vehicle_gprs_response_api[g]['VehicleID']
                tr_vendor_date_time = vehicle_gprs_response_api[g]['GpsTime']
                tr_system_date_time = vehicle_gprs_response_api[g]['RecTime']
                ### TRACKER RAW DATA (START)

                GetTrackerRawData = TrackerRawData.objects.filter(vehicle_code_id=ck_vehicle_code,
                                                                  system_date_time=tr_system_date_time,
                                                                  vendor_date_time=tr_vendor_date_time)
                if len(GetTrackerRawData) == 0:
                    auto_gprs_code = AutoGenerateCodeForModel(TrackerRawData, "gprs_raw_code", "GPRS-")

                    vt_latitude = vehicle_gprs_response_api[g]['Lat']
                    vt_longitude = vehicle_gprs_response_api[g]['Long']
                    get_feature_coordinate = "POINT(" + str(vt_longitude) + " " + str(vt_latitude) + ")"

                    InstTrackerRawData = TrackerRawData(
                        gprs_raw_code=auto_gprs_code,
                        vehicle_code_id=ck_vehicle_code,
                        terminal_no=vehicle_gprs_response_api[g]['TerminalNo'],
                        geom=get_feature_coordinate,
                        latitude=vt_latitude,
                        longitude=vt_longitude,
                        g_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        vehicle_status=vehicle_gprs_response_api[g]['Veh_Status'],
                        device_status=vehicle_gprs_response_api[g]['Dev_status'],
                        vendor_date_time=vehicle_gprs_response_api[g]['GpsTime'],
                        system_date_time=vehicle_gprs_response_api[g]['RecTime'],
                        speed=vehicle_gprs_response_api[g]['Speed'],
                        distance=vehicle_gprs_response_api[g]['Distance'],
                        direction=vehicle_gprs_response_api[g]['Direction'],
                        mileage=vehicle_gprs_response_api[g]['Mileage_Val'],
                        created_at=dateTime,
                        created_by="admin"
                    )
                    InstTrackerRawData.save()
                    # TRACKER RAW DATA (END)

                    # GENERATE VEHICLE OBJECT
                    gprs_dict = dict()
                    gprs_dict["vehicle_code"] = ck_vehicle_code
                    gprs_dict["longitude"] = vt_longitude
                    gprs_dict["latitude"] = vt_latitude
                    gprs_dict["geom"] = get_feature_coordinate
                    gprs_dict["g_status"] = vehicle_gprs_response_api[g]['Veh_Status']

                    item_double = "NO"
                    gprs_vehicle_obj_count = len(gprs_vehicle_obj)
                    if gprs_vehicle_obj_count > 0:
                        for v in range(int(gprs_vehicle_obj_count)):
                            obj_vehicle_code = gprs_vehicle_obj[v]['vehicle_code']
                            if obj_vehicle_code == ck_vehicle_code:
                                item_double = "YES"

                    if item_double == "NO":
                        gprs_vehicle_obj.append(gprs_dict)

                    # GENERATE VEHICLE OBJECT (END)
                # GREATER THEN ZERO
            ### API RESPONSE LOOP ###

            ### UPDATE LIVE MONITORING LOCATION START
            if len(gprs_vehicle_obj) > 0:
                for gl in range(len(gprs_vehicle_obj)):
                    ck_live_vehicle_code = gprs_vehicle_obj[gl]['vehicle_code']
                    # VEHICLE LIVE MONITORING (START)
                    VehicleLiveMonitor.objects.filter(vehicle_code_id=ck_live_vehicle_code).update(
                        geom=gprs_vehicle_obj[gl]['geom'],
                        longitude=gprs_vehicle_obj[gl]['longitude'],
                        latitude=gprs_vehicle_obj[gl]['latitude'],
                        g_status=gprs_vehicle_obj[gl]['g_status'],
                        updated_at=dateTime,
                        updated_by="admin"
                    )
                ### LOOP END
            ### UPDATE LIVE MONITORING LOCATION START
        ### RECORD EXIST ###
    ### API RESPONSE CHECK (END) ###

    return True


def GetVehicleDataExcel(request):
    # Get the exact same data as AllVehicleMonitoringView
    message = ""
    cursor = connections['default'].cursor()
    dateTime = datetime.datetime.now()

    # Get the same filters
    search_term = request.GET.get('vehicle_search', '')
    vehicle_type_filter = request.GET.get('cmd_vehicle_type')
    if vehicle_type_filter == 'NA':
        vehicle_type_filter = None
    vehicle_code_filter = request.GET.get('cmd_vehicle_list')
    if vehicle_code_filter == 'NA':
        vehicle_code_filter = None
    get_vehicle_status = request.GET.get('vehicle_status', 'NA')

    # Use the exact same query as AllVehicleMonitoringView
    query_feature = "SELECT ST_X(geom) as x, ST_Y(geom) as y, chasis_no, register_no, COALESCE(pitb_code, 'Waiting') AS pitb_code, vehicle_type, g_status, direction, speed, COALESCE(device_status, 'NA') AS device_status, COALESCE(ignition_status, 'NA') AS ignition_status, COALESCE(geo_location, 'NA') AS geo_location, vendor_date_time, COALESCE(duration, 'NA') AS duration FROM tbl_vehicle_live_monitor AS vlm LEFT OUTER JOIN tbl_vehicle_data AS veh ON vlm.vehicle_code_id = veh.vehicle_code WHERE 1=1"

    if get_vehicle_status != "NA":
        query_feature += f" AND g_status = '{get_vehicle_status}'"
    if search_term:
        query_feature += f" AND (veh.vehicle_code LIKE '%{search_term}%') "
    if vehicle_type_filter:
        query_feature += f" AND veh.vehicle_type = '{vehicle_type_filter}' "
    if vehicle_code_filter:
        query_feature += f" AND veh.vehicle_code = '{vehicle_code_filter}' "

    query_feature += ";"
    cursor.execute(query_feature)
    vehicle_live_lists = DictinctFetchAll(cursor)

    # Add the exact same distance and working hours calculation as the main view
    format_date = '%Y-%m-%d'
    today_date = dateTime.strftime(format_date)

    for idx, vehicle in enumerate(vehicle_live_lists):
        # Extract vehicle_code from the result
        vehicle_code = vehicle['vehicle_code_id'] if 'vehicle_code_id' in vehicle else vehicle.get('chasis_no')

        if vehicle_code:
            # Set up filters for today's data
            tracker_filters = {
                'vendor_date_time__date__range': (today_date, today_date),
                'vehicle_code_id': vehicle_code
            }

            # Get tracker records for this vehicle today
            vehicle_tracker_records = TrackerRawData.objects.filter(**tracker_filters).order_by('vendor_date_time')

            # Calculate distance
            distance_km = DistanceFinder(vehicle_tracker_records)

            # Calculate working hours
            try:
                set_working_hours = CalculateSingleVehicleWorkingHour_Function(vehicle_tracker_records)
                if set_working_hours and len(set_working_hours) > 0 and vehicle_code in set_working_hours[0]:
                    delta_working = set_working_hours[0][vehicle_code]['working']
                    working_hours_only = round(delta_working.total_seconds() / 60, 2)  # Convert to minutes
                else:
                    working_hours_only = 0.0
            except (IndexError, KeyError, TypeError):
                working_hours_only = 0.0

            # Add to vehicle data
            vehicle_live_lists[idx]['distance_km'] = distance_km
            vehicle_live_lists[idx]['working_hours'] = working_hours_only
        else:
            # Default values if no vehicle code found
            vehicle_live_lists[idx]['distance_km'] = "0.00"
            vehicle_live_lists[idx]['working_hours'] = 0.0

    # Create Excel workbook
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Vehicle Monitoring')

    # Define column headers that match the HTML table
    headers = [
        'Register No', 'PITB Code', 'Status', 'Location',
        'Engine Status', 'Distance (km)', 'Working Minutes', 'Speed (km/h)', 'Trip Date'
    ]

    # Apply styles to header row
    header_style = xlwt.easyxf(
        'font: bold on; align: wrap on, horiz center; '
        'pattern: pattern solid, fore_colour gray25'
    )

    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
        worksheet.col(col).width = 256 * 20

    # Write the data (same as what's displayed in HTML)
    for row, vehicle in enumerate(vehicle_live_lists, 1):
        # Engine status extraction
        engine_status = "Unknown"
        if vehicle.get('device_status'):
            import re
            m = re.findall(r'\b(Off|On)\b', vehicle['device_status'])
            engine_status = m[0] if m else "Unknown"

        # Write data to Excel
        worksheet.write(row, 0, vehicle.get('register_no', ''))
        worksheet.write(row, 1, vehicle.get('pitb_code', ''))
        worksheet.write(row, 2, vehicle.get('g_status', ''))
        worksheet.write(row, 3, vehicle.get('geo_location', ''))
        worksheet.write(row, 4, engine_status)
        worksheet.write(row, 5, vehicle.get('distance_km', '0.00'))
        worksheet.write(row, 6, f"{vehicle.get('working_hours', 0.0):.2f}")
        worksheet.write(row, 7, f"{vehicle.get('speed', 0)}")
        worksheet.write(row, 8, str(vehicle.get('vendor_date_time', '')))

    # Create response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Vehicle_Monitoring.xls"'
    workbook.save(response)
    return response


def export_to_excel(duration_data):
    # Convert to DataFrame
    df = pd.DataFrame(duration_data)

    # Ensure all expected columns are present and ordered
    columns = ['Vehicle Code', 'Pitb Code', 'Reg No', 'Working', 'Moving', 'Parked', 'Idle', 'Offline', 'No Response']

    # Create a new DataFrame with the desired columns
    result_df = pd.DataFrame(columns=columns)

    # Map the existing columns to the new DataFrame
    column_mapping = {
        'vehicle_code': 'Vehicle Code',
        'pitb_code': 'Pitb Code',
        'reg_no': 'Reg No',
        'working': 'Working',
        'moving': 'Moving',
        'parked': 'Parked',
        'idle': 'Idle',
        'offline': 'Offline',
        'no_response': 'No Response'
    }

    # Copy data from original DataFrame to new DataFrame
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            # Special handling for vehicle_code and pitb_code
            if old_col in ['vehicle_code', 'pitb_code']:
                # Replace empty strings and NaN with 'Null'
                result_df[new_col] = df[old_col].replace(['', 'nan', 'NaN', 'None', 'none'], 'Null')
                result_df[new_col] = result_df[new_col].fillna('Null')
            else:
                result_df[new_col] = df[old_col]
        else:
            # Use 'Null' for missing vehicle_code and pitb_code, '0.00h' for others
            if new_col in ['Vehicle Code', 'Pitb Code']:
                result_df[new_col] = 'Null'
            else:
                result_df[new_col] = '0.00h'

    # Fill NaN with appropriate defaults
    for col in result_df.columns:
        if col in ['Vehicle Code', 'Pitb Code']:
            result_df[col] = result_df[col].fillna('Null')
        else:
            result_df[col] = result_df[col].fillna('0.00h')

    # Export to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Vehicle Status Report')

        # Auto-adjust column widths
        worksheet = writer.sheets['Vehicle Status Report']
        for idx, col in enumerate(result_df.columns):
            max_length = max(
                result_df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2

    # Return response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vehicle_status_report.xlsx"'
    return response


# Helper Function

def export_to_excel(duration_data):
    print(f"Received duration data: {duration_data[:2]}")  # Print first two records

    # Convert to DataFrame
    df = pd.DataFrame(duration_data)
    print(f"DataFrame columns: {df.columns.tolist()}")
    print(f"DataFrame shape: {df.shape}")

    # Ensure all expected columns are present and ordered
    columns = ['Vehicle Code', 'Pitb Code', 'Reg No', 'Distance', 'Working', 'Moving', 'Parked', 'Idle', 'Offline',
               'No Response']

    # Create a new DataFrame with the desired columns
    result_df = pd.DataFrame(columns=columns)

    # Map the existing columns to the new DataFrame
    column_mapping = {
        'vehicle_code': 'Vehicle Code',
        'pitb_code': 'Pitb Code',
        'reg_no': 'Reg No',
        'distance': 'Distance',
        'working': 'Working',
        'moving': 'Moving',
        'parked': 'Parked',
        'idle': 'Idle',
        'offline': 'Offline',
        'no_response': 'No Response'
    }

    # Copy data from original DataFrame to new DataFrame
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            # Special handling for vehicle_code and pitb_code
            if old_col in ['vehicle_code', 'pitb_code']:
                # Replace empty strings and NaN with 'Null'
                result_df[new_col] = df[old_col].replace(['', 'nan', 'NaN', 'None', 'none'], 'Null')
                result_df[new_col] = result_df[new_col].fillna('Null')
            else:
                result_df[new_col] = df[old_col]
            print(f"Copied {old_col} to {new_col}")
        else:
            # Use 'Null' for missing vehicle_code and pitb_code, '0.00h' for others
            if new_col in ['Vehicle Code', 'Pitb Code']:
                result_df[new_col] = 'Null'
            else:
                result_df[new_col] = '0.00h'
            print(f"Column {old_col} not found, using default value")

    # Fill NaN with appropriate defaults
    for col in result_df.columns:
        if col in ['Vehicle Code', 'Pitb Code']:
            result_df[col] = result_df[col].fillna('Null')
        else:
            result_df[col] = result_df[col].fillna('0.00h')

    print(f"Final DataFrame shape: {result_df.shape}")
    print(f"Sample data:\n{result_df.head()}")

    # Export to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Vehicle Status Report')

        # Auto-adjust column widths
        worksheet = writer.sheets['Vehicle Status Report']
        for idx, col in enumerate(result_df.columns):
            max_length = max(
                result_df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2

    # Return response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vehicle_status_report.xlsx"'
    return response


def format_duration_hours_minutes(duration):
    """Convert timedelta to HHh.MMm format"""
    total_seconds = duration.total_seconds()
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours:02d}H.{minutes:02d}M"


def format_distance_km(distance_meters):
    """Convert meters to kilometers with 2 decimal places"""
    return f"{distance_meters / 1000:.2f}KM"


# For Ajax Request
def FetchVehicleCodes(request):
    vehicle_type = request.GET.get('vehicle_type')

    # Start with all vehicle codes
    vehicle_codes = VehicleData.objects.all()

    # Apply vehicle type filter if provided
    if vehicle_type:
        vehicle_codes = vehicle_codes.filter(vehicle_type=vehicle_type)

    # Get distinct vehicle codes
    codes = vehicle_codes.values_list('vehicle_code', flat=True).distinct()

    return JsonResponse({
        'vehicle_codes': list(codes),
        'count': len(codes)
    })


def GetByRecord_VehicleStatus_Function(acc_status, speed, ext_bat_voltage):
    if acc_status == "on" and speed > 0 and ext_bat_voltage > 9:
        return "moving"
    elif acc_status == "on" and speed == 0 and ext_bat_voltage > 9:
        return "idle"
    elif acc_status == "off" and speed == 0 and ext_bat_voltage > 9:
        return "parked"
    elif acc_status == "off" and speed == 0 and ext_bat_voltage == 0:
        return "offline"
    elif ext_bat_voltage == -1000:
        return "offline"


def VehicleGeoStatusMovementReportView(request):
    g_status = request.POST.get('g_status')
    vehicle_type = request.POST.get('vehicle_type')
    # vehicle_code = "278903"
    vehicle_code = request.POST.get('vehicle_code')
    end_date = request.POST.get('end_date', localdate())
    if isinstance(end_date, str):
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

    # start_date = request.POST.get('start_date', end_date - timedelta(days=1))
    start_date = request.POST.get('start_date', localdate())
    if isinstance(start_date, str):
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

    export = request.POST.get('export')

    # New filters for working and no-response vehicles
    selected_vehicle_status = request.POST.get('vehicle_status', 'all')

    filters = {}
    if vehicle_type:
        filters['vehicle_code__vehicle_type'] = vehicle_type
    if vehicle_code:
        filters['vehicle_code__vehicle_code'] = vehicle_code
    if start_date and end_date:
        filters['vendor_date_time__date__range'] = (start_date, end_date)

    queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by(
        'vehicle_code__vehicle_code', 'vendor_date_time'
    )

    duration_data = CalculateVehicleGeoStatusDuration_Function(queryset)

    # Apply additional filtering based on working/no-response status
    if selected_vehicle_status == 'working':
        # Filter for vehicles that have any working time
        duration_data = [vehicle for vehicle in duration_data
                         if 'working' in vehicle and float(vehicle['working'].rstrip('h')) > 0]

    elif selected_vehicle_status == 'no_response':
        # Filter for vehicles that have any no-response time
        duration_data = [vehicle for vehicle in duration_data
                         if 'no_response' in vehicle and float(vehicle['no_response'].rstrip('h')) > 0]

    # Check for export parameter
    if export == 'excel':
        return export_to_excel(duration_data)

    # Dropdown options - distinct values
    vehicle_types = VehicleData.objects.values_list('vehicle_type', flat=True).distinct()
    g_statuses = VehicleLiveMonitor.objects.values_list('g_status', flat=True).distinct()

    vehicle_status_options = [
        ('all', 'All Vehicles'),
        ('working', 'Working Vehicles'),
        ('no_response', 'No Response Vehicles'),
    ]

    context = {
        'data': duration_data,  # filtered result
        'vehicle_types': vehicle_types,
        'g_statuses': g_statuses,
        # For preserving form state
        'selected_g_status': g_status,
        'selected_vehicle_type': vehicle_type,
        'selected_vehicle_code': vehicle_code,
        'start_date': start_date,
        'end_date': end_date,
        'vehicle_status_options': vehicle_status_options,
        'selected_vehicle_status': selected_vehicle_status,
    }
    return render(request, "VehicleLiveMonitorReport.html", context)


def SingleVehicleGeoStatusMovementReportView(request, vehicle_code):
    """
    Generate a single-vehicle live monitoring report (HTML or Excel).

    This view:
      - Fetches tracker data for the vehicle (filtered by optional start & end date).
      - Calculates status durations (Working, Moving, Idle, Parked, Offline).
      - Splits data into trips based on >30 min gaps and assigns status IDs for each segment.
      - Builds:
          * status_sub_category → trip-wise segments with status_id, start, end, duration.
          * all_tracker_points → point-wise tracker data (with trip_id & status_id).
      - Supports Excel export (summary + detailed logs).

    Args:
        request (HttpRequest): GET params: start_date, end_date, export ('excel_detail').
        vehicle_code (str): Target vehicle's unique code.

    Returns:
        HttpResponse:
            - Excel file if `export=excel_detail`.
            - Otherwise, renders 'SingleVehicleStatusReport.html' with trip flow and status IDs.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    start_page_date = request.POST.get('start_date')
    end_page_date = request.POST.get('end_date')
    export = request.GET.get('export')

    # Initialize filters
    filters = {'vehicle_code__vehicle_code': vehicle_code}
    if start_date and end_date:
        filters['vendor_date_time__date__range'] = (start_date, end_date)
    if start_page_date and end_page_date:
        filters['vendor_date_time__date__range'] = (start_page_date, end_page_date)

    print(filters)
    queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by('vendor_date_time')

    # if queryset.exists():
    #     null_status_qs = queryset.filter(Q(gis_geo_status__isnull=True) | Q(gis_geo_status=""))
    #     if null_status_qs.exists():
    #         print("Found records with [NULL]")
    #         status_category = CalculateVehicleGeoStatusDuration_Function(queryset)
    #         vehicle_status_duration = status_category[0]

    #         # Refresh queryset after updating gis_geo_status
    #         queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by('vendor_date_time')

    status_category = CalculateVehicleGeoStatusDuration_Function(queryset)
    vehicle_status_duration = status_category[0]

    # Status mapping
    STATUS_ID_MAP = {
        "working": 1,
        "moving": 2,
        "parked": 3,
        "idle": 4,
        "offline": 5,
        "waiting": 6,
    }

    # Tracking variables
    status_sub_category = []
    all_tracker_points = []

    # Count of trips
    total_trips = 0

    total_distance = 0
    total_trip_distance = 0

    # Tracking variables
    consecutive_period_id = 0
    current_gis_geo_status = None
    gis_geo_status_start_time = None
    prev_vendor_date_time = None
    prev_gis_geo_status = None
    prev_geom = None

    # status_category = None
    # vehicle_status_duration = None

    for record in queryset:
        current_vendor_date_time = record.vendor_date_time
        gis_geo_status = (record.gis_geo_status or "unknown").lower()

        ### Processing All Trip Calculation (START)
        if current_gis_geo_status is None:
            # Set First record/Point
            current_gis_geo_status = gis_geo_status
            gis_geo_status_start_time = current_vendor_date_time

            ## Offline Duration at Start of the day (START)
            start_of_day = timezone.make_aware(
                datetime.datetime.combine(current_vendor_date_time.date(), time.min),  # 00:00:00
                timezone.get_current_timezone()
            )
            start_time = start_of_day
            end_time = current_vendor_date_time
            time_diff = end_time - start_time  # 05:02:00 - 00:00:00 = 5 Hours (offline)
            duration_gap = format_duration_hours_minutes(time_diff)

            consecutive_period_id += 1
            total_trips += 1

            status_sub_category.append({
                'trip_id': consecutive_period_id,
                'status_id': vehicle_status_duration.get("offline", {}).get("id", 0),
                # 'status_id': STATUS_ID_MAP.get("offline", 0),
                'status': 'Offline',
                'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                'duration': duration_gap,
                'distance': "0"
            })
            ## Offline Duration at Start of the day (END)

        else:

            gis_geo_status_end_time = current_vendor_date_time

            # -------- OFFLINE GAP CHECK (applies always, not only on status change (START)) --------
            time_diff = current_vendor_date_time - prev_vendor_date_time
            if time_diff.total_seconds() > 60 * 60:  # gap threshold = 1 hour (you can change)
                # Close previous segment up to prev_vendor_date_time
                end_time = prev_vendor_date_time
                duration_gap = format_duration_hours_minutes(end_time - gis_geo_status_start_time)

                consecutive_period_id += 1
                total_trips += 1
                status_sub_category.append({
                    'trip_id': consecutive_period_id,
                    'status_id': vehicle_status_duration.get(current_gis_geo_status.lower(), {}).get("id", 0),
                    'status': current_gis_geo_status,
                    'start_time': gis_geo_status_start_time.strftime("%Y-%m-%d %H:%M"),
                    'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                    'duration': duration_gap,
                    'distance': "0"
                })

                # Insert offline trip
                consecutive_period_id += 1
                total_trips += 1
                offline_duration = format_duration_hours_minutes(time_diff)
                status_sub_category.append({
                    'trip_id': consecutive_period_id,
                    'status_id': STATUS_ID_MAP.get('offline', 0),
                    'status': 'offline',
                    'start_time': prev_vendor_date_time.strftime("%Y-%m-%d %H:%M"),
                    'end_time': current_vendor_date_time.strftime("%Y-%m-%d %H:%M"),
                    'duration': offline_duration,
                    'distance': "0"
                })

                # Reset for new segment (after offline gap)
                current_gis_geo_status = gis_geo_status
                gis_geo_status_start_time = current_vendor_date_time
                total_trip_distance = 0.0
            # -------- OFFLINE GAP CHECK (applies always, not only on status change (END)) --------

            if gis_geo_status != current_gis_geo_status:  # Moving not equal to Moving == True (Mean Gis-Geo-Status Change)
                status_id = vehicle_status_duration.get(gis_geo_status.lower(), {}).get("id", 0)
                start_time = gis_geo_status_start_time
                end_time = gis_geo_status_end_time
                time_diff = end_time - start_time
                duration_gap = format_duration_hours_minutes(time_diff)

                # Calculate distance for the period
                total_distance += total_trip_distance
                distance_km = (float(total_trip_distance) / 1000)
                round_distance_km = "{:.2f}".format(distance_km)

                consecutive_period_id += 1
                total_trips += 1

                if duration_gap != "00H.00M":
                    status_sub_category.append({
                        'trip_id': consecutive_period_id,
                        'status_id': status_id,
                        'status': current_gis_geo_status,
                        'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                        'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                        'duration': duration_gap,
                        'distance': round_distance_km
                    })
                    # total_durations += time_diff
                    total_trip_distance = 0.0  # Reset total_trip_distance for the next trip calculation

                # Start new status
                current_gis_geo_status = gis_geo_status
                gis_geo_status_start_time = current_vendor_date_time

        # Add record to current period with IDs
        all_tracker_points.append({
            'trip_id': consecutive_period_id + 1,
            'status_id': vehicle_status_duration.get(gis_geo_status.lower(), {}).get("id", 0),
            # 'status_id': STATUS_ID_MAP.get(current_gis_geo_status.lower(), 0),
            'tracker_data': {
                "vehicle_code": vehicle_code,
                'pitb_code': record.vehicle_code.pitb_code,
                'register_no': record.vehicle_code.register_no,
                'chasis_no': record.vehicle_code.chasis_no,
                'gis_geo_status': current_gis_geo_status,
                'vendor_date_time': current_vendor_date_time.strftime("%Y-%m-%d %H:%M"),
                'speed': record.speed,
                'mileage': record.mileage,
                'latitude': record.latitude,
                'longitude': record.longitude,
                'device_status': record.device_status,
            }
        })

        prev_vendor_date_time = current_vendor_date_time
        prev_gis_geo_status = gis_geo_status

    else:
        # Handle the last period if it wasn't closed
        if current_gis_geo_status and gis_geo_status_start_time:
            start_time = gis_geo_status_start_time
            end_time = queryset.last().vendor_date_time
            time_diff = end_time - start_time
            duration_gap = format_duration_hours_minutes(time_diff)

            # Calculate distance for the period
            total_distance += total_trip_distance
            distance_km = (float(total_trip_distance) / 1000)
            round_distance_km = "{:.2f}".format(distance_km)

            consecutive_period_id += 1
            total_trips += 1

            status_sub_category.append({
                'trip_id': consecutive_period_id,
                'status_id': vehicle_status_duration.get(gis_geo_status.lower(), {}).get("id", 0),
                # 'status_id': STATUS_ID_MAP.get(current_gis_geo_status.lower(), 0),
                'status': current_gis_geo_status,
                'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                'duration': duration_gap,
                'distance': round_distance_km
            })

        ## No Response Trip End of the day (Offline)
        end_of_day = timezone.make_aware(
            datetime.datetime.combine(prev_vendor_date_time.date(), time.max),  # 23:59:59
            timezone.get_current_timezone()
        )
        start_time = prev_vendor_date_time
        end_time = end_of_day
        time_diff = end_time - start_time  # 23:59:59 - 17:59:59 = 18 Hours (offline)
        duration_gap = format_duration_hours_minutes(time_diff)

        consecutive_period_id += 1
        total_trips += 1
        status_sub_category.append({
            'trip_id': consecutive_period_id,
            'status_id': vehicle_status_duration.get("offline", {}).get("id", 0),
            # 'status_id': STATUS_ID_MAP.get("offline", 0),
            'status': 'Offline',
            'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
            'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
            'duration': duration_gap,
            'distance': "0"
        })
        ### Processing All Trip Calculation (END)

    # Handle Excel export
    if export == 'excel_detail':
        # Create Excel writer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # First table data
            summary_data = []
            for row in status_category:
                summary_data.append({
                    'Vehicle Code': row.get('vehicle_code', 'N/A'),
                    'PITB Code': row.get('pitb_code', 'N/A'),
                    'Reg. No': row.get('register_no', 'N/A'),
                    'Working Hour': row.get('working', {}).get('duration', '0H.0M'),
                    'Moving': row.get('moving', {}).get('duration', '0H.0M'),
                    'Parked': row.get('parked', {}).get('duration', '0H.0M'),
                    'Idle': row.get('idle', {}).get('duration', '0H.0M'),
                    'Offline': row.get('offline', {}).get('duration', '0H.0M') if 'offline' in row else '0H.0M',
                    'No Response': row.get('no_response', {}).get('duration',
                                                                  '0H.0M') if 'no_response' in row else '0H.0M'
                })

            # Second table data
            detail_data = []
            for data in all_tracker_points:
                tracker = data['tracker_data']
                detail_data.append({
                    'Pitb Code': tracker['pitb_code'],
                    'Reg. No': tracker['register_no'],
                    'Status': tracker['g_status'],
                    'Time': tracker['time'],
                    'Speed': tracker['speed'],
                    'Mileage': tracker['mileage'],
                    'Device Status': tracker['device_status'],
                    # 'Working Hours': tracke_data['working_hour']
                })

            # Create DataFrames
            df_summary = pd.DataFrame(summary_data)
            df_detail = pd.DataFrame(detail_data)

            # Write to Excel
            df_summary.to_excel(writer, sheet_name='Vehicle Report', index=False, startrow=0)
            df_detail.to_excel(writer, sheet_name='Vehicle Report', index=False, startrow=len(df_summary) + 3)

            # Auto-adjust column widths
            worksheet = writer.sheets['Vehicle Report']
            for idx, col in enumerate(df_summary.columns):
                max_length = max(
                    df_summary[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2

        # Prepare response
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="vehicle_report_{vehicle_code}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    # Serialize the data for JavaScript
    status_category_json = json.dumps(status_category if status_category else [], cls=DjangoJSONEncoder)

    status_sub_category_json = json.dumps(status_sub_category if status_sub_category else [], cls=DjangoJSONEncoder)

    all_tracker_points_json = json.dumps(all_tracker_points if all_tracker_points else [], cls=DjangoJSONEncoder)

    # print(f"First JSON : {status_category}")
    # print(f"2nd JSON : {status_sub_category}")
    # print(f"3rd JSON : {all_tracker_points}")

    context = {
        'vehicle_duration': status_category,
        'status_sub_category': status_sub_category,
        'all_tracker_points': all_tracker_points,
        'start_date': start_date,
        'end_date': end_date,
        'vehicle_code': vehicle_code,

        'status_category_json': status_category_json,
        'status_sub_category_json': status_sub_category_json,
        'all_tracker_points_json': all_tracker_points_json,
    }
    return render(request, "SingleVehicleMonitorStatusReport.html", context)


# def SingleVehicleGeoStatusMovementReportView(request, vehicle_code):
#     """
#         Generate a single-vehicle live monitoring report (HTML or Excel).
#
#         This view:
#           - Fetches tracker data for the vehicle (filtered by optional start & end date).
#           - Calculates status durations (Working, Moving, Idle, Parked, Offline).
#           - Splits data into trips based on >30 min gaps and assigns status IDs for each segment.
#           - Builds:
#               * status_sub_category → trip-wise segments with status_id, start, end, duration.
#               * all_tracker_points → point-wise tracker data (with trip_id & status_id).
#           - Supports Excel export (summary + detailed logs).
#
#         Args:
#             request (HttpRequest): GET params: start_date, end_date, export ('excel_detail').
#             vehicle_code (str): Target vehicle's unique code.
#
#         Returns:
#             HttpResponse:
#                 - Excel file if `export=excel_detail`.
#                 - Otherwise, renders 'SingleVehicleStatusReport.html' with trip flow and status IDs.
#         """
#     template_name = "Single.html"
#
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     start_page_date = request.POST.get('start_date')
#     end_page_date = request.POST.get('end_date')
#     export = request.GET.get('export')
#
#     # Initialize filters
#     filters = {'vehicle_code__vehicle_code': vehicle_code}
#     if start_date and end_date:
#         filters['vendor_date_time__date__range'] = (start_date, end_date)
#     if start_page_date and end_page_date:
#         filters['vendor_date_time__date__range'] = (start_page_date, end_page_date)
#
#     queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by(
#         'vehicle_code__vehicle_code', 'vendor_date_time'
#     )
#
#     records_list = list(queryset)
#
#     status_category = CalculateVehicleGeoStatusDuration_Function(queryset)
#     vehicle_status_duration = status_category[0]
#
#     def convert_duration_H_M(duration):
#         # Convert string hour to float
#         hour_float = float(duration)
#         # Extract hours and minutes
#         hours = int(hour_float)  # Whole hours
#         minutes = int((hour_float - hours) * 60)
#
#         if minutes >= 60:
#             hours += 1
#             minutes = 0
#
#         # Format as H:M
#         working_hour_format = f"{hours}H:{minutes}M"
#         return working_hour_format
#
#     # Tracking variables
#     consecutive_period_id = 1
#     current_status = None
#     status_start_time = None
#     status_sub_category = []
#     all_tracker_points = []
#     previous_record = None
#
#     for idx, record in enumerate(records_list):
#         # status = record.g_status.lower()
#         current_time = record.vendor_date_time
#         acc_status = (record.acc_status or "unknown").lower()
#         speed = record.speed
#         ext_bat_voltage = record.ext_bat_voltage
#         gis_geo_status = (record.gis_geo_status or "unknown").lower()
#
#         # Determine current status based on conditions
#         if acc_status == "on" and speed > 0 and ext_bat_voltage > 9 and gis_geo_status == 'moving':
#             status = "moving"
#         elif acc_status == "on" and speed == 0 and ext_bat_voltage > 9 and gis_geo_status == 'idle':
#             status = "idle"
#         elif acc_status == "off" and speed == 0 and ext_bat_voltage > 9 and gis_geo_status == 'parked':
#             status = "parked"
#         elif acc_status == "off" and speed == 0 and ext_bat_voltage == 0 and gis_geo_status == 'offline':
#             status = "offline"
#         else:
#             status = "unknown"
#
#         # Detect No Response Gap (Insert as a separate period)
#         if previous_record:
#             gap_seconds = (current_time - previous_record.vendor_date_time).total_seconds()
#             if gap_seconds > 1800:  # >30 min gap
#                 consecutive_period_id += 1
#                 status_sub_category.append({
#                     'trip_id': consecutive_period_id,
#                     'status_id': vehicle_status_duration.get("no_response", {}).get("id", 0),
#                     'status': 'No Response',
#                     'start_time': previous_record.vendor_date_time.strftime("%Y-%m-%d %H:%M"),
#                     'end_time': current_time.strftime("%Y-%m-%d %H:%M"),
#                     'duration': convert_duration_H_M(gap_seconds / 3600)
#                 })
#
#         # Handle Status Change Periods (Moving, Idle, Parked, Offline)
#         if current_status is None:
#             # First record
#             current_status = status
#             status_start_time = current_time
#         else:
#             if status != "unknown" and status != current_status:
#                 status_id = vehicle_status_duration.get(current_status.lower(), {}).get("id", 0)
#                 end_time = previous_record.vendor_date_time
#                 duration_gap = convert_duration_H_M(
#                     (end_time - status_start_time).total_seconds() / 3600
#                 )
#
#                 status_sub_category.append({
#                     'trip_id': consecutive_period_id,
#                     'status_id': status_id,
#                     'status': current_status,
#                     'start_time': status_start_time.strftime("%Y-%m-%d %H:%M"),
#                     'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
#                     'duration': duration_gap,
#                     # 'record_count': len(all_tracker_points),
#                     # 'records': period_records
#                 })
#                 consecutive_period_id += 1
#
#                 # Start new status
#                 current_status = status
#                 status_start_time = current_time
#                 # period_records = []
#
#         status_id = vehicle_status_duration.get(current_status.lower(), {}).get("id", 0)
#
#         # Add record to current period with IDs
#         all_tracker_points.append({
#             'trip_id': consecutive_period_id,
#             'status_id': status_id,
#             'tracker_data': {
#                 "vehicle_code": vehicle_code,
#                 'pitb_code': record.vehicle_code.pitb_code,
#                 'register_no': record.vehicle_code.register_no,
#                 'chasis_no': record.vehicle_code.chasis_no,
#                 'g_status': record.g_status,
#                 'time': record.vendor_date_time.strftime("%Y-%m-%d %H:%M"),
#                 'speed': record.speed,
#                 'mileage': record.mileage,
#                 'latitude': record.latitude,
#                 'longitude': record.longitude,
#                 'device_status': record.device_status,
#             }
#         })
#
#         previous_record = record
#
#     # Add final period if exists
#     if current_status and status_start_time:
#         last_record_time = queryset.last().vendor_date_time
#         status_sub_category.append({
#             'period_id': consecutive_period_id,
#             'status_id': status_id,
#             'status': current_status,
#             'start_time': status_start_time.strftime("%Y-%m-%d %H:%M"),
#             'end_time': previous_record.vendor_date_time.strftime("%Y-%m-%d %H:%M"),
#             'duration': convert_duration_H_M((last_record_time - status_start_time).total_seconds() / 3600),
#             # 'record_count': len(period_records),
#             # 'records': period_records
#         })
#
#     # Handle Excel export
#     if export == 'excel_detail':
#         # Create Excel writer
#         output = BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # First table data
#             summary_data = []
#             for row in status_category:
#                 summary_data.append({
#                     'Vehicle Code': row.get('vehicle_code', 'N/A'),
#                     'PITB Code': row.get('pitb_code', 'N/A'),
#                     'Reg. No': row.get('register_no', 'N/A'),
#                     'Working Hour': row.get('working', {}).get('duration', '0H.0M'),
#                     'Moving': row.get('moving', {}).get('duration', '0H.0M'),
#                     'Parked': row.get('parked', {}).get('duration', '0H.0M'),
#                     'Idle': row.get('idle', {}).get('duration', '0H.0M'),
#                     'Offline': row.get('offline', {}).get('duration', '0H.0M') if 'offline' in row else '0H.0M',
#                     'No Response': row.get('no_response', {}).get('duration',
#                                                                   '0H.0M') if 'no_response' in row else '0H.0M'
#                 })
#
#             # Second table data
#             detail_data = []
#             for data in all_tracker_points:
#                 tracker = data['tracker_data']
#                 detail_data.append({
#                     'Pitb Code': tracker['pitb_code'],
#                     'Reg. No': tracker['register_no'],
#                     'Status': tracker['g_status'],
#                     'Time': tracker['time'],
#                     'Speed': tracker['speed'],
#                     'Mileage': tracker['mileage'],
#                     'Device Status': tracker['device_status'],
#                     # 'Working Hours': tracke_data['working_hour']
#                 })
#
#             # Create DataFrames
#             df_summary = pd.DataFrame(summary_data)
#             df_detail = pd.DataFrame(detail_data)
#
#             # Write to Excel
#             df_summary.to_excel(writer, sheet_name='Vehicle Report', index=False, startrow=0)
#             df_detail.to_excel(writer, sheet_name='Vehicle Report', index=False, startrow=len(df_summary) + 3)
#
#             # Auto-adjust column widths
#             worksheet = writer.sheets['Vehicle Report']
#             for idx, col in enumerate(df_summary.columns):
#                 max_length = max(
#                     df_summary[col].astype(str).apply(len).max(),
#                     len(col)
#                 )
#                 worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
#
#         # Prepare response
#         output.seek(0)
#         response = HttpResponse(
#             output.getvalue(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response[
#             'Content-Disposition'] = f'attachment; filename="vehicle_report_{vehicle_code}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
#         return response
#
#     # print(f"First JSON : {status_category}")
#     # print(f"2nd JSON : {status_sub_category}")
#     # print(f"3rd JSON : {all_tracker_points}")
#
#     context = {
#         'vehicle_duration': status_category,
#         'status_sub_category': status_sub_category,
#         'all_tracker_points': all_tracker_points,
#         'start_date': start_date,
#         'end_date': end_date,
#         'vehicle_code': vehicle_code
#     }
#     return render(request, template_name, context)


def SingleVehicleFleetActivityReportView(request, pitb_code):
    template_name = "Report/SingleVehicleFleetActivity.html"
    cursor = connections['default'].cursor()

    url_pitb_code = pitb_code
    # Filter by vehicle code if provided
    vehicle_obj = VehicleData.objects.filter(pitb_code=url_pitb_code).only('vehicle_code').first()
    # logger.info(f"Selected vehicle code: {vehicle_obj.vehicle_code}")
    get_vehicle_id = vehicle_obj.vehicle_code

    format_str_date = "%Y-%m-%d"
    current_data_time = datetime.datetime.now()
    today_date = current_data_time.date()
    get_selected_date = today_date.strftime(format_str_date)  # e.g., "2025-08-16"

    start_of_day_time_str = '00:00:01'  # '00:00:01'

    current_time_str = datetime.datetime.now().strftime('%H:%M')
    end_of_day_time_str = current_time_str + ':00'  # '23:59:00'

    # ### SYNC TRACKER GPRS VEHICLE DATA BY VENDOR SYSTEM
    # SyncTrackerGPRSVehicleData_By_Vendor_Function(request,
    #                                               get_vehicle_id,
    #                                               get_selected_date,
    #                                               start_of_day_time_str,
    #                                               end_of_day_time_str)

    query_feature = f"""
    WITH veh_live AS (
        SELECT
            ST_X(geom) AS x,
            ST_Y(geom) AS y,
            chasis_no,
            register_no,
            vehicle_code_id,
            COALESCE(pitb_code, 'Waiting') AS pitb_code,
            vehicle_type,
            g_status,
            direction,
            speed,
            COALESCE(device_status, 'NA') AS device_status,
            COALESCE(ignition_status, 'NA') AS ignition_status,
            COALESCE(geo_location, 'NA') AS geo_location,
            vendor_date_time,
            COALESCE(duration, 'NA') AS duration
        FROM
            tbl_vehicle_live_monitor AS vlm
        LEFT OUTER JOIN
            tbl_vehicle_data AS veh
        ON
            vlm.vehicle_code_id = veh.vehicle_code WHERE veh.status = 'Active'
    )
    SELECT
        x,
        y,
        chasis_no,
        register_no,
        veh_live.vehicle_code_id,
        pitb_code,
        vehicle_type,
        g_status,
        direction,
        speed,
        device_status,
        ignition_status,
        geo_location,
        vendor_date_time,
        COALESCE(distance, 0) AS distance,
        COALESCE(working_hours, 0) AS working_hours
    FROM
        veh_live
    LEFT OUTER JOIN (
        SELECT
            vehicle_code_id,
            COALESCE(distance, 0) AS distance,
            COALESCE(working_hours, 0) AS working_hours
        FROM
            tbl_vehicle_schedule_gprs_api
        WHERE
            veh_sch_date::date = '{today_date}'
    ) AS veh_sche
    ON
        veh_live.vehicle_code_id = veh_sche.vehicle_code_id
    WHERE
        1=1
    """

    if get_vehicle_id != "NA":
        query_feature += f" AND (veh_live.vehicle_code_id = '{get_vehicle_id}')"

    query_feature += """
    ORDER BY
        distance DESC,
        vendor_date_time DESC,
        duration;
    """

    cursor.execute(query_feature)
    vehicle_live_lists = DictinctFetchAll(cursor)

    ### COMBINE DATE AND TIME INTO STRING
    datetime_from = get_selected_date + " " + start_of_day_time_str  ## "2025-02-27 00:00:00"
    datetime_to = get_selected_date + " " + end_of_day_time_str  ## "2025-02-27 23:59:59"

    django_datetime_from = parse_datetime(datetime_from)
    django_datetime_to = parse_datetime(datetime_to)

    tracker_raw_gprs_lists = list(
        TrackerRawData.objects.filter(
            vehicle_code_id=get_vehicle_id,
            vendor_date_time__gte=django_datetime_from,
            vendor_date_time__lte=django_datetime_to
        ).annotate(
            x=RawSQL("ST_X(geom)", []),
            y=RawSQL("ST_Y(geom)", [])
        ).values(
            "x", "y", "vehicle_code_id", "g_status", "system_date_time",
            "vendor_date_time", "device_status", "max_speed", "speed",
            "vehicle_status", "direction", "distance", "mileage"
        ).order_by('id')
    )

    tracker_raw_gprs_lists_json = json.dumps(tracker_raw_gprs_lists, cls=DjangoJSONEncoder)

    ### VEHICLE TRACK CALCULATE LENGTH FROM GPRS DATA
    VehicleTracks = TrackerRawData.objects.filter(
        vehicle_code_id=get_vehicle_id,
        vendor_date_time__gte=django_datetime_from,
        vendor_date_time__lte=django_datetime_to
    ).values('vehicle_code_id').annotate(
        line=MakeLine('geom'),  # Create a line from grouped points
        length=Length(MakeLine('geom'))  # Calculate line length
    )

    # Convert data into JSON response
    VehicleTracks_Length = [
        {
            "vehicle_code_id": track["vehicle_code_id"],
            "length_meters": track["length"].m,  # Convert to meters
            "line_geojson": track["line"].geojson  # Convert to GeoJSON
        }
        for track in VehicleTracks
    ]

    page_title = "Vehicle Fleet Activity"
    params = {
        'page_title': page_title,
        'vehicle_live_lists': vehicle_live_lists,
        'tracker_raw_gprs_lists': tracker_raw_gprs_lists,
        'tracker_raw_gprs_lists_json': tracker_raw_gprs_lists_json,
        'vehicle_tracks_length': VehicleTracks_Length,
        # 'current_date': current_date,
        # 'message': message,
        # "chart_data": json.dumps(chart_data),
        # "all_vehicles_today_count": chart_data['current_count'],
        # "vendor_date_groups_count": chart_data['no_response_count'],
        # "total_vehicles_count": chart_data['total_count'],
        # "current_vehicles": current_vehicles['current_vehicles'],
        # "vehicle_type_list": vehicle_type_list,
        # "device_status": device_status,
        # "working_status_options": working_status_options,
        # "no_response_options": no_response_options,
        # "g_statuses": g_statuses,
        # "vehicle_codes": vehicle_codes,
        # "request": request,
        # "start_date": start_dt,
        # "end_date": end_dt,
        # "start_date_report": start_dt_report,
        # # Selected filter values for maintaining state
        # "selected_device_status": device_status_param,
        # "selected_status": vehicle_status,
        # "selected_vehicle_type": vehicle_type,
        "selected_vehicle_code": get_vehicle_id,
    }

    return render(request, template_name, params)


@require_POST
def connect_vehicle_to_route(request):
    """Check PITB code availability and connect vehicle to PITB route"""
    vehicle_code = request.POST.get('vehicle_code')
    pitb_code = request.POST.get('pitb_code')

    if not vehicle_code or not pitb_code:
        return JsonResponse({
            'success': False,
            'message': 'Vehicle code and PITB code are required'
        })

    try:
        # First, check if PITB code already exists
        if VehicleData.objects.filter(pitb_code=pitb_code).exists():
            return JsonResponse({
                'success': False,
                'message': 'This PITB code already exists. Please use a different PITB code.'
            })

        # If PITB code is available, proceed with connection
        vehicle = VehicleData.objects.get(vehicle_code=vehicle_code)
        vehicle.pitb_code = pitb_code
        vehicle.updated_at = localtime()
        vehicle.updated_by = "admin"
        vehicle.save()

        return JsonResponse({
            'success': True,
            'message': f'Vehicle {vehicle_code} successfully connected to route {pitb_code}'
        })

    except VehicleData.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Vehicle not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error connecting vehicle: {str(e)}'
        })


# Add this new function at the end of the file

def create_complaint_form_view(request, complaint_id):
    """Show complaint form and check if complaint already exists for the vehicle"""
    vehicle_code = complaint_id

    vehicle = get_object_or_404(VehicleData, vehicle_code=vehicle_code)

    # Check for existing complaint for this vehicle today
    existing_complaint = Complaint.objects.filter(
        vehicle_code=vehicle,
        created_at__date=datetime.datetime.now().date(),
        status=Complaint.PENDING
    ).order_by('-created_at').first()

    if request.method == 'POST':
        complaint_type = request.POST.get('complaint_type')
        description = request.POST.get('description')

        # Process complaint creation
        complaint = Complaint.objects.create(
            vehicle_code=vehicle_code,
            complaint_type=complaint_type,
            description=description,
            created_by="admin",
        )
        messages.success(request, 'Complaint created successfully')

    return render(request, 'create_complaint_form.html', {
        'vehicle_code': vehicle_code,
        'existing_complaint': existing_complaint,
    })


# Add this new function to view a single complaint

def ViewComplaintDetailView(request, complaint_id):
    """View details of a specific complaint"""

    # Get the complaint by ID
    complaint = get_object_or_404(Complaint, complaint_code=complaint_id)

    # Render the complaint details template
    return render(request, 'view_complaint.html', {
        'complaint': complaint,
    })


def ViewAndUpdateComplaintsView(request):
    template_name = 'Complaint_Manager.html'

    if request.method == 'GET':
        # Get all complaints
        complaints_list = Complaint.objects.all().order_by('-created_at')

        # Apply filters
        status_filter = request.GET.get('status')
        complaint_type_filter = request.GET.get('complaint_type')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        vehicle_code_filter = request.GET.get('vehicle_code')

        if status_filter:
            complaints_list = complaints_list.filter(status=status_filter)

        if complaint_type_filter:
            complaints_list = complaints_list.filter(complaint_type__icontains=complaint_type_filter)

        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                complaints_list = complaints_list.filter(created_at_date_gte=from_date)
            except ValueError:
                pass

        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                complaints_list = complaints_list.filter(created_at_date_lte=to_date)
            except ValueError:
                pass

        if vehicle_code_filter:
            complaints_list = complaints_list.filter(
                Q(vehicle_code_vehicle_code_icontains=vehicle_code_filter) |
                Q(vehicle_code_register_no_icontains=vehicle_code_filter)
            )

        # Get statistics
        total_complaints = complaints_list.count()
        pending_complaints = complaints_list.filter(status='Pending').count()
        completed_complaints = complaints_list.filter(status='Completed').count()

        # Get unique complaint types for filter
        complaint_types = Complaint.objects.values_list('complaint_type', flat=True).distinct().order_by(
            'complaint_type')

        # Get recent complaints for summary
        recent_complaints = Complaint.objects.all().order_by('-created_at')[:5]

        context = {
            'complaints': complaints_list,
            'total_complaints': total_complaints,
            'pending_complaints': pending_complaints,
            'completed_complaints': completed_complaints,
            'complaint_types': complaint_types,
            'recent_complaints': recent_complaints,
            'page_title': 'Complaint Manager'
        }
        return render(request, template_name, context)

    elif request.method == 'POST':
        complaint_id = request.POST.get('complaint_id')
        remarks = request.POST.get('remarks')
        status = request.POST.get('status')
        try:
            updated_complaint = get_object_or_404(Complaint, complaint_code=complaint_id)
            updated_complaint.status = status
            updated_complaint.remarks = remarks
            updated_complaint.updated_by = "admin"
            updated_complaint.updated_at = datetime.datetime.now()
            updated_complaint.save()
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })


### VEHICLE REPORT ####
def VehicleThresholdReportView(request):
    template_name = "Report/VehicleThresholdReport.html"
    format_date = '%Y-%m-%d'
    current_date = datetime.datetime.now().date()

    get_vehicle_type = request.POST.get('cmd_vehicle_type')
    get_selected_date = request.POST.get('selected_date')
    selected_vehicle_code = request.POST.get('cmd_vehicle_list')
    if not get_selected_date:
        get_selected_date = current_date.strftime(format_date)

    ### APPLY FILTER ON QUERY
    filters = {}
    filters['status'] = "Active"
    if get_vehicle_type not in (None, "NA"):
        filters['vehicle_type'] = get_vehicle_type

    # Then we will call UpdateVehicleScheduleGPRSAPIByThreshold function
    veh_sche_gprs_api_json = CalculateVehicleThresholdSelectedDate_Function(get_selected_date)

    ### APPLY FILTER ON QUERY
    vehicle_data_list = list(VehicleData.objects.filter(**filters).order_by('vehicle_code'))
    count = 0
    total_distance = 0.0
    total_working_hour = 0.0
    completed_count = 0
    pending_count = 0
    for vehicle in vehicle_data_list:

        ### FETCH ACTUAL DISTANCE AND WORKING HOURS AND THRESHOLD
        vehicle_code = vehicle.vehicle_code
        grprs_result = veh_sche_gprs_api_json.get(vehicle_code)
        if grprs_result:
            actual_threshold = grprs_result['threshold']
            actual_distance_km = grprs_result['distance']
            actual_working_hours = grprs_result['working_hour']
            actual_distance_m = actual_distance_km * 1000

            threshold_distance = grprs_result['exp_distance'] or 0
            expected_threshold = {
                "distance": grprs_result['exp_distance'] or 0,
                "min_distance": grprs_result['exp_distance_min'] or 0,
                "working_hours": grprs_result['exp_working_hour'] or 0,
                "ignition_status": grprs_result['exp_ignition_status'] or 'No',
            }
            setattr(vehicle_data_list[count], 'expected_threshold', expected_threshold)

            ### COMPLETED VEHICLE THRESHOLD COUNT
            if actual_threshold == "Yes":
                completed_count += 1
            else:
                pending_count += 1

        setattr(vehicle_data_list[count], 'distance_km', actual_distance_km or 0)
        setattr(vehicle_data_list[count], 'distance_m', actual_distance_m or 0)
        setattr(vehicle_data_list[count], 'working_hours', actual_working_hours or 0)
        setattr(vehicle_data_list[count], 'threshold', actual_threshold or 'No')  # Add threshold status
        total_count = completed_count + pending_count
        completed_percent = round((completed_count / total_count) * 100, 2) if total_count else 0
        pending_percent = round((pending_count / total_count) * 100, 2) if total_count else 0

        ### SET PRIORITY ORDER WITH VALUE CONDITION
        if actual_threshold == "Yes":
            priority_order = 3  # Last priority
        elif actual_distance_km <= (actual_distance_m - 2) and actual_threshold != "Yes":
            priority_order = 1  # Top priority
        elif actual_distance_m <= actual_distance_km <= threshold_distance and actual_threshold != "Yes":
            priority_order = 2  # Middle priority
        else:
            priority_order = 4  # Others
        setattr(vehicle_data_list[count], 'priority_order', priority_order)
        ### SET PRIORITY ORDER WITH VALUE CONDITION

        count += 1  # Increment count if condition is met
        total_distance += float(actual_distance_km)  # Increment count if condition is met
        total_working_hour += float(actual_working_hours)  # Increment count if condition is met

    ### SORT BASE ON PRIORITY ORDER
    vehicle_data_list = sorted(vehicle_data_list, key=lambda x: x.priority_order)

    ### VEHICLE DATA DETAIL WITH VEHICLE STATUS AND VEHICLE TYPE
    cmd_vehicle_data = list(
        # VehicleData.objects.filter(**filters_vehicle_data)
        VehicleData.objects.values('vehicle_code', 'vehicle_type', 'register_no', 'chasis_no')
        .annotate(
            pitb_code=Coalesce('pitb_code', Value('Waiting'), output_field=CharField())
        )
        .order_by('register_no')
        .distinct()
    )

    vehicle_type_list = VehicleData.objects.filter(status="Active").values('vehicle_type').annotate(
        count=Count('id')).order_by('vehicle_type')
    # New getting count of those vehicles that complated the threshold by vehicle type
    vehicle_threshold_dict = {}
    logger.info(
        f"First 2 vehicles: {[(v.vehicle_code, v.vehicle_type, v.distance_km, v.working_hours) for v in vehicle_data_list[:9]]}")
    logger.info(f"Vehicle types:{vehicle_type_list}")
    for v_type in vehicle_type_list:
        # Fix: Access dictionary value with key instead of attribute
        vehicle_type_name = v_type['vehicle_type']
        vehicle_threshold_dict[vehicle_type_name] = {"Yes": 0, "No": 0}
        for data in vehicle_data_list:
            if data.vehicle_type == vehicle_type_name:
                if data.threshold == "Yes":
                    vehicle_threshold_dict[vehicle_type_name]["Yes"] += 1
                else:
                    vehicle_threshold_dict[vehicle_type_name]["No"] += 1
    logger.info(f"Vehicle threshold count below:\n")
    logger.info(f"Vehicle threshold dict: {vehicle_threshold_dict}")
    # ------------------------------------

    ### STACK BAR GRAPH TWO DIFFERENT VALUE COMPARISION
    vehicle_type_name_hierarchy = list(vehicle_threshold_dict.keys())

    # 1. Extract the ordered list of type‑names:
    type_names = [vt['vehicle_type'] for vt in vehicle_type_list]

    # 2. Build the two series (Yes / No):
    threshold_yes_data = [vehicle_threshold_dict[t]['Yes'] for t in type_names]
    threshold_no_data = [vehicle_threshold_dict[t]['No'] for t in type_names]

    # 3. Assemble into the TownWiseDetail list:
    VehicleThresholdStatusGraph = [
        {
            'name': 'Completed',
            'data': threshold_yes_data,
            'color': '#96CC39'
        },
        {
            'name': 'Pending',
            'data': threshold_no_data,
            'color': 'orange'
        }
    ]

    page_title = "Create Network"

    params = {
        'vehicle_type_list': vehicle_type_list,
        'vehicle_data_list': vehicle_data_list,
        'total_distance': total_distance,
        'total_working_hour': total_working_hour,
        "get_vehicle_type": get_vehicle_type,
        "str_selected_date": get_selected_date,
        "vehicle_type_name_hierarchy": vehicle_type_name_hierarchy,
        "VehicleThresholdStatusGraph": VehicleThresholdStatusGraph,
        'page_title': page_title,
        'vehicle_threshold_dict': json.dumps(vehicle_threshold_dict),  # Convert to JSON
        'completed_count': completed_count,
        'pending_count': pending_count,
        'completed_percent': completed_percent,
        'pending_percent': pending_percent,
        'cmd_vehicle_data': cmd_vehicle_data,
        'selected_vehicle_code': selected_vehicle_code,
    }

    return render(request, template_name, params)


def CalculateVehicleThresholdSelectedDate_Function(selected_date: str):
    format_date = '%Y-%m-%d'
    current_date = datetime.datetime.now().date()
    str_current_date = current_date.strftime(format_date)

    vehicle_json_dict = {}  # This will store JSON by vehicle code

    ### Get all vehicle schedule GPS data for the given date
    vehicle_schedule_qs = VehicleScheduleGPRSApi.objects.filter(veh_sch_date=selected_date).order_by('vehicle_code_id')
    for vehicle in vehicle_schedule_qs:

        vsgs_id = vehicle.id
        set_vehicle_code = vehicle.vehicle_code_id
        set_vehicle_type = vehicle.vehicle_code.vehicle_type
        set_threshold = vehicle.threshold
        vsgs_distance = vehicle.distance or 0
        vsgs_working_hour = vehicle.working_hours or 0

        ### FETCH VEHICLE TYPE THRESHOLD RECORD BY VEHICLE TYPE
        vehicle_threshold_record = VehicleThreshold.objects.filter(vehicle_type=set_vehicle_type).first()
        if vehicle_threshold_record:
            vtr_distance = vehicle_threshold_record.distance or 0
            vtr_distance_min = vehicle_threshold_record.min_distance or 0
            vtr_working_hour = vehicle_threshold_record.working_hours or 0
            vtr_ignition_status = vehicle_threshold_record.ignition_status or 'No'
        else:
            vtr_distance = 0
            vtr_distance_min = 0
            vtr_working_hour = 0
            vtr_ignition_status = 'No'

        if set_threshold == "No" or set_threshold is None:
            ## WHICH VEHICLE HAVING DISTANCE AND WORKING HOUR
            if vtr_ignition_status == "No" and (float(vsgs_distance) > vtr_distance) and (
                    float(vsgs_working_hour) >= vtr_working_hour):
                set_threshold = "Yes"

            ## VEHICLE ONLY WORKING HOUR
            if vtr_ignition_status == "Yes" and (float(vsgs_working_hour) > vtr_working_hour):
                set_threshold = "Yes"

            if set_threshold == "Yes":
                # logger.info(f"Vehicle Threshold Status Change {vsgs_working_hour} on date {date}")
                VehicleScheduleGPRSApi.objects.filter(id=vsgs_id).update(threshold="Yes")

        ### CREATE VEHICLE GPRS OBJECT
        vehicle_json_dict[set_vehicle_code] = {
            'threshold': set_threshold or 'No',
            'distance': vsgs_distance,
            'working_hour': vsgs_working_hour,
            'exp_distance': vtr_distance,
            'exp_distance_min': vtr_distance_min,
            'exp_working_hour': vtr_working_hour,
            'exp_ignition_status': vtr_ignition_status,
        }

    return vehicle_json_dict


def VehicleThresholdSingle(filters, tracker_filters):
    vehicle_data_list = list(VehicleData.objects.filter(**filters).order_by('vehicle_type'))
    count = 0
    for vehicle in vehicle_data_list:
        see_vehicle_code = vehicle.vehicle_code
        tracker_filters['vehicle_code_id'] = see_vehicle_code

        # Get all records for this vehicle on this date for duration calculation
        vehicle_tracker_records = TrackerRawData.objects.filter(**tracker_filters).order_by('vendor_date_time')

        distance_km = DistanceFinder(vehicle_tracker_records)
        distance_m = float(distance_km) * 1000  # Convert km to m

        ### FUNCTION USED FOR CALCULATE SINGLE VEHICLE WORKING HOURS
        working_hours_only = CalculateSingleVehicleWorkingHour_Function(vehicle_tracker_records)

        total_ignition_off = ""
        if working_hours_only > 0:
            total_ignition_off = "Yes"

        setattr(vehicle_data_list[count], 'distance_km', distance_km)
        setattr(vehicle_data_list[count], 'distance_m', distance_m)
        setattr(vehicle_data_list[count], 'ignition', total_ignition_off)
        setattr(vehicle_data_list[count], 'working_hours', working_hours_only)

        count += 1  # Increment count if condition is met

    return vehicle_data_list


def VehicleThresholdGroup(single_records):
    """
    Group pre‐enriched single_records by (vehicle_type, date),
    summing distance and working hours.
    """
    grouped = defaultdict(lambda: {
        'vehicle_type': None,
        'date': None,
        'total_distance_km': 0.0,
        'total_distance_m': 0.0,
        'total_working_hours': 0.0,
        'vehicle_count': 0
    })

    for rec in single_records:
        vtype = rec.vehicle_type
        date = rec.vendor_date_time.date()
        key = (vtype, date)
        grp = grouped[key]
        grp['vehicle_type'] = vtype
        grp['date'] = date
        grp['total_distance_km'] += rec.distance_km
        grp['total_distance_m'] += rec.distance_m
        grp['total_working_hours'] += rec.working_hours
        grp['vehicle_count'] += 1

    return list(grouped.values())


def DistanceFinder(vehicle_tracker_records):
    """
    Calculate total distance (in km) for a vehicle on a given date.
    Skips records without geometry.
    """
    total_distance = 0.0
    previous_point = None

    for row in vehicle_tracker_records:
        point = GEOSGeometry(row.geom)
        if previous_point:
            # geodesic needs (lat, lon)
            dist = geodesic(
                (previous_point.y, previous_point.x),
                (point.y, point.x)
            ).meters
            total_distance += dist
        previous_point = point
    # print(f"Total distance (meters): {total_distance}")

    distance_km = (float(total_distance) / 1000)
    round_distance_km = "{:.2f}".format(distance_km)

    return round_distance_km


def working_idle_buffer_status_func(
        v_code,
        idle_buffer_between_moving_sum,
        vehicle_status_durations,
):
    """
    CALCULATE TIME DURATION START MOVING AND END MOVING
    1 - LESS THAN 3 MINUTES MOVING
    2 - GREATE THAN 3 AND LESS THAN 31 MIN = WAITING
    3 - GREATE THAN 31 AND LESS THAN 60:01 MIN = IDLE
    4 - GREATE THAN 1 HOURS = PARKED
    """

    ### 1 - LESS THAN 3 MINUTES MOVING
    if idle_buffer_between_moving_sum < 3 * 60:
        updated_status = 'Moving'
        vehicle_status_durations[v_code]['working'] += timedelta(seconds=idle_buffer_between_moving_sum)
        vehicle_status_durations[v_code]['moving'] += timedelta(seconds=idle_buffer_between_moving_sum)

    ### 2 - GREATE THAN 3 AND LESS THAN 31 MIN = WAITING
    elif 3 * 60 < idle_buffer_between_moving_sum < 31 * 60:
        updated_status = 'Waiting'
        vehicle_status_durations[v_code]['waiting'] += timedelta(seconds=idle_buffer_between_moving_sum)

    ### 3 - GREATE THAN 31 AND LESS THAN 60:01 MIN = IDLE
    elif 31 * 60 < idle_buffer_between_moving_sum < 61 * 60:
        updated_status = 'Idle'
        vehicle_status_durations[v_code]['idle'] += timedelta(seconds=idle_buffer_between_moving_sum)

    ### 4 - GREATE THAN 1 HOURS = PARKED
    elif idle_buffer_between_moving_sum > 61 * 60:  ## change
        updated_status = 'Parked'
        vehicle_status_durations[v_code]['parked'] += timedelta(seconds=idle_buffer_between_moving_sum)

    return updated_status


# Update distance , Threshold Status, gis_geo_status, working_hour
from django.db.models import Case, When, Value, FloatField


def CalculateVehicleGeoStatusDurationThreshold_Function(queryset):
    """
    Process raw vehicle tracking data to determine geo-status (Moving, Idle, Waiting, Parked, Offline),
    calculate cumulative working hours and distance, and apply threshold rules.

    Args:
        queryset (QuerySet): Ordered queryset of TrackerRawData records for a vehicle.

    Workflow:
        - Iterates through each tracking record.
        - Determines current geo-status based on ACC, speed, and battery voltage.
        - Tracks transitions between statuses (Moving → Idle, Idle → Moving, Moving → Parked, etc.).
        - Accumulates working hours (engine ON & moving/idle) and travelled distance (geo-coordinates).
        - Groups consecutive idle records and later bulk-updates them with the correct status
          (Moving/Waiting/Idle/Parked) depending on idle duration.
        - Applies business thresholds for minimum distance, working hours, and ignition rules,
          and updates the `threshold_status` flag accordingly.

    Returns:
        bool: Always True on successful processing and database updates.
    """

    # --- Initialization of state variables ---
    vehicle_type = None
    previous_status = None  # Holds the last determined status
    previous_geom = None  # Holds last GPS coordinate
    previous_vendor_date_time = None  # Holds last timestamp
    previous_id = None  # Holds last record ID

    # Flags used for handling transitions
    last_moving_time = False
    next_moving_time = False
    idle_buffer_between_moving_sum = 0  # Accumulated idle duration (seconds)

    # Temporary storage for bulk updates
    idle_status_ids_obj = []  # Collects record IDs during idle
    working_hours_obj = []  # Collects cumulative working hours aligned to record IDs
    distance_obj = []  # Collects cumulative distances aligned to record IDs

    # Cumulative totals
    total_distance = 0.0
    total_working_hour = 0.0

    # --- Main loop: process each tracking record sequentially ---
    for record in queryset:
        current_id = record.id
        vehicle_code = record.vehicle_code
        vehicle_type = record.vehicle_code.vehicle_type
        acc_status = record.acc_status.lower() if record.acc_status else None  ## Update and below
        speed = record.speed if record.speed is not None else None
        ext_bat_voltage = record.ext_bat_voltage if record.ext_bat_voltage is not None else None
        current_vendor_date_time = localtime(record.vendor_date_time)  ## remove local_time
        current_geom = record.geom

        # --- Determine geo-status based on ACC, speed, and battery ---
        if acc_status == "on" and speed > 0 and ext_bat_voltage > 0:
            status = "moving"
        elif acc_status == "on" and speed == 0 and ext_bat_voltage > 0:
            status = "idle"
        elif acc_status == "off" and speed == 0 and (ext_bat_voltage > 0 or ext_bat_voltage < 0):
            status = "parked"
        elif acc_status == "off" and speed == 0 and ext_bat_voltage == 0:
            status = "offline"
        elif acc_status is None or speed is None or ext_bat_voltage is None:
            # If data is missing, fallback to previous status
            status = previous_status if previous_status else None

        # Collect idle records ids for later classification
        if status == 'idle':
            idle_status_ids_obj.append(current_id)

        # Start of the idles are converted to Parked (ONLY FIRST/ONE TIME)
        if status in ['parked', 'moving'] and idle_status_ids_obj and last_moving_time == False:
            TrackerRawData.objects.filter(
                id__in=idle_status_ids_obj
            ).update(gis_geo_status="Parked", working_hours=total_working_hour, distance=total_distance)

            idle_status_ids_obj.clear()

        # Mark first "Moving" event after idle
        if status == 'moving' and next_moving_time == False:
            last_moving_time = True
            TrackerRawData.objects.filter(id=current_id).update(gis_geo_status='Moving',
                                                                working_hours=total_working_hour,
                                                                distance=total_distance)

        # Handle transition: Moving → Parked
        if status == 'parked' and previous_status == 'moving' and len(idle_status_ids_obj) == 0:
            last_moving_time = False
            TrackerRawData.objects.filter(
                id=current_id).update(gis_geo_status='Parked', working_hours=total_working_hour,
                                      distance=total_distance)

            # Handle transition: Idle -> Moving/Parked
            if previous_status == 'idle' and status in ['moving', 'parked'] and last_moving_time == False:
                if idle_status_ids_obj:
                    TrackerRawData.objects.filter(
                        id__in=idle_status_ids_obj).update(gis_geo_status='Idle', working_hours=total_working_hour,
                                                           distance=total_distance)
                    idle_status_ids_obj.clear()

                if status == 'parked':
                    TrackerRawData.objects.filter(
                        id=current_id).update(gis_geo_status='Parked', working_hours=total_working_hour,
                                              distance=total_distance)

                # --- Duration and distance and geo status calculations only after first Moving is detected ---
        if last_moving_time and previous_vendor_date_time and previous_status:

            # Calculate elapsed time between current and previous records
            time_diff = current_vendor_date_time - previous_vendor_date_time
            time_diff_seconds = time_diff.total_seconds()

            # Convert to hours.minutes (e.g., 1.05 = 1 hr 5 mins)
            # working_duration_hours = int(time_diff_seconds // 3600)
            # working_duration_minutes = int((time_diff_seconds % 3600) // 60)
            # working_hours_minutes = float(f"{working_duration_hours}.{working_duration_minutes:02d}")

            # Calculate elapsed time between current and previous records
            time_diff = current_vendor_date_time - previous_vendor_date_time
            time_diff_seconds = time_diff.total_seconds()

            # Convert to hours.minutes (e.g., 1.05 = 1 hr 5 mins)
            working_duration_hours = time_diff_seconds / 3600
            working_hours_minutes = float("{:.2f}".format(working_duration_hours))

            # --- Distance calculation between consecutive GPS points (Start) ---
            if current_geom and previous_geom:
                try:
                    prev_point = previous_geom
                    current_point = current_geom
                    if prev_point and current_point:
                        distance = geodesic(
                            (prev_point.y, prev_point.x),
                            (current_point.y, current_point.x)
                        ).meters
                    distance_km = (float(distance) / 1000) if distance > 0 else 0.0
                    round_distance_km = float("{:.1f}".format(distance_km))

                except Exception as e:
                    print(f"Error calculating distance for vehicle {vehicle_code}: {e}")
            # --- Distance calculation between consecutive GPS points (Start) ---

            # Handle continous transition: Moving → Moving -> Moving
            if previous_status == 'moving' and status == 'moving':
                total_working_hour += working_hours_minutes
                total_distance += round_distance_km
                # Get hours
                hours = int(total_working_hour)
                # Convert fractional part to minutes
                minutes = int(round((total_working_hour - hours) * 60))
                ft_working_hours_minutes = Decimal(f"{hours}.{minutes:02d}")
                TrackerRawData.objects.filter(id=current_id).update(gis_geo_status='Moving',
                                                                    working_hours=ft_working_hours_minutes,
                                                                    distance=total_distance)

            # Track idle buffer + accumulate totals temporarily
            if status == 'idle':
                idle_buffer_between_moving_sum += time_diff_seconds
                total_working_hour += working_hours_minutes
                # Get hours
                hours = int(total_working_hour)
                # Convert fractional part to minutes
                minutes = int(round((total_working_hour - hours) * 60))
                ft_working_hours_minutes = Decimal(f"{hours}.{minutes:02d}")
                working_hours_obj.append(ft_working_hours_minutes)
                distance_obj.append(total_distance)

            # Transition from idle back to moving (Mean Next Moving -> Flagged)
            if status == 'moving' and previous_status == 'idle' and len(idle_status_ids_obj) > 0:
                idle_buffer_between_moving_sum += time_diff_seconds
                next_moving_time = True

            # --- If idle group ends, decide correct status for collected records ---
            if next_moving_time or (time_diff_seconds > 60 * 60 and idle_status_ids_obj):

                if idle_buffer_between_moving_sum < 3 * 60:
                    updated_status = 'Moving'

                elif 3 * 60 < idle_buffer_between_moving_sum < 31 * 60:
                    updated_status = 'Waiting'

                elif 31 * 60 < idle_buffer_between_moving_sum < 61 * 60:
                    updated_status = 'Idle'

                elif idle_buffer_between_moving_sum > 61 * 60:
                    updated_status = 'Parked'

                else:
                    updated_status = "Unknown"

                if updated_status in ['Moving', 'Waiting']:
                    # Bulk update with Case/When for accurate per-record WH & distance
                    tids = idle_status_ids_obj
                    whs = working_hours_obj
                    dists = distance_obj

                    TrackerRawData.objects.filter(id__in=tids).update(
                        gis_geo_status=updated_status,
                        working_hours=Case(
                            *[When(id=tid, then=Value(wh)) for tid, wh in zip(tids, whs)],
                            output_field=FloatField()
                        ),
                        distance=Case(
                            *[When(id=tid, then=Value(dist)) for tid, dist in zip(tids, dists)],
                            output_field=FloatField()
                        ),
                    )

                else:
                    # Single bulk update with same WH & distance
                    TrackerRawData.objects.filter(
                        id__in=idle_status_ids_obj
                    ).update(gis_geo_status=updated_status, working_hours=total_working_hour, distance=total_distance)

                # Reset group variables
                idle_status_ids_obj.clear()
                last_moving_time = False if time_diff_seconds > 60 * 60 else True
                next_moving_time = False
                idle_buffer_between_moving_sum = 0
                working_hours_obj.clear()
                distance_obj.clear()

            # --- If idle persists into parked state --- (Moving -> Idle -> Idle -> .... -> Parked)
            if status == 'parked' and idle_status_ids_obj:
                idle_status_ids_obj.append(current_id)
                TrackerRawData.objects.filter(
                    id__in=idle_status_ids_obj
                ).update(gis_geo_status="Parked", working_hours=total_working_hour, distance=total_distance)

                # Reset everything
                last_moving_time = False
                next_moving_time = False
                idle_buffer_between_moving_sum = 0
                idle_status_ids_obj.clear()
                working_hours_obj.clear()
                distance_obj.clear()

        # --- Update reference for next iteration ---
        previous_id = current_id
        previous_status = status
        previous_geom = current_geom
        previous_vendor_date_time = current_vendor_date_time

    # --- End of loop: Final check for leftover idle records ---
    else:
        if idle_status_ids_obj:
            TrackerRawData.objects.filter(
                id__in=idle_status_ids_obj
            ).update(gis_geo_status="Parked", working_hours=total_working_hour, distance=total_distance)

            idle_status_ids_obj.clear()

    # --- Apply Vehicle Threshold rules ---
    if vehicle_type:
        vehicle_threshold_record = VehicleThreshold.objects.filter(
            vehicle_type=vehicle_type
        ).first()

        if vehicle_threshold_record:
            vtr_min_distance = vehicle_threshold_record.min_distance or 0
            vtr_working_hour = vehicle_threshold_record.working_hours or 0
            vtr_ignition_status = vehicle_threshold_record.ignition_status or 'No'

            # Distance + working hour condition
            if vtr_ignition_status == "No" and (float(total_distance) > vtr_min_distance) and (
                    float(total_working_hour) >= vtr_working_hour):
                set_threshold = "Yes"

                # Update threshold in DB
                updated_count = TrackerRawData.objects.filter(
                    vehicle_code=vehicle_code,
                    working_hours__gt=vtr_working_hour,
                    distance__gt=vtr_min_distance
                ).update(threshold_status=set_threshold)

                print(f"Updated rows (Distance + Working Hour): {updated_count}")

            # Only working hour condition
            if vtr_ignition_status == "Yes" and (float(total_working_hour) > vtr_working_hour):
                set_threshold = "Yes"

                # Update threshold in DB
                updated_count = TrackerRawData.objects.filter(
                    vehicle_code=vehicle_code,
                    working_hours__gt=vtr_working_hour
                ).update(threshold_status=set_threshold)

                print(f"Updated rows (Working Hour only): {updated_count}")

    return True


def test_view(request):
    vehicle_code = '279115'
    vendor_date_time = '2025-09-26'

    filters = {}
    if vehicle_code:
        filters['vehicle_code__vehicle_code'] = vehicle_code
    if vendor_date_time:
        filters['vendor_date_time__date__range'] = (vendor_date_time, vendor_date_time)

    queryset = TrackerRawData.objects.filter(**filters).order_by("vendor_date_time")

    print(len(queryset))


    CalculateVehicleGeoStatusDurationThreshold_Function(queryset)

    return HttpResponse("Done")

def CalculateVehicleGeoStatusDuration_Function(queryset):
    """
    Calculate duration of different statuses for vehicles, incorporating the working hour calculation function.

    Args:
        queryset: QuerySet of TrackerRawData records ordered by vehicle and time
        status_filter: Optional status to filter by (only calculate for this status)

    Returns:
        List of dictionaries with vehicle info and status durations in HHh.MMm format
    """

    # Status mapping
    STATUS_ID_MAP = {
        "working": 1,
        "moving": 2,
        "parked": 3,
        "idle": 4,
        "offline": 5,
        "waiting": 6,
    }

    # Initialize data structures
    vehicle_status_durations = defaultdict(lambda: defaultdict(timedelta))
    previous_records = {}
    distance_accumulator = defaultdict(float)
    vehicle_status_counts = defaultdict(lambda: defaultdict(int))

    # Working hour specific variables per vehicle
    vehicle_working_data = defaultdict(lambda: {
        'last_moving_time': False,
        'next_moving_time': False,
        'last_moving_vendor_date_time': None,
        'next_moving_vendor_date_time': None,
        'idle_buffer_between_moving': [],
        'working_seconds': 0,
        'prev_vendor_date_time': None,
        'prev_status': None
    })

    idle_status_obj = defaultdict(lambda: {
        'tracking_ids': [],
        'total_seconds': 0,
    })

    gis_geo_status_logic_obj = defaultdict(lambda: {
        'tracking_ids': [],
        'total_seconds': 0,
        'status': None

    })

    for i, current_record in enumerate(queryset):
        vehicle = current_record.vehicle_code
        v_code = vehicle.vehicle_code
        acc_status = current_record.acc_status.lower() if current_record.acc_status else None
        speed = current_record.speed if current_record.speed is not None else None
        ext_bat_voltage = current_record.ext_bat_voltage if current_record.ext_bat_voltage is not None else None
        current_vendor_date_time = current_record.vendor_date_time
        distance = current_record.distance
        tracking_id = current_record.id
        working_data = vehicle_working_data[v_code]

        # Determine current status based on conditions
        if acc_status == "on" and speed > 0 and ext_bat_voltage > 0:
            status = "moving"
        elif acc_status == "on" and speed == 0 and ext_bat_voltage > 0:
            status = "idle"
        elif acc_status == "off" and speed == 0 and (ext_bat_voltage > 0 or ext_bat_voltage < 0):
            status = "parked"
        elif acc_status == "off" and speed == 0 and ext_bat_voltage == 0:
            status = "offline"
        elif acc_status is None or speed is None or ext_bat_voltage is None:
            status = previous_records[v_code]['status'] if previous_records[v_code] else None

        if v_code in previous_records:
            prev_record = previous_records[v_code]
            time_diff = current_vendor_date_time - prev_record['vendor_date_time']
            diff_vendor_time_seconds = time_diff.total_seconds()
            prev_status = prev_record['status']

            if status == 'idle':
                idle_status_obj[v_code]['tracking_ids'].append(tracking_id)
                idle_status_obj[v_code]['total_seconds'] += diff_vendor_time_seconds

            ### NO RESPONSE VEHICLE TIME DURATION GREATER THAN 1 HOUR
            if time_diff.total_seconds() > 60 * 60:  # 60 Minutes
                vehicle_status_durations[v_code]['offline'] += time_diff

                ## Remove the previous tracking id & total seconds from idle list if gap between points is > 60 to maintain teh gis_geo_status offline , otherwise all idles are convert into PARKED  with gap of 1 hour
                if prev_record['id'] in idle_status_obj[v_code]['tracking_ids']:
                    idle_status_obj[v_code]['tracking_ids'].remove(prev_record['id'])
                    idle_status_obj[v_code]['total_seconds'] -= diff_vendor_time_seconds

                TrackerRawData.objects.filter(
                    id=tracking_id
                ).update(gis_geo_status=status.capitalize())

                if working_data['last_moving_time']:
                    idle_buffer_between_moving_sum = sum(working_data['idle_buffer_between_moving'])

                    ### CALCULATE TIME DURATION START MOVING AND END MOVING
                    updated_status = working_idle_buffer_status_func(
                        v_code,
                        idle_buffer_between_moving_sum,
                        vehicle_status_durations,
                    )
                    for v_code, changes in gis_geo_status_logic_obj.items():
                        TrackerRawData.objects.filter(
                            id__in=changes["tracking_ids"]
                        ).update(gis_geo_status=updated_status)

                    vehicle_status_counts[v_code][updated_status.lower()] += 1

                    working_data['last_moving_time'] = False
                    working_data['next_moving_time'] = False
                    working_data['idle_buffer_between_moving'].clear()
                    gis_geo_status_logic_obj[v_code]['tracking_ids'].clear()

                else:
                    if idle_status_obj[v_code]['total_seconds'] > 0 and not working_data['last_moving_time']:
                        vehicle_status_durations[v_code]['parked'] += timedelta(
                            seconds=idle_status_obj[v_code]['total_seconds'])

                        for v_code, changes in idle_status_obj.items():
                            TrackerRawData.objects.filter(
                                id__in=changes["tracking_ids"]
                            ).update(gis_geo_status="Parked")

                        vehicle_status_counts[v_code]['parked'] += 1

                        idle_status_obj[v_code]['tracking_ids'].clear()
                        idle_status_obj[v_code]['total_seconds'] = 0

                # Store current record for next iteration
                previous_records[v_code] = {
                    'id': tracking_id,
                    'vendor_date_time': current_vendor_date_time,
                    'status': status,
                    'distance': distance
                }
                continue

            # ------------- WORKING AND MOVING DURATION CALCULATION (START)----------  C
            if working_data['last_moving_time']:
                ## Moving -> Moving
                if prev_status == 'moving' and status == 'moving':
                    working_time_diff = current_vendor_date_time - working_data['prev_vendor_date_time']
                    vehicle_status_durations[v_code]['working'] += working_time_diff
                    vehicle_status_durations[v_code]['moving'] += working_time_diff
                    working_data['last_moving_vendor_date_time'] = current_vendor_date_time
                    # gis_geo_status_logic_obj[v_code]['tracking_ids'].append(tracking_id)

                    # Update both current and previous records as "Moving"
                    TrackerRawData.objects.filter(
                        id__in=[prev_record['id'], tracking_id]
                    ).update(gis_geo_status="Moving")

                    if prev_status != status:
                        vehicle_status_counts[v_code]['working'] += 1
                        vehicle_status_counts[v_code]['moving'] += 1

                ## Moving -> (Idle , Parked, offline)
                ### PREV STATUS MOVING AND NEXT STATUS CHANGE THEN CALCULATE GAP TIME DURATION
                ### MOVING = 07:00:01 - (IDLE) 07:01:23 = 1:22 MINUTES (MOVING < 3 MINUTES ) CONSIDER MOVING
                if prev_status == 'moving' and status in ['idle', 'parked'] and len(
                        working_data["idle_buffer_between_moving"]) == 0:
                    if status == 'idle':
                        gis_geo_status_logic_obj[v_code]['tracking_ids'].append(tracking_id)
                        gis_geo_status_logic_obj[v_code]['tracking_ids'].append(prev_record['id'])

                    working_time_diff = (
                            current_vendor_date_time - working_data['prev_vendor_date_time']).total_seconds()
                    if working_time_diff < 3 * 60:
                        vehicle_status_durations[v_code]['working'] += timedelta(seconds=working_time_diff)
                        vehicle_status_durations[v_code]['moving'] += timedelta(seconds=working_time_diff)


                ## Moving (Start) .....(idles) -> idle ...
                elif status == 'idle' and working_data['last_moving_time']:
                    working_time_diff = (
                            current_vendor_date_time - working_data['prev_vendor_date_time']).total_seconds()
                    working_data['idle_buffer_between_moving'].append(working_time_diff)
                    gis_geo_status_logic_obj[v_code]['tracking_ids'].append(tracking_id)

                ## Moving (Start) .....(idles) -> idle ... Moving(End)
                ### NOW MOVING START AGAIN
                if status == 'moving' and prev_status == 'idle' and working_data['last_moving_time']:
                    working_time_diff = (
                            current_vendor_date_time - working_data['prev_vendor_date_time']).total_seconds()
                    working_data['idle_buffer_between_moving'].append(working_time_diff)
                    gis_geo_status_logic_obj[v_code]['tracking_ids'].append(tracking_id)
                    working_data['next_moving_time'] = True

                if working_data['next_moving_time']:
                    idle_buffer_between_moving_sum = sum(working_data['idle_buffer_between_moving'])

                    ### CALCULATE TIME DURATION START MOVING AND END MOVING
                    updated_status = working_idle_buffer_status_func(
                        v_code,
                        idle_buffer_between_moving_sum,
                        vehicle_status_durations,
                    )
                    for v_code, changes in gis_geo_status_logic_obj.items():
                        TrackerRawData.objects.filter(
                            id__in=changes["tracking_ids"]
                        ).update(gis_geo_status=updated_status)

                    vehicle_status_counts[v_code][updated_status.lower()] += 1

                    ### ALL VARIABLE AND OBJECT RESET
                    working_data['idle_buffer_between_moving'].clear()
                    gis_geo_status_logic_obj[v_code]['tracking_ids'].clear()
                    working_data['last_moving_time'] = True
                    working_data['next_moving_time'] = False
                    idle_status_obj[v_code]['tracking_ids'].clear()
                    idle_status_obj[v_code]['total_seconds'] = 0


                elif status == 'parked':
                    working_time_diff = (
                            current_vendor_date_time - working_data['prev_vendor_date_time']).total_seconds()
                    working_data['idle_buffer_between_moving'].append(working_time_diff)
                    idle_buffer_between_moving_sum = sum(working_data['idle_buffer_between_moving'])
                    vehicle_status_durations[v_code]['parked'] += timedelta(seconds=idle_buffer_between_moving_sum)

                    for v_code, changes in idle_status_obj.items():
                        TrackerRawData.objects.filter(
                            id__in=changes["tracking_ids"]
                        ).update(gis_geo_status="Parked")

                    vehicle_status_counts[v_code]['parked'] += 1

                    working_data['last_moving_time'] = False
                    working_data['next_moving_time'] = False
                    working_data['idle_buffer_between_moving'].clear()
                    gis_geo_status_logic_obj[v_code]['tracking_ids'].clear()

                    idle_status_obj[v_code]['tracking_ids'].clear()
                    idle_status_obj[v_code]['total_seconds'] = 0

            # ----------- PARKED AND OFFLINE DURATION CALCULATION  (END)----------
            if status == 'parked':
                if idle_status_obj[v_code]['total_seconds'] > 0:
                    vehicle_status_durations[v_code]['parked'] += time_diff
                    vehicle_status_durations[v_code]['parked'] += timedelta(
                        seconds=idle_status_obj[v_code]['total_seconds'])

                    for v_code, changes in idle_status_obj.items():
                        TrackerRawData.objects.filter(
                            id__in=changes["tracking_ids"]
                        ).update(gis_geo_status="Parked")

                    vehicle_status_counts[v_code]['parked'] += 1

                    idle_status_obj[v_code]['tracking_ids'].clear()
                    idle_status_obj[v_code]['total_seconds'] = 0

                updated_status = TrackerRawData.objects.filter(
                    id=tracking_id
                ).update(gis_geo_status='Parked')

            if prev_status == 'parked':
                vehicle_status_durations[v_code]['parked'] += time_diff

                # Count new segment if status changes
                if prev_status != status:
                    vehicle_status_counts[v_code]['parked'] += 1

                updated_status = TrackerRawData.objects.filter(
                    id=tracking_id
                ).update(gis_geo_status='Parked')

            # ----------- PARKED AND OFFLINE DURATION CALCULATION  (END)----------

            distance_accumulator[v_code] += distance

            working_data['prev_status'] = status

            if status == 'moving' and working_data['next_moving_time'] == False:
                working_data['last_moving_time'] = True
                working_data['prev_vendor_date_time'] = current_vendor_date_time
                # gis_geo_status_logic_obj[v_code]['tracking_ids'].append(tracking_id)

                vehicle_status_durations[v_code]['parked'] += timedelta(
                    seconds=idle_status_obj[v_code]['total_seconds'])

                for v_code, changes in idle_status_obj.items():
                    TrackerRawData.objects.filter(
                        id__in=changes["tracking_ids"]
                    ).update(gis_geo_status="Parked")

                vehicle_status_counts[v_code]['parked'] += 1

                idle_status_obj[v_code]['tracking_ids'].clear()
                idle_status_obj[v_code]['total_seconds'] = 0

            else:
                working_data['prev_vendor_date_time'] = current_vendor_date_time

        else:
            # ----------START-OF-DAY NO RESPONSE CHECK (START)----------
            start_of_day = timezone.make_aware(
                datetime.datetime.combine(current_vendor_date_time.date(), time.min),
                timezone.get_current_timezone()
            )

            initial_gap = current_vendor_date_time - start_of_day  # 05:02:00 - 00:00:00 = 5 Hours (NO Response)

            if initial_gap.total_seconds() > 300:  # 300 = 5 Minutes
                vehicle_status_durations[v_code]['offline'] += initial_gap
                vehicle_status_counts[v_code]['offline'] += 1
            # ----------START-OF-DAY NO RESPONSE CHECK (END)----------

            # # Check if first record is idle
            if status == 'idle':  ## change
                idle_status_obj[v_code]['tracking_ids'].append(tracking_id)

            if status == 'moving':
                working_data['last_moving_time'] = True
                working_data['prev_vendor_date_time'] = current_vendor_date_time
                gis_geo_status_logic_obj[v_code]['tracking_ids'].append(tracking_id)

        # Store current record for next iteration
        previous_records[v_code] = {
            'id': tracking_id,
            'vendor_date_time': current_vendor_date_time,
            'status': status,
            'distance': distance
        }

    else:  ### AFTER MOVING END THEN ALL POLE INFORMATION CONSIDER AS PARK
        print('Existance of Idles', idle_status_obj, len(idle_status_obj))
        print('Previous_Record : ', 0, len(previous_records))
        if idle_status_obj and idle_status_obj[v_code]['total_seconds'] > 0:
            vehicle_status_durations[v_code]['parked'] += timedelta(seconds=idle_status_obj[v_code]['total_seconds'])

            for v_code, changes in idle_status_obj.items():
                TrackerRawData.objects.filter(
                    id__in=changes["tracking_ids"]
                ).update(gis_geo_status="Parked")

            vehicle_status_counts[v_code]['parked'] += 1

        last_vendor_time = previous_records[v_code]['vendor_date_time']

        end_of_day = timezone.make_aware(
            datetime.datetime.combine(last_vendor_time.date(), time.max),  # 23:59:59
            timezone.get_current_timezone()
        )

        if last_vendor_time < end_of_day:
            offline_diff = end_of_day - last_vendor_time
            vehicle_status_durations[v_code]['offline'] += offline_diff

        vehicle_status_counts[v_code]['offline'] += 1

        idle_status_obj[v_code]['tracking_ids'].clear()
        idle_status_obj[v_code]['total_seconds'] = 0

    # Prepare the final output
    duration_data = []
    for v_code, status_durations in vehicle_status_durations.items():
        # Get the vehicle object from any record
        vehicle = next(r.vehicle_code for r in queryset if r.vehicle_code.vehicle_code == v_code)

        vehicle_data = {
            'vehicle_code': vehicle.vehicle_code,
            'pitb_code': vehicle.pitb_code,
            'register_no': vehicle.register_no,
            'chasis_no': vehicle.chasis_no,
            'vehicle_type': vehicle.vehicle_type,
            'distance': distance_accumulator.get(v_code, 0),
        }

        # Format all durations
        for status, duration in status_durations.items():
            status_id = STATUS_ID_MAP.get(status, None)
            if status_id:
                vehicle_data[status] = {
                    'id': status_id,
                    'duration': format_duration_hours_minutes(duration),
                    'count': vehicle_status_counts[v_code].get(status, 0)
                }

        # Ensure working is added even if no records
        if 'working' not in vehicle_data and 'working' in STATUS_ID_MAP:
            if 'working' in status_durations:
                vehicle_data['working'] = {
                    'id': STATUS_ID_MAP['working'],
                    'duration': format_duration_hours_minutes(status_durations['working']),
                    'count': vehicle_status_counts[v_code].get('working', 0)
                }

        # Ensure waiting is added even if no records
        if 'waiting' not in vehicle_data and 'waiting' in STATUS_ID_MAP:
            if 'waiting' in status_durations:
                vehicle_data['waiting'] = {
                    'id': STATUS_ID_MAP['waiting'],
                    'duration': format_duration_hours_minutes(status_durations['waiting']),
                    'count': vehicle_status_counts[v_code].get('waiting', 0)
                }

        duration_data.append(vehicle_data)

    return duration_data


### Calculate Working Hour of Single Vehicle (Using GIS_GEO_STATUS Column)
def CalculateSingleVehicleWorkingHour(querset):
    working_duration = timedelta()
    working_seconds = 0
    prev_vendor_date_time = None
    prev_gis_geo_status = None

    for record in querset:
        gis_geo_status = record.gis_geo_status.lower()
        current_vendor_date_time = record.vendor_date_time

        if prev_gis_geo_status == 'moving':
            time_diff = current_vendor_date_time - prev_vendor_date_time
            time_diff_seconds = time_diff.total_seconds()
            working_duration += time_diff
            working_seconds += time_diff_seconds

        prev_vendor_date_time = current_vendor_date_time
        prev_gis_geo_status = gis_geo_status

    # Convert total working seconds into integer hours
    working_hours = int(working_seconds // 3600)
    round_working_hours = "{:.2f}".format(working_hours)

    working_time = format_duration_hours_minutes(working_duration)

    return working_time


def TripReportView(request):
    template_name = 'TripReport.html'

    # Check if any filters are applied
    filters_applied = False

    ## Get filter values from POST request
    vehicle_code = request.POST.get('vehicle_code')
    vehicle_type = request.POST.get('vehicle_type')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    export = request.POST.get('export')

    # Convert date strings to DJANGO DATE
    if isinstance(end_date, str):
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
    if isinstance(start_date, str):
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

    # Initialize filters
    filters = {}
    if vehicle_type:
        filters['vehicle_code__vehicle_type'] = vehicle_type
    if vehicle_code:
        filters['vehicle_code__vehicle_code'] = vehicle_code
    if start_date and end_date:
        filters['vendor_date_time__date__range'] = (start_date, end_date)

    if filters:
        filters_applied = True
        # Query the filtered data and order by vehicle code and datetime
        queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by(
            'vendor_date_time'
        )
    else:
        queryset = TrackerRawData.objects.none()

    if queryset.exists():

        ### IF GEO GIS STATUS NULL THEN UPDATE ALL VALUE (START)
        null_status_qs = queryset.filter(Q(gis_geo_status__isnull=True) | Q(gis_geo_status=""))
        if null_status_qs.exists():
            CalculateVehicleGeoStatusDuration_Function(queryset)
        ### IF GEO GIS STATUS NULL THEN UPDATE ALL VALUE (END)

        # filters['gis_geo_status'] = 'Moving'
        # Refresh queryset after updating gis_geo_status
        queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by(
            'vehicle_code__vehicle_code', 'vendor_date_time'
        )

    ### Initialized Start Date & End Date for Frontend Fields(start_date/end_date) on page load
    if not start_date and not end_date:
        start_date = localdate()
        end_date = localdate()

    # Status mapping
    STATUS_ID_MAP = {
        "working": 1,
        "moving": 2,
        "parked": 3,
        "idle": 4,
        "offline": 5,
        "waiting": 6,
    }

    # Count of trips
    total_trips = 0

    # Total distance for each trip
    total_trip_distance = 0.0

    # Count Total Working Hours
    working_duration = timedelta()
    working_hours = "0H.0M"

    # Summary Stats
    total_distance = 0.0

    # Tracking variables
    consecutive_period_id = 0
    current_gis_geo_status = None
    gis_geo_status_start_time = None
    gis_geo_status_end_time = None
    prev_vendor_date_time = None
    prev_gis_geo_status = None
    prev_geom = None

    ## Result lists
    all_trip_duration = []
    all_tracker_points = []

    if queryset:
        for record in queryset:

            current_vendor_date_time = record.vendor_date_time
            gis_geo_status = (record.gis_geo_status or 'Status Missing').lower()
            current_geom = GEOSGeometry(record.geom) if record.geom else None

            ## Calculate Working Hour (START)
            if prev_gis_geo_status == 'moving':
                time_diff = current_vendor_date_time - prev_vendor_date_time
                working_duration += time_diff
                working_hours = format_duration_hours_minutes(working_duration)
            ## Calculate Working Hour (END)

            ## Calculate the distance (START)
            if prev_geom and current_geom:
                # geodesic needs (lat, lon)
                distance = geodesic(
                    (prev_geom.y, prev_geom.x),
                    (current_geom.y, current_geom.x)
                ).meters
                total_trip_distance += distance
            ## Calculate the distance (END)

            ## Initialized first record (START)
            if current_gis_geo_status is None:
                # First record
                current_gis_geo_status = gis_geo_status
                gis_geo_status_start_time = current_vendor_date_time
                prev_geom = GEOSGeometry(record.geom) if record.geom else None
            ## Initialized first record (END)

            else:

                gis_geo_status_end_time = current_vendor_date_time

                geo_status_offline = "No"
                time_diff = current_vendor_date_time - prev_vendor_date_time
                if time_diff.total_seconds() > 60 * 60:
                    geo_status_offline = "Yes"
                    gis_geo_status_end_time = prev_vendor_date_time

                ## Create second json [all_trip_duration] (START)
                if gis_geo_status != current_gis_geo_status:  # Moving not equal to Moving == True (Mean Gis-Geo-Status Change)
                    status_id = STATUS_ID_MAP.get(current_gis_geo_status.lower(), 0)
                    start_time = gis_geo_status_start_time
                    end_time = gis_geo_status_end_time
                    time_diff = end_time - start_time

                    duration_gap = format_duration_hours_minutes(time_diff)

                    # Calculate distance for the period
                    total_distance += total_trip_distance
                    distance_km = (float(total_trip_distance) / 1000)
                    round_distance_km = "{:.2f}".format(distance_km)

                    consecutive_period_id += 1
                    total_trips += 1

                    ### THIS CODE USED FOR TRIP JSON (SECOND CATEGORY)
                    if duration_gap != "00H.00M":
                        all_trip_duration.append({
                            'trip_id': consecutive_period_id,
                            'status_id': status_id,
                            'status': current_gis_geo_status,
                            'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                            'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                            'duration': duration_gap,
                            'distance': round_distance_km
                        })
                        # total_durations += time_diff
                        total_trip_distance = 0.0  # Reset total_trip_distance for the next trip calculation

                    if geo_status_offline == "Yes":
                        time_diff = current_vendor_date_time - gis_geo_status_end_time
                        offline_gap_duration = format_duration_hours_minutes(time_diff)
                        all_trip_duration.append({
                            'trip_id': consecutive_period_id + 1,  # CATEGORY WISE
                            'status_id': STATUS_ID_MAP.get('offline', 0),
                            'status': 'Offline',
                            'start_time': gis_geo_status_end_time.strftime("%Y-%m-%d %H:%M"),
                            'end_time': current_vendor_date_time.strftime("%Y-%m-%d %H:%M"),
                            'duration': offline_gap_duration,
                            'distance': '0'
                        })
                        # total_durations += time_diff
                        total_trip_distance = 0.0  # Reset total_trip_distance for the next trip calculation

                    ## Create second json [all_trip_duration] (END)

                    # Start new status
                    current_gis_geo_status = gis_geo_status
                    gis_geo_status_start_time = current_vendor_date_time

            ## Create third json [all_tracker_points] (START)
            # Add record to current period with IDs
            all_tracker_points.append({
                'trip_id': consecutive_period_id + 1,
                'status_id': STATUS_ID_MAP.get(current_gis_geo_status.lower(), 0),
                'tracker_data': {
                    "vehicle_code": vehicle_code,
                    'pitb_code': record.vehicle_code.pitb_code,
                    'register_no': record.vehicle_code.register_no,
                    'chasis_no': record.vehicle_code.chasis_no,
                    'gis_geo_status': current_gis_geo_status,
                    'vendor_date_time': current_vendor_date_time.strftime("%Y-%m-%d %H:%M"),
                    'device_status': record.device_status,
                    'speed': record.speed,
                    'location': record.location,
                }
            })
            ## Create third json [all_tracker_points] (END)

            prev_vendor_date_time = current_vendor_date_time
            prev_gis_geo_status = gis_geo_status
            prev_geom = current_geom

        else:
            ## Handle the last Trip if it wasn't closed (START)
            if current_gis_geo_status and gis_geo_status_start_time:
                start_time = gis_geo_status_start_time
                end_time = queryset.last().vendor_date_time
                time_diff = end_time - start_time
                duration_gap = format_duration_hours_minutes(time_diff)

                # Calculate distance for the period
                total_distance += total_trip_distance
                distance_km = (float(total_trip_distance) / 1000)
                round_distance_km = "{:.2f}".format(distance_km)

                consecutive_period_id += 1
                total_trips += 1

                all_trip_duration.append({
                    'trip_id': consecutive_period_id,
                    'status_id': STATUS_ID_MAP.get(current_gis_geo_status.lower(), 0),
                    'status': current_gis_geo_status,
                    'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                    'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                    'duration': duration_gap,
                    'distance': round_distance_km
                })
            ## Handle the last Trip if it wasn't closed (END)

    # Dropdown options - Distinct values
    vehicle_types = VehicleData.objects.values_list('vehicle_type', flat=True).distinct()
    g_statuses = VehicleLiveMonitor.objects.values_list('g_status', flat=True).distinct()
    ## Vehicle Code : Dropdown is get by using FetchVehicleCodes function (Ajax function)

    total_distance_km = (float(total_distance) / 1000)
    round_distance_km = "{:.2f}".format(total_distance_km)

    # Filter only "moving" records
    moving_trip_duration = [t for t in all_trip_duration if t.get("status") == "moving"]
    moving_tracker_points = [t for t in all_tracker_points if
                             t.get("tracker_data", {}).get("gis_geo_status") == "moving"]

    # Count moving trips
    total_moving_trips = len(moving_trip_duration)

    # Serialize the data for JavaScript
    all_trips_json = json.dumps(moving_trip_duration if filters_applied else [], cls=DjangoJSONEncoder)
    all_tracker_points_json = json.dumps(moving_tracker_points if filters_applied else [], cls=DjangoJSONEncoder)

    context = {
        'filters_applied': filters_applied,
        'total_trips': total_moving_trips if filters_applied else 0,
        'working_hours': working_hours if filters_applied else "0H.0M",
        'total_distance': round_distance_km if filters_applied else "0.00",

        'dropdown_vehicle_types': vehicle_types,
        'dropdown_g_statuses': g_statuses,

        'selected_vehicle_code': vehicle_code,
        'selected_vehicle_type': vehicle_type,

        'start_date': start_date,
        'end_date': end_date,

        'all_trip_points_json': all_trips_json,
        'all_tracker_points_json': all_tracker_points_json,

        'all_trip_points': moving_trip_duration if filters_applied else [],
        'all_tracker_points': moving_tracker_points

    }

    return render(request, template_name, context)


def OfflineTripReportView(request):
    template_name = 'OfflineTripReport.html'

    # Check if any filters are applied
    filters_applied = False

    ## Get filter values from POST request
    vehicle_code = request.POST.get('vehicle_code')
    vehicle_type = request.POST.get('vehicle_type')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    export = request.POST.get('export')

    # Convert date strings to date objects
    if isinstance(end_date, str):
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
    if isinstance(start_date, str):
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

    # Initialize filters
    filters = {}
    if vehicle_type:
        filters['vehicle_code__vehicle_type'] = vehicle_type
    if vehicle_code:
        filters['vehicle_code__vehicle_code'] = vehicle_code
    if start_date and end_date:
        filters['vendor_date_time__date__range'] = (start_date, end_date)

    if filters:
        filters_applied = True
        # Query the filtered data and order by vehicle code and datetime
        queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by('vendor_date_time')
    else:
        queryset = TrackerRawData.objects.none()

    ## Initialized Start Date & End Date for Frontend Fields(start_date/end_date) on page load
    if not start_date and not end_date:
        start_date = localdate()
        end_date = localdate()

    # Status mapping
    STATUS_ID_MAP = {
        "offline": 5,
    }

    # Count of offline trips
    total_offline_trips = 0

    # Total offline duration
    total_offline_duration = timedelta()

    # Tracking variables
    consecutive_period_id = 0
    prev_vendor_date_time = None
    prev_record = None

    ## Result lists
    all_offline_duration = []
    all_tracker_points = []

    if queryset:
        for i, record in enumerate(queryset):
            current_vendor_date_time = record.vendor_date_time

            # No Response Trip Start of the day (Offline)
            if prev_record is None:
                start_of_day = timezone.make_aware(
                    datetime.combine(current_vendor_date_time.date(), time.min)
                    # 00:00:00 timezone.get_current_timezone()
                )
                start_time = start_of_day
                end_time = current_vendor_date_time
                offline_duration = end_time - start_time  # 05:02:00 - 00:00:00 = 5 Hours (offline)

                consecutive_period_id += 1
                total_offline_trips += 1
                total_offline_duration += offline_duration

                all_offline_duration.append({
                    'trip_id': consecutive_period_id,
                    'status_id': STATUS_ID_MAP.get("offline", 0),
                    'status': 'offline',
                    'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                    'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                    'duration': format_duration_hours_minutes(offline_duration),
                    'distance': "0"
                })

            # Check for time gap offline (more than 60 minutes between records)
            if prev_record and prev_vendor_date_time:
                time_diff = current_vendor_date_time - prev_vendor_date_time

                # If time gap is more than 60 minutes, create an offline period
                if time_diff.total_seconds() > 60 * 60:  # 60 minutes
                    offline_start = prev_vendor_date_time
                    offline_end = current_vendor_date_time
                    offline_duration = offline_end - offline_start

                    consecutive_period_id += 1
                    total_offline_trips += 1
                    total_offline_duration += offline_duration

                    # Add offline period to duration list
                    all_offline_duration.append({
                        'trip_id': consecutive_period_id,
                        'status_id': STATUS_ID_MAP.get("offline", 5),
                        'status': "offline",
                        'start_time': offline_start.strftime("%Y-%m-%d %H:%M"),
                        'end_time': offline_end.strftime("%Y-%m-%d %H:%M"),
                        'duration': format_duration_hours_minutes(offline_duration),
                        'distance': "0.00"
                    })

                    # Add tracker points for the offline period
                    # Start of offline period (last known point before going offline)
                    all_tracker_points.append({
                        'trip_id': consecutive_period_id,
                        'status_id': STATUS_ID_MAP.get("offline", 5),
                        'tracker_data': {
                            "vehicle_code": prev_record.vehicle_code.vehicle_code,
                            'pitb_code': prev_record.vehicle_code.pitb_code,
                            'register_no': prev_record.vehicle_code.register_no,
                            'chasis_no': prev_record.vehicle_code.chasis_no,
                            'gis_geo_status': prev_record.gis_geo_status or "[Null]",
                            'vendor_date_time': offline_start.strftime("%Y-%m-%d %H:%M"),
                            'device_status': prev_record.device_status,
                            'speed': prev_record.speed,
                            'location': prev_record.location,
                        }
                    })

                    # End of offline period (first point after coming back online)
                    all_tracker_points.append({
                        'trip_id': consecutive_period_id,
                        'status_id': STATUS_ID_MAP.get("offline", 5),
                        'tracker_data': {
                            "vehicle_code": record.vehicle_code.vehicle_code,
                            'pitb_code': record.vehicle_code.pitb_code,
                            'register_no': record.vehicle_code.register_no,
                            'chasis_no': record.vehicle_code.chasis_no,
                            'gis_geo_status': record.gis_geo_status,
                            'vendor_date_time': offline_end.strftime("%Y-%m-%d %H:%M"),
                            'device_status': record.device_status,
                            'speed': record.speed,
                            'location': record.location,
                        }
                    })

            prev_vendor_date_time = current_vendor_date_time
            prev_record = record
        else:
            ## No Response Trip End of the day (Offline)
            end_of_day = timezone.make_aware(
                datetime.combine(prev_vendor_date_time.date(), time.max)  # 23:59:59 timezone.get_current_timezone()
            )
            start_time = prev_vendor_date_time
            end_time = end_of_day
            offline_duration = end_time - start_time  # 23:59:59 - 17:59:59 = 18 Hours (offline)

            consecutive_period_id += 1
            total_offline_trips += 1
            total_offline_duration += offline_duration

            all_offline_duration.append({
                'trip_id': consecutive_period_id,
                'status_id': STATUS_ID_MAP.get("offline", 0),
                'status': 'offline',
                'start_time': start_time.strftime("%Y-%m-%d %H:%M"),
                'end_time': end_time.strftime("%Y-%m-%d %H:%M"),
                'duration': format_duration_hours_minutes(offline_duration),
                'distance': "0"
            })
    # Dropdown options - Distinct values
    vehicle_types = VehicleData.objects.values_list('vehicle_type', flat=True).distinct()
    g_statuses = VehicleLiveMonitor.objects.values_list('g_status', flat=True).distinct()

    # Format total offline duration
    offline_hours = format_duration_hours_minutes(total_offline_duration)

    # Serialize the data for JavaScript
    all_offline_json = json.dumps(all_offline_duration if filters_applied else [], cls=DjangoJSONEncoder)
    all_tracker_points_json = json.dumps(all_tracker_points if filters_applied else [], cls=DjangoJSONEncoder)

    context = {
        'filters_applied': filters_applied,
        'total_trips': total_offline_trips if filters_applied else 0,
        'offline_hours': offline_hours if filters_applied else "0H.0M",
        'total_distance': "0.00",

        'dropdown_vehicle_types': vehicle_types,
        'dropdown_g_statuses': g_statuses,

        'selected_vehicle_code': vehicle_code,
        'selected_vehicle_type': vehicle_type,

        'start_date': start_date,
        'end_date': end_date,

        'all_trip_points_json': all_offline_json,
        'all_tracker_points_json': all_tracker_points_json,

        'all_trip_points': all_offline_duration if filters_applied else [],
        'all_tracker_points': all_tracker_points if filters_applied else []
    }

    return render(request, template_name, context)


def CalculateSingleVehicleWorkingHour_Function(tracker_records):
    """
    Calculate the total working hours for a single vehicle based on its tracker records.

    The function:
      - Iterates through sequential tracker data records for one vehicle.
      - Uses vehicle status (moving, idle, parked) to determine working time.
      - Handles short idle/parked intervals between moving sessions as part of working time.
      - Returns total working hours (integer hours).

    Args:
        tracker_records: Iterable of tracker data objects sorted by `vendor_date_time`.
                         Each record must have:
                            - acc_status (string)
                            - speed (numeric)
                            - ext_bat_voltage (numeric)
                            - vendor_date_time (datetime)

    Returns:
        int: Total working hours (integer part only).
    """

    # Timestamp of the previous record
    last_moving_vendor_date_time = None
    prev_status = None

    # Flag to track if vehicle has started moving at least once
    last_moving_time = False
    next_moving_time = False

    # Buffers to store short idle periods between moving sessions
    idle_buffer_between_moving = []

    # Total working time in seconds
    working_seconds = 0

    # Iterate through each tracker record in chronological order
    for current_record in tracker_records:
        acc_status = (current_record.acc_status or "off").lower()
        speed = (current_record.speed or 0)
        ext_bat_voltage = (current_record.ext_bat_voltage or 12)

        # Determine the current vehicle status based on record data
        status = "offline"
        if acc_status == "on" and speed > 0 and ext_bat_voltage > 9:
            status = "moving"
        elif acc_status == "on" and speed == 0 and ext_bat_voltage > 9:
            status = "idle"
        elif acc_status == "off" and speed == 0 and ext_bat_voltage > 9:
            status = "parked"
        elif acc_status == "off" and speed == 0 and ext_bat_voltage == 0:
            status = "offline"

        ### UPDATE VEHICLE STATUS BY GIS DEPT
        TrackerRawData.objects.filter(
            vehicle_code__vehicle_code=current_record.vehicle_code_id,
            vendor_date_time=current_record.vendor_date_time
        ).update(gis_geo_status=status)

        # Only process if we have a previous record to compare against
        if last_moving_time:

            ## LAST MOVING STATUS AND CURRENT STATUS  == MOVING
            if prev_status == 'moving' and status == 'moving':
                # If previous status was MOVING and current is also MOVING
                time_diff = (current_record.vendor_date_time - prev_vendor_date_time).total_seconds()
                working_seconds += time_diff
                last_moving_time = current_record.vendor_date_time

            # Moving -> (Idle , Parked, offline)
            # LAST STATUS = MOVING AND CURRENT STATUS = IDLE OR PARKED OFFLINE
            if prev_status == 'moving' and status in ['idle', 'parked', 'offline'] and len(
                    idle_buffer_between_moving) == 0:
                time_diff = (current_record.vendor_date_time - prev_vendor_date_time).total_seconds()
                if time_diff <= 20 * 60:  # 20 MIN
                    working_seconds += time_diff

            ## Moving (Start) .....(idles) -> idle ...
            if status == 'idle':
                time_diff = (current_record.vendor_date_time - prev_vendor_date_time).total_seconds()
                if time_diff <= 1 * 60:  # 1 MIN
                    idle_buffer_between_moving.append(time_diff)

            ## Moving (Start) .....(idles) -> idle ... Moving(End)
            # IF MOVING STOP AND CURRENT STATUS IDLE BUT WAITING TIME IS LESS THEN THREE MINT ALL TIME DURATION COUNT
            ### STATUS = MOVING AND ADD MOVING TIME
            if status == 'moving' and prev_status == 'idle':
                time_diff = (current_record.vendor_date_time - prev_vendor_date_time).total_seconds()
                if time_diff <= 3 * 60:  # 3 MIN
                    idle_buffer_between_moving.append(time_diff)
                    next_moving_time = True
                else:
                    idle_buffer_between_moving.clear()

            if next_moving_time:
                idle_buffer_between_moving_sum = sum(idle_buffer_between_moving)
                working_seconds += idle_buffer_between_moving_sum
                idle_buffer_between_moving.clear()
                next_moving_time = False
            ### STATUS = MOVING AND ADD MOVING TIME (END)

            elif status in ['parked', 'offline']:
                # If status PARKED or OFFLINE
                last_moving_time = False
                # Reset next moving time
                next_moving_time = False
                # Reset idle buffer
                idle_buffer_between_moving.clear()

            # Update previous status and record for next iteration
        prev_status = status

        # WHEN MOVING START AND NEXT MOVING FALSE
        if status == 'moving' and next_moving_time == False:
            last_moving_time = True
            prev_vendor_date_time = current_record.vendor_date_time

        else:
            prev_vendor_date_time = current_record.vendor_date_time

    # Convert total working seconds into integer hours
    working_hours = int(working_seconds // 3600)
    round_working_hours = "{:.2f}".format(working_hours)

    return round_working_hours


# Functin for getting excel
def ExportVehicleThresholdExcel(request):
    from datetime import datetime
    import xlwt
    from django.http import HttpResponse
    from .models import VehicleScheduleGPRSApi, VehicleThreshold  # adjust if necessary

    format_date = '%Y-%m-%d'
    today_date = datetime.now().strftime(format_date)
    get_vehicle_type = request.POST.get('cmd_vehicle_type')
    get_date = request.POST.get('date') or today_date

    filters = {'veh_sch_date': get_date}
    if get_vehicle_type and get_vehicle_type != "NA":
        filters['vehicle_code__vehicle_type'] = get_vehicle_type

    records = VehicleScheduleGPRSApi.objects.select_related('vehicle_code').filter(
        **filters
    ).order_by('vehicle_code__vehicle_type')

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Vehicle Threshold Data')

    header_style = xlwt.easyxf(
        'font: bold on; align: wrap on, horiz center; pattern: pattern solid, fore_colour gray25')

    headers = [
        '#', 'PITB Code', 'Register No', 'Type',
        'Exp Distance (km)', 'Exp Min Distance (km)', 'Exp Working Hours', 'Exp Ignition',
        'Act Distance (km)', 'Act Min Distance (km)', 'Act Working Hours', 'Act Ignition',
        'Status'
    ]

    # Write headers
    for col, header in enumerate(headers):
        ws.write(0, col, header, header_style)
        ws.col(col).width = 256 * 20  # Adjust column width
    ws.col(1).width = 256 * 30  # Vehicle column wider

    for row, rec in enumerate(records, 1):
        vehicle = rec.vehicle_code
        ws.row(row).height = 400 * 2

        threshold_obj = VehicleThreshold.objects.filter(vehicle_type=vehicle.vehicle_type).first()

        # Expected Threshold (from threshold table)
        exp_distance = threshold_obj.distance if threshold_obj else 'N/A'
        exp_min_distance = threshold_obj.min_distance if threshold_obj else 'N/A'
        exp_working_hours = threshold_obj.working_hours if threshold_obj and threshold_obj.working_hours is not None else 'N/A'
        exp_ignition = threshold_obj.ignition_status if threshold_obj else 'N/A'

        # Actual Threshold (from schedule record)
        act_distance = rec.distance if rec.distance is not None else (
            threshold_obj.distance if threshold_obj else 'N/A')
        act_min_distance = threshold_obj.min_distance if threshold_obj else 'N/A'
        act_working_hours = rec.working_hours if rec.working_hours is not None else (
            threshold_obj.working_hours if threshold_obj else 'N/A')
        act_ignition = rec.threshold if rec.threshold is not None else (
            threshold_obj.ignition_status if threshold_obj else 'N/A')

        ws.write(row, 0, row)
        ws.write(row, 1, vehicle.pitb_code or 'N/A')
        ws.write(row, 2, vehicle.register_no or 'N/A')
        ws.write(row, 3, vehicle.vehicle_type or 'N/A')

        ws.write(row, 4, exp_distance)
        ws.write(row, 5, exp_min_distance)
        ws.write(row, 6, exp_working_hours)
        ws.write(row, 7, exp_ignition)

        ws.write(row, 8, act_distance)
        ws.write(row, 9, act_min_distance)
        ws.write(row, 10, act_working_hours)
        ws.write(row, 11, act_ignition)

        ws.write(row, 12, "Yes" if rec.threshold == "Yes" else "No")

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Vehicle_Threshold_{today_date}.xls"'
    wb.save(response)
    return response


def VTMSReportView(request):
    template_name = "Report/VTMSReport.html"

    # Getting today vehicles distance and working hour
    today_schedule = VehicleScheduleGPRSApi.objects.filter(
        veh_sch_date=timezone.now().date(),
        vehicle_code_id=OuterRef('vehicle_code')
    )
    # Get filter parameters
    vehicle_type_filter = request.POST.get('cmd_vehicle_type')
    vehicle_code_filter = request.POST.get('vehicle_code')
    status_filter = request.POST.get('status_filter')
    delay_range_filter = request.POST.get('cmd_time_class')  # Add this for delay range

    # Use Django ORM to get ALL vehicles from VehicleData with LEFT JOIN to VehicleLiveMonitor
    queryset = VehicleData.objects.annotate(
        veh_code=F('vehicle_code'),
        vehicle_code_field=F('vehicle_code'),
        g_status=F('live_monitor__g_status'),
        speed=F('live_monitor__speed'),
        direction=F('live_monitor__direction'),
        device_status=F('live_monitor__device_status'),
        ignition_status=F('live_monitor__ignition_status'),
        geo_location=F('live_monitor__geo_location'),
        vendor_date_time=ExpressionWrapper(
            Cast(F('live_monitor__vendor_date_time'), output_field=DateTimeField()) - timedelta(hours=5),
            output_field=DateTimeField()
        ),
        duration=F('live_monitor__duration'),
        latitude=F('live_monitor__latitude'),
        longitude=F('live_monitor__longitude'),

        delay_hours=ExpressionWrapper(
            (Extract(Now() - F('vendor_date_time'), 'epoch') / 3600),
            output_field=FloatField()
        ),
        delay_minutes=ExpressionWrapper(
            (Extract(Now() - F('vendor_date_time'), 'epoch') / 60),
            output_field=FloatField()
        ),

        # Add a field to indicate if vehicle has live data
        has_live_data=Case(
            When(~Q(live_monitor__id__isnull=True), then=Value(True)),
            default=Value(False),
            output_field=models.BooleanField()
        ),
        vendor_date=TruncDate(Cast(F('live_monitor__vendor_date_time'), DateTimeField())),
        # status_display=Case(
        #     # Offline: No live data or not today
        #     When(Q(vendor_date__isnull=True) | ~Q(vendor_date=timezone.now().date()), then=Value('Offline')),
        #     When(Q(vendor_date=timezone.now().date()), then=Value('Working')),
        #     # # Delay: Today and time_diff_hours > 1
        #     default=Value('Offline'),
        #     output_field=models.CharField()
        # ),
        status_display=Case(
            # Offline: No live data or not today
            When(Q(vendor_date__isnull=True) | ~Q(vendor_date=timezone.now().date()), then=Value('Offline')),
            # Working: Today and time_diff_hours <= 1
            When(Q(vendor_date=timezone.now().date()) & Q(delay_hours__lte=1), then=Value('Working')),
            # # Delay: Today and time_diff_hours > 1
            When(Q(vendor_date=timezone.now().date()) & Q(delay_hours__gt=1), then=Value('Delay')),
            default=Value('Offline'),
            output_field=models.CharField()
        ),
        working_hours=Subquery(today_schedule.values('working_hours')[:1]),
        distance=Subquery(today_schedule.values('distance')[:1]),
    )

    ### WITHOUT NOT REPONSE VEHICLE DATA
    vtms_data_list = queryset.exclude(status_display='Offline').order_by(
        Case(
            When(delay_hours__gt=60, then=Value(1)),
            When(Q(delay_hours__gt=40) & Q(delay_hours__lte=60), then=Value(2)),
            When(Q(delay_hours__gt=20) & Q(delay_hours__lte=40), then=Value(3)),
            When(Q(delay_hours__gte=0) & Q(delay_hours__lte=20), then=Value(4)),
            When(delay_hours__isnull=True, then=Value(5)),
            default=Value(5),
            output_field=IntegerField()
        ),
        'vendor_date_time',
    )

    # Working count
    working_count = queryset.filter(status_display='Working').count()
    if working_count == 0:
        working_count = 1
    # Delay count
    delay_count = queryset.filter(status_display='Delay').count()
    # No Response count
    no_response_count = queryset.filter(status_display='Offline').count()

    total_vehicles_count = len(vtms_data_list)

    ### CODE FOR GRAPH SUMMARY DATASET
    graph_summary_list = []
    for vehicle in vtms_data_list:
        if vehicle.delay_hours is not None:
            summary_title = ""
            delay_minutes = round(vehicle.delay_hours * 60)  # Convert hours to minutes
            if delay_minutes > 60:
                summary_title = "Poor"
            elif 41 <= delay_minutes <= 60:
                summary_title = "Fair"
            elif 21 <= delay_minutes <= 40:
                summary_title = "Good"
            elif 0 <= delay_minutes <= 20:
                summary_title = "Excellent"

            ## CREATE VEHICLE DICT
            vehicle_dict = {
                'summary': summary_title
            }
            graph_summary_list.append(vehicle_dict)

    ### CODE FOR GRAPH SUMMARY DATASET
    color_progress = {
        'Excellent': '#65c15c',  # green
        'Poor': '#ff2b2b',  # red
        'Fair': '#ffb829',  # yellow
        'Good': '#7366FF',  # blue
        'NR': '#6c757d',  # gray
    }

    summary_counts = Counter(item['summary'] for item in graph_summary_list)
    graph_summary_list = []
    for key, value in summary_counts.items():
        percent = round((value / working_count) * 100, 2)
        color = color_progress.get(key, 'gray')  # default color if not in map
        graph_summary_list.append({
            'summary': key,
            'count': value,
            'percentage': percent,
            'color': color
        })

    # ✅ Labels
    set_pie_labels = [item['summary'] for item in graph_summary_list]

    # ✅ Counts (series)
    set_pie_series = [item['count'] for item in graph_summary_list]

    # ✅ Colors
    set_pie_colors = [item['color'] for item in graph_summary_list]

    ### CODE FOR GRAPH SUMMARY DATASET

    # Apply filters
    if vehicle_type_filter and vehicle_type_filter != "NA":
        vtms_data_list = vtms_data_list.filter(vehicle_type=vehicle_type_filter)

    if vehicle_code_filter and vehicle_code_filter != "NA":
        vtms_data_list = vtms_data_list.filter(vehicle_code=vehicle_code_filter)

    if delay_range_filter and delay_range_filter != "NA":
        if delay_range_filter == "1-20":
            vtms_data_list = vtms_data_list.filter(delay_hours__gte=1 / 60, delay_hours__lte=20 / 60)  # 1 to 20 minutes
        elif delay_range_filter == "21-40":
            vtms_data_list = vtms_data_list.filter(delay_hours__gte=21 / 60,
                                                   delay_hours__lte=40 / 60)  # 21 to 40 minutes
        elif delay_range_filter == "41-60":
            vtms_data_list = vtms_data_list.filter(delay_hours__gte=41 / 60,
                                                   delay_hours__lte=60 / 60)  # 41 to 60 minutes
        elif delay_range_filter == "above-60":
            vtms_data_list = vtms_data_list.filter(delay_hours__gt=60 / 60)  # above 60 minutes

    if status_filter and status_filter != "NA":
        vtms_data_list = vtms_data_list.filter(status_display=status_filter)
        vtms_data_list = vtms_data_list.exclude(status_display='').order_by(
            Case(
                When(delay_hours__gt=60, then=Value(1)),
                When(Q(delay_hours__gt=40) & Q(delay_hours__lte=60), then=Value(2)),
                When(Q(delay_hours__gt=20) & Q(delay_hours__lte=40), then=Value(3)),
                When(Q(delay_hours__gte=0) & Q(delay_hours__lte=20), then=Value(4)),
                When(delay_hours__isnull=True, then=Value(5)),
                default=Value(5),
                output_field=IntegerField()
            ),
            'vendor_date_time',

        )

    ### CODE FOR GRAPH SUMMARY DATASET

    # SET VTMS FILTER DROPDOWN
    vehicle_type_list = VehicleData.objects.values('vehicle_type').annotate(count=Count('id')).order_by('vehicle_type')
    vehicle_codes = VehicleData.objects.values('vehicle_code', 'register_no', 'chasis_no', 'pitb_code',
                                               'vehicle_type').annotate(
        veh_code=F('vehicle_code')
    ).order_by('vehicle_code')
    # Getting delay range count
    delay_range_counts = vtms_data_list.aggregate(
        range_1_20=Count('id', filter=Q(delay_hours__gte=1 / 60, delay_hours__lte=20 / 60)),
        range_21_40=Count('id', filter=Q(delay_hours__gte=21 / 60, delay_hours__lte=40 / 60)),
        range_41_60=Count('id', filter=Q(delay_hours__gte=41 / 60, delay_hours__lte=60 / 60)),
        range_above_60=Count('id', filter=Q(delay_hours__gt=60 / 60)),
    )

    context = {
        'vtms_data_list': vtms_data_list,
        'vehicle_type_list': vehicle_type_list,
        'vehicle_codes': vehicle_codes,
        'selected_vehicle_type': vehicle_type_filter,
        'selected_vehicle_code': vehicle_code_filter,
        'selected_status': status_filter,
        'selected_delay_range': delay_range_filter,  # <-- Add this line
        'page_title': "VTMS Report",
        # New code here *******
        'working_count': working_count,
        'delay_count': delay_count,
        'no_response_count': no_response_count,
        'total_vehicles_count': total_vehicles_count,
        # End of new code *******
        'graph_summary_list': graph_summary_list,
        'set_pie_labels': set_pie_labels,
        'set_pie_series': set_pie_series,
        'set_pie_colors': set_pie_colors,
        'delay_range_counts': delay_range_counts,
    }

    return render(request, template_name, context)


@require_POST
def connect_vehicle_to_route(request):
    """Check PITB code availability and connect vehicle to PITB route"""
    vehicle_code = request.POST.get('vehicle_code')
    pitb_code = request.POST.get('pitb_code')

    if not vehicle_code or not pitb_code:
        return JsonResponse({
            'success': False,
            'message': 'Vehicle code and PITB code are required'
        })

    try:
        # First, check if PITB code already exists
        if VehicleData.objects.filter(pitb_code=pitb_code).exists():
            return JsonResponse({
                'success': False,
                'message': 'This PITB code already exists. Please use a different PITB code.'
            })

        # If PITB code is available, proceed with connection
        vehicle = VehicleData.objects.get(vehicle_code=vehicle_code)
        vehicle.pitb_code = pitb_code
        vehicle.updated_at = datetime.datetime.now()
        vehicle.updated_by = "admin"
        vehicle.save()

        return JsonResponse({
            'success': True,
            'message': f'Vehicle {vehicle_code} successfully connected to route {pitb_code}'
        })

    except VehicleData.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Vehicle not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error connecting vehicle: {str(e)}'
        })


def ExportVTMSReportExcel(request):
    """Export VTMS Report data to Excel with the exact same data as displayed in the table"""

    # vehicle_type_filter = request.POST.get('vehicle_type')
    # vehicle_code_filter = request.POST.get('vehicle_code')
    # status_filter = request.POST.get('status_filter')
    vehicle_type_filter = request.GET.get('vehicle_type')
    vehicle_code_filter = request.GET.get('vehicle_code')
    status_filter = request.GET.get('status_filter')

    # Use Django ORM to get ALL vehicles from VehicleData with LEFT JOIN to VehicleLiveMonitor
    queryset = VehicleData.objects.annotate(
        veh_code=F('vehicle_code'),
        vehicle_code_field=F('vehicle_code'),
        g_status=F('live_monitor__g_status'),
        speed=F('live_monitor__speed'),
        direction=F('live_monitor__direction'),
        device_status=F('live_monitor__device_status'),
        ignition_status=F('live_monitor__ignition_status'),
        geo_location=F('live_monitor__geo_location'),
        vendor_date_time=Cast(F('live_monitor__vendor_date_time'), output_field=DateTimeField()),
        duration=F('live_monitor__duration'),
        latitude=F('live_monitor__latitude'),
        longitude=F('live_monitor__longitude'),

        delay_hours=Case(
            When(
                Q(pitb_code__isnull=True) & ~Q(live_monitor__vendor_date_time__isnull=True),
                then=ExpressionWrapper(
                    (Extract(Now() - Cast(F('live_monitor__vendor_date_time'), DateTimeField()), 'epoch') / 3600),
                    output_field=FloatField()
                )
            ),
            When(
                Q(pitb_code='') & ~Q(live_monitor__vendor_date_time__isnull=True),
                then=ExpressionWrapper(
                    (Extract(Now() - Cast(F('live_monitor__vendor_date_time'), DateTimeField()), 'epoch') / 3600),
                    output_field=FloatField()
                )
            ),
            default=None,
            output_field=FloatField()
        ),

        has_live_data=Case(
            When(~Q(live_monitor__id__isnull=True), then=Value(True)),
            default=Value(False),
            output_field=models.BooleanField()
        ),
        time_diff_hours=ExpressionWrapper(
            Extract(Now() - Cast(F('live_monitor__vendor_date_time'), DateTimeField()), 'epoch') / 3600,
            output_field=FloatField()
        ),
        vendor_date=TruncDate(Cast(F('live_monitor__vendor_date_time'), DateTimeField())),
        status_display=Case(
            When(Q(vendor_date__isnull=True) | ~Q(vendor_date=timezone.now().date()), then=Value('Offline')),
            When(Q(vendor_date=timezone.now().date()) & Q(time_diff_hours__lte=1.0), then=Value('Working')),
            When(Q(vendor_date=timezone.now().date()) & Q(time_diff_hours__gt=1.0), then=Value('Delay')),
            default=Value('Offline'),
            output_field=models.CharField()
        ),
    )

    if vehicle_type_filter:
        queryset = queryset.filter(vehicle_type=vehicle_type_filter)
    if vehicle_code_filter:
        queryset = queryset.filter(vehicle_code=vehicle_code_filter)
    if status_filter:
        queryset = queryset.filter(status_display=status_filter)

        data = queryset.order_by(
            Case(
                When(delay_hours__isnull=True, then=Value(4)),
                When(delay_hours__gt=60, then=Value(1)),
                When(Q(delay_hours__gte=20) & Q(delay_hours__lte=40), then=Value(3)),
                When(delay_hours__lt=20, then=Value(2)),
                default=Value(4)
            )
        )
    else:
        data = queryset.order_by(
            Case(
                When(delay_hours__isnull=True, then=Value(4)),
                When(delay_hours__gt=60, then=Value(1)),
                When(Q(delay_hours__gte=20) & Q(delay_hours__lte=40), then=Value(3)),
                When(delay_hours__lt=20, then=Value(2)),
                default=Value(4)
            )
        ).order_by('-has_live_data', '-vendor_date_time', 'vehicle_code').filter(
            vendor_date_time__date=timezone.now().date())

    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('VTMS Report')

    headers = [
        'S.No', 'Vehicle Code', 'PITB Code', 'Vehicle Type',
        'Vendor Date Time', 'Push Time', 'Delay Time (hrs)', 'Status'
    ]

    header_style = xlwt.easyxf(
        'font: bold on; align: wrap on, horiz center; '
        'pattern: pattern solid, fore_colour gray25; '
        'borders: left thin, right thin, top thin, bottom thin'
    )

    data_style = xlwt.easyxf(
        'borders: left thin, right thin, top thin, bottom thin'
    )

    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
        worksheet.col(col).width = 256 * 20

    for row, item in enumerate(data, 1):
        worksheet.write(row, 0, row, data_style)
        worksheet.write(row, 1, item.veh_code or "N/A", data_style)
        worksheet.write(row, 2, item.pitb_code or "Waiting", data_style)
        worksheet.write(row, 3, item.vehicle_type or "N/A", data_style)
        if item.vendor_date_time:
            local_dt = timezone.localtime(item.vendor_date_time).replace(tzinfo=None)
            worksheet.write(row, 4, local_dt.strftime("%d %b %Y, %H:%M"), data_style)

        else:
            worksheet.write(row, 4, "No Live Data", data_style)

        if item.vendor_date_time is not None:
            if isinstance(item.vendor_date_time, timedelta):
                worksheet.write(row, 5, str(item.vendor_date_time), data_style)
            elif isinstance(item.vendor_date_time, datetime):
                # Convert to naive (remove timezone info safely)
                local_dt = timezone.localtime(item.vendor_date_time).replace(tzinfo=None)
                worksheet.write(row, 5, local_dt.strftime("%d %b %Y, %H:%M"), data_style)
            else:
                worksheet.write(row, 5, str(item.vendor_date_time), data_style)
        else:
            worksheet.write(row, 5, "-", data_style)

        if item.delay_hours:
            worksheet.write(row, 6, f"{item.delay_hours:.2f}h", data_style)
        else:
            worksheet.write(row, 6, "-", data_style)

        if item.has_live_data:
            status = "Working" if item.pitb_code else "Delay"
        else:
            status = "No Response"
        worksheet.write(row, 7, status, data_style)

    current_time = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"VTMS_Report_{current_time}.xls"
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)

    return response


# 7/4/2025
# New code here
def VehicleThresholdView(request):
    template_name = 'VehicleThreshold.html'

    vehicle_types = VehicleData.objects.values_list('vehicle_type', flat=True).distinct()

    logger.info(f"Vehicle types: {vehicle_types}")

    existing_thresholds = VehicleThreshold.objects.all()

    logger.info(f"Existing thresholds: {existing_thresholds.count()}")

    context = {
        'vehicle_types': vehicle_types,
        'existing_thresholds': existing_thresholds,
    }
    return render(request, template_name, context)


# For Saving and Updating data in Vehicle Threshold
# def CreateUpdateVehicleThreshold(request):
#     if request.method == 'POST':
#         vehicle_type = request.POST.get('vehicle_type')
#         distance=request.POST.get('distance')
#         working_hours=request.POST.get('working_hours')
#         ignition=request.POST.get('ignition')
#         description=request.POST.get('description')

#         VehicleThreshold.objects.update_or_create(
#             vehicle_type=vehicle_type,
#             distance=distance,
#             working_hours=working_hours,
#             ignition_status=ignition,
#             description=description
#         )

#         return JsonResponse({'success': True, 'message': 'Vehicle threshold saved successfully'})
#     else:
#         return JsonResponse({'success': False, 'message': 'Invalid request method'})

def CreateUpdateVehicleThreshold(request):
    if request.method == 'POST':
        # Check if we're getting a single threshold or multiple
        if 'thresholds' in request.POST:
            # Handle multiple thresholds from JSON
            import json
            thresholds = json.loads(request.POST.get('thresholds'))

            for threshold in thresholds:
                vehicle_type = threshold.get('vehicle_type')
                distance = threshold.get('distance')
                distance = float(distance) if distance not in (None, '', 'null') else 0
                min_distance = threshold.get('min_distance')
                min_distance = float(min_distance) if min_distance not in (None, '', 'null') else 0
                working_hours = threshold.get('working_hours') if threshold.get('working_hours') else None
                ignition = threshold.get('ignition') if threshold.get('ignition') else 'No'
                description = threshold.get('description')

                # Use defaults parameter to update existing or create new
                VehicleThreshold.objects.update_or_create(
                    vehicle_type=vehicle_type,
                    defaults={
                        'distance': distance,
                        'min_distance': min_distance,
                        'working_hours': working_hours,
                        'ignition_status': ignition,
                        'description': description
                    }
                )

            return JsonResponse({'success': True, 'message': 'Vehicle thresholds saved successfully'})
        else:
            # Handle single threshold (original implementation)
            vehicle_type = request.POST.get('vehicle_type')
            distance = request.POST.get('distance')
            distance = float(distance) if distance not in (None, '', 'null') else 0
            min_distance = request.POST.get('min_distance')
            min_distance = float(min_distance) if min_distance not in (None, '', 'null') else 0
            working_hours = request.POST.get('working_hours') if request.POST.get('working_hours') else None
            ignition = request.POST.get('ignition') if request.POST.get('ignition') else 'No'
            description = request.POST.get('description')

            VehicleThreshold.objects.update_or_create(
                vehicle_type=vehicle_type,
                defaults={
                    'distance': distance,
                    'min_distance': min_distance,
                    'working_hours': working_hours,
                    'ignition_status': ignition,
                    'description': description
                }
            )

            return JsonResponse({'success': True, 'message': 'Vehicle threshold saved successfully'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})


def FetchVehicleRouteDataByDateandId(request):
    """
    Simplified version of FetchSingleVehicleRouteDataView that only requires
    vehicle_id and date parameters.
    """
    # Extract required parameters
    get_vehicle_id = request.POST.get('vehicle_id')
    get_date = request.POST.get('date')

    if not get_vehicle_id or not get_date:
        return HttpResponse(json.dumps({
            'message': 'Error: vehicle_id and date parameters are required',
            'tracker_raw_gprs_lists': [],
            'vehicle_tracks_length': []
        }, default=date_handler))

    # Define default time range - full day (00:00:00 to 23:59:59)
    get_time_from = "00:00"
    get_time_to = "23:59"

    format_str_date = "%Y-%m-%d"
    format_str_date_time = "%Y-%m-%d %H:%M:%S"
    current_data_time = datetime.datetime.now()

    try:
        # Parse date
        dt_selected_date = datetime.strptime(get_date, format_str_date).date()

        # Combine date and time into strings
        start_of_day_time_str = get_time_from + ':00'
        end_of_day_time_str = get_time_to + ':59'

        datetime_from = get_date + " " + start_of_day_time_str
        datetime_to = get_date + " " + end_of_day_time_str

        # Generate working schedule
        GenerateWorkingWithVehicleSchedule_Function(dt_selected_date)

        # Check vehicle vendor record status
        try:
            fetch_vehicle_sche_record = VehicleScheduleGPRSApi.objects.get(
                veh_sch_date=dt_selected_date,
                vehicle_code_id=get_vehicle_id
            )
            ck_process_status = fetch_vehicle_sche_record.process_status

            # If status is pending, fetch data from vendor API
            if ck_process_status == "Pending":
                # Get vendor server response
                response_gprs_api = ResponseTrackerGPRSRawApi_By_Vendor_Function(
                    get_vehicle_id,
                    datetime_from,
                    datetime_to
                )

                # Process API response if records exist
                if len(response_gprs_api) > 0 and 'Table' in response_gprs_api:
                    total_vendor_record = len(response_gprs_api['Table'])

                    vehicle_gprs_response_api = response_gprs_api['Table']
                    vehicle_gprs_response_api_len = len(vehicle_gprs_response_api)

                    # Process each record from the API
                    for g in range(vehicle_gprs_response_api_len):
                        ck_vehicle_code = vehicle_gprs_response_api[g]['VehicleID']
                        ck_vendor_date_time = str(vehicle_gprs_response_api[g]['GpsTime'])
                        ck_api_latitude = vehicle_gprs_response_api[g]['Lat']
                        ck_api_longitude = vehicle_gprs_response_api[g]['Long']

                        # Process date/time format
                        try:
                            if '.' in ck_vendor_date_time:
                                ck_vendor_date_time = ck_vendor_date_time.split('.')[0]

                            _ck_vendor_date_time = datetime.fromisoformat(ck_vendor_date_time)
                            formatted_vendor_date_time = _ck_vendor_date_time.strftime(format_str_date_time)
                        except ValueError:
                            formatted_vendor_date_time = datetime.strptime(ck_vendor_date_time, format_str_date_time)

                        # Check if record already exists, create if not
                        GetTrackerRawData = TrackerRawData.objects.filter(
                            vehicle_code_id=ck_vehicle_code,
                            vendor_date_time=ck_vendor_date_time,
                            latitude=ck_api_latitude,
                            longitude=ck_api_longitude
                        )

                        if len(GetTrackerRawData) == 0:
                            gprs_code_number = int(
                                TrackerRawData.objects.order_by('-id').values_list('id', flat=True).first()
                            )
                            auto_gprs_code = f"GPRS-{gprs_code_number}"

                            # Create geometry point
                            get_feature_coordinate = f"POINT({ck_api_longitude} {ck_api_latitude})"

                            # Create new tracker raw data record
                            TrackerRawData.objects.create(
                                gprs_raw_code=auto_gprs_code,
                                vehicle_code_id=ck_vehicle_code,
                                terminal_no=vehicle_gprs_response_api[g]['TerminalNo'],
                                geom=get_feature_coordinate,
                                latitude=ck_api_latitude,
                                longitude=ck_api_longitude,
                                g_status=vehicle_gprs_response_api[g]['Veh_Status'],
                                vehicle_status=vehicle_gprs_response_api[g]['Veh_Status'],
                                device_status=vehicle_gprs_response_api[g]['Dev_status'],
                                vendor_date_time=vehicle_gprs_response_api[g]['GpsTime'],
                                system_date_time=vehicle_gprs_response_api[g]['RecTime'],
                                speed=vehicle_gprs_response_api[g]['Speed'],
                                distance=vehicle_gprs_response_api[g]['Distance'],
                                direction=vehicle_gprs_response_api[g]['Direction'],
                                mileage=vehicle_gprs_response_api[g]['Mileage_Val'],
                                created_at=current_data_time,
                                created_by="admin"
                            )

                    # Count tracker records for the vehicle and date
                    tracker_gprs_length = TrackerRawData.objects.filter(
                        vehicle_code_id=get_vehicle_id,
                        vendor_date_time__date=dt_selected_date
                    ).aggregate(
                        total_count=Count('id')
                    )['total_count']

                    # Update vehicle schedule record
                    UpdateVehiclScheduleGPRSApi = VehicleScheduleGPRSApi.objects.get(
                        veh_sch_date=dt_selected_date,
                        vehicle_code_id=get_vehicle_id
                    )
                    UpdateVehiclScheduleGPRSApi.retrieve_record = tracker_gprs_length

                    if tracker_gprs_length >= total_vendor_record:
                        UpdateVehiclScheduleGPRSApi.process_status = "Completed"

                    UpdateVehiclScheduleGPRSApi.save()

        except VehicleScheduleGPRSApi.DoesNotExist:
            # If no schedule exists, create one
            logger.warning(f"No vehicle schedule found for vehicle {get_vehicle_id} on {dt_selected_date}")
            pass

        # Parse datetime strings to Django datetime objects
        django_datetime_from = parse_datetime(datetime_from)
        django_datetime_to = parse_datetime(datetime_to)

        # Query tracker data for the vehicle within the time range
        tracker_raw_gprs_lists = list(TrackerRawData.objects.filter(
            vehicle_code_id=get_vehicle_id,
            vendor_date_time__gte=django_datetime_from,
            vendor_date_time__lte=django_datetime_to
        ).annotate(
            x=RawSQL("ST_X(geom)", []),
            y=RawSQL("ST_Y(geom)", [])
        ).values(
            "x", "y", "vehicle_code_id", "g_status", "system_date_time",
            "vendor_date_time", "device_status", "max_speed", "speed",
            "vehicle_status", "direction", "distance", "mileage"
        ).order_by('id'))

        # Calculate vehicle track length from GPRS data
        VehicleTracks = TrackerRawData.objects.filter(
            vehicle_code_id=get_vehicle_id,
            vendor_date_time__gte=django_datetime_from,
            vendor_date_time__lte=django_datetime_to
        ).values('vehicle_code_id').annotate(
            line=MakeLine('geom'),
            length=Length(MakeLine('geom'))
        )

        # Convert track data to JSON format
        VehicleTracks_Length = [
            {
                "vehicle_code_id": track["vehicle_code_id"],
                "length_meters": track["length"].m if track["length"] else 0,
                "line_geojson": track["line"].geojson if track["line"] else None
            }
            for track in VehicleTracks
        ]

        # Prepare response
        message = "Success"
        params = {
            'message': message,
            'tracker_raw_gprs_lists': tracker_raw_gprs_lists,
            'vehicle_tracks_length': VehicleTracks_Length,
        }

        return HttpResponse(json.dumps(params, default=date_handler))

    except Exception as e:
        logger.error(f"Error in FetchVehicleRouteDataByDate: {str(e)}")
        return HttpResponse(json.dumps({
            'message': f'Error: {str(e)}',
            'tracker_raw_gprs_lists': [],
            'vehicle_tracks_length': []
        }, default=date_handler))


# TRACKERDATA VIEW LIST/SAVE/UPDATE (START)
def TerminalDataView(request):
    template = "TerminalInstallation.html"
    companies = TrackerCompany.objects.all()

    # Get filters from POST
    filter_terminal_type = request.POST.get("terminal_type")
    filter_terminal_status = request.POST.get("terminal_status")

    # Initialize Q object
    filters = Q()

    # Apply terminal type filter if provided and not 'NA'
    if filter_terminal_type and filter_terminal_type != "NA":
        filters &= Q(terminal_type=filter_terminal_type)

    # Apply terminal status filter if provided and not 'NA'
    if filter_terminal_status and filter_terminal_status != "NA":
        filters &= Q(status=filter_terminal_status)
    else:
        filters &= Q(status="Active")

    # Apply filters to query
    terminals = TrackerData.objects.filter(filters).order_by("-id")

    context = {
        'terminals': terminals,
        'selected_type': filter_terminal_type,
        'selected_status': filter_terminal_status,
        'companies': companies
    }

    return render(request, template, context)


@require_POST
def SaveTerminalDataView(request):
    # Handle form submission For Advance/Installed Terminal
    if request.method == "POST":
        terminal_code = request.POST.get('terminal_code')
        modal_no = request.POST.get('modal_no')
        sim_no = request.POST.get('sim_no')
        company_code = request.POST.get('company_code')
        terminal_type = request.POST.get('terminal_type')
        gsm_co = request.POST.get('gsm_co')

        tracker_company = TrackerCompany.objects.filter(company_code=company_code).first()

        # Save TelecomData
        telecom = TelecomData.objects.create(
            sim_no=sim_no,
            gsm_co=gsm_co,
            created_at=localtime(),
            created_by="admin"
        )
        # Save TerminalData
        terminal = TrackerData.objects.create(
            terminal_code=terminal_code,
            modal_no=modal_no,
            terminal_type=terminal_type,
            tracker_company_code=tracker_company,
            status="Blocked" if terminal_type == "Advanced" else "Active",
            created_at=localtime(),
            created_by="admin"
        )
        # Bridge table
        TrackerTelecomData.objects.create(
            terminal=terminal,
            sim=telecom,
            assigned_at=localtime(),
            assigned_by="admin",
            status="Blocked" if terminal_type == "Advanced" else "Active"
        )
        messages.success(request, "Terminal Added Successfully")
        return redirect('terminal-data')  # Redirect to avoid re-submission on refresh


@require_POST
def UpdateTerminalDataView(request):
    if request.method == 'POST':
        terminal_id = request.POST.get('terminal_id')
        terminal_code = request.POST.get('terminal_code')
        modal_no = request.POST.get('modal_no')
        terminal_type = request.POST.get('terminal_type')
        terminal = get_object_or_404(TrackerData, id=terminal_id)

        terminal.terminal_code = terminal_code
        terminal.modal_no = modal_no
        terminal.terminal_type = terminal_type
        terminal.status = "Blocked" if terminal_type == "Advanced" else "Active"
        terminal.updated_at = localtime()
        terminal.updated_by = 'admin'
        terminal.save()

        messages.success(request, "Terminal updated successfully.")
        return redirect('terminal-data')
    else:
        messages.error(request, "Invalid request method.")
        return redirect('terminal-data')


# TRACKERDATA VIEW LIST/SAVE/UPDATE (END)


# TELECOMDATA VIEW LIST/SAVE/UPDATE (START)
def TelecomDataView(request):
    template = "Telecomdata.html"
    context = {"sim": TelecomData.objects.all().order_by("-created_at")}

    if request.method == "POST":
        try:
            sim_no = request.POST.get('sim_no', '').strip()
            gsm_co = request.POST.get('gsm_co', '').strip()

            # Validate required fields
            if not sim_no or not gsm_co:
                messages.error(request, "Both SIM number and GSM company are required.")
                return render(request, template, context)

            # Check for existing SIM number
            if TelecomData.objects.filter(sim_no=sim_no).exists():
                messages.warning(request, f"SIM number {sim_no} already exists.")
                return render(request, template, context)

            # Create new record
            TelecomData.objects.create(
                sim_no=sim_no,
                gsm_co=gsm_co,
                created_at=localdate(),
                created_by=request.user.username if request.user.is_authenticated else "system"
            )
            messages.success(request, "Telecom data added successfully.")

            # Refresh the data after successful creation
            context["sim"] = TelecomData.objects.all().order_by("-created_at")

            return redirect('telecom-data')

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            # Log the error for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in TelecomDataView: {str(e)}", exc_info=True)

    return render(request, template, context)


@require_POST
def UpdateTelecomDataView(request):
    try:
        telecom_id = request.POST.get('telecom_id')
        sim_no = request.POST.get('sim_no', '').strip()
        gsm_co = request.POST.get('gsm_co', '').strip()

        # Validate required fields
        if not all([telecom_id, sim_no, gsm_co]):
            messages.error(request, "All fields are required.")
            return redirect('telecom-data')

        # Get the telecom record
        telecom = get_object_or_404(TelecomData, telecom_id=telecom_id)

        # Check if SIM number is being changed to an existing one
        if telecom.sim_no != sim_no and TelecomData.objects.filter(sim_no=sim_no).exists():
            messages.warning(request, f"SIM number {sim_no} already exists.")
            return redirect('telecom-data')

        # Update the record
        telecom.sim_no = sim_no
        telecom.gsm_co = gsm_co
        telecom.updated_at = localtime()
        telecom.updated_by = 'admin'
        telecom.save()

        messages.success(request, "Telecom data updated successfully.")
        return redirect('telecom-data')

    except Exception as e:
        messages.error(request, f"Failed to update telecom data: {str(e)}")
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in UpdateTelecomDataView: {str(e)}", exc_info=True)
        return redirect('telecom-data')


# TELECOMDATA VIEW LIST/SAVE/UPDATE (END)


# TRACKERTELECOMDATA VIEW (BRIDGE) LIST/SAVE/UPDATE (START)
def TrackerTelecomDataView(request):
    template = "TrackerTelecomData.html"

    # Initialize filter variables
    selected_type = request.POST.get('terminal_type', 'NA')
    selected_status = request.POST.get('terminal_status', 'NA')
    selected_vehicle_type = request.POST.get('vehicle_type', 'NA')
    export = request.POST.get('export', False)

    filters = Q()

    # Apply terminal type filter if provided and not 'NA'
    if selected_type and selected_type != "NA":
        filters &= Q(terminal__terminal_type=selected_type)

    if selected_status and selected_status != "NA":
        filters &= Q(terminal__status=selected_status)

    if selected_vehicle_type != "NA":
        filters &= Q(terminal__vehicle_code__vehicle_type=selected_vehicle_type)

    if filters:
        tracker_telecoms = TrackerTelecomData.objects.filter(filters).order_by('-assigned_at')
    else:
        # Get all tracker-telecom associations with related data
        tracker_telecoms = TrackerTelecomData.objects.select_related(
            'terminal', 'sim'
        ).order_by('-assigned_at')
    # Check if Excel download is requested
    if export:
        # Prepare data for Excel
        data = []
        for terminal in tracker_telecoms:
            data.append({
                'Vehicle Code': terminal.terminal.vehicle_code.vehicle_code if terminal.terminal.vehicle_code else "Wating",
                'Registration no': terminal.terminal.vehicle_code.register_no if terminal.terminal.vehicle_code.register_no else "Wating",
                'Chasis no': terminal.terminal.vehicle_code.chasis_no if terminal.terminal.vehicle_code.chasis_no else "Wating",
                'PITB Code': terminal.terminal.vehicle_code.pitb_code if terminal.terminal.vehicle_code.pitb_code else "Wating",
                'Vehicle Type': terminal.terminal.vehicle_code.vehicle_type if terminal.terminal.vehicle_code.vehicle_type else "Wating",
                'Terminal No': terminal.terminal.terminal_code,
                'Sim No': terminal.sim.sim_no,
                'Terminal Type': "Connected " if terminal.terminal.terminal_code.terminal_type == "Installed" else terminal.terminal.terminal_code.terminal_type,
                'Status': terminal.terminal.status
            })

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, sheet_name='TrackerTelecomData', index=False)

        # Auto-adjust column widths
        workbook = writer.book
        worksheet = writer.sheets['TrackerTelecomData']

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # e.g., 'A', 'B', etc.
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

        writer.close()
        output.seek(0)

        # Prepare response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=NoResponseVehicles.xlsx'
        return response

    # Get all vehicle types for dropdown
    vehicle_types = VehicleData.objects.values_list('vehicle_type', flat=True).distinct()

    context = {
        'tracker_telecoms': tracker_telecoms,
        'selected_type': selected_type,
        'selected_status': selected_status,
        'selected_vehicle_type': selected_vehicle_type,
        'vehicle_types': vehicle_types,
    }
    return render(request, template, context)


# TRACKERTELECOMDATA VIEW (BRIDGE) LIST/SAVE/UPDATE (END)


def VehicleTerminalPairingView(request):
    template = 'VehicleTerminalPairingForm.html'

    # Get all vehicle PITB codes for dropdown
    vehicle_codes = VehicleData.objects.all()

    # Get all Customer no. for dropdown
    customers = Customer.objects.all()

    # Get terminal data for dropdown (Advanced/Installed BOTH)
    terminals = TrackerData.objects.all()
    # terminals = TrackerData.objects.filter(status='Active')

    # Handle AJAX request for vehicle data after selecting the Vehcile Pitb Code from dropdown to showing vehicle data
    vehicle_pitb_code = request.GET.get("pitb_code")
    is_ajax_request = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if vehicle_pitb_code and is_ajax_request:
        vehicle_data = VehicleData.objects.filter(pitb_code=vehicle_pitb_code).first()

        if vehicle_data:
            # Return JSON response with vehicle data
            return JsonResponse({
                'success': True,
                'vehicle_data': {
                    'vehicle_code': vehicle_data.vehicle_code,
                    'register_no': vehicle_data.register_no,
                    'make': vehicle_data.make,
                    'engine_no': vehicle_data.engine_no,
                    'chasis_no': vehicle_data.chasis_no,
                    'color': vehicle_data.color,
                    'model': vehicle_data.model,
                    'cc': vehicle_data.cc,
                    'fuel_type': vehicle_data.fuel_type,
                    'vehicle_type': vehicle_data.vehicle_type,
                    'status': vehicle_data.status
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Vehicle not found'
            })

    context = {
        'vehicle_codes': vehicle_codes,
        'terminals': terminals,
        'customers': customers
    }
    return render(request, template, context)


# This function will check if its local ip or server ip
def Retrieve_IP_Address(request):
    """
    Check if the request is coming from a server IP.
    Returns True if the client IP is in the list of server IPs, False otherwise.
    """
    # Get the client IP
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]  # First IP in the list
    else:
        ip = request.META.get('REMOTE_ADDR')

    # Detect if local
    if ip in ("127.0.0.1", "::1") or ip.startswith("192.168.") or ip.startswith("10."):
        env_type = "Local Development"
    else:
        env_type = "Server / Production"

    # Check if client IP is in the list of server IPs
    return env_type


def PushToPITB(request):
    selected_date = request.POST.get('selected_date')
    get_vehicle_code = request.POST.get('vehicle_code')
    submission_type = request.POST.get('submission_type')  # Auto/Manual

    format_str = "%Y-%m-%d %H:%M:%S"
    current_date_time = datetime.datetime.now()
    current_date = current_date_time.date()

    # # Check if request is from a server IP
    # server_request = Retrieve_IP_Address(request)
    # if server_request == "Local Development":
    #     # Add 5 hours
    #     formatted_vendor_date_time = (_ck_vendor_date_time + timedelta(hours=5)).strftime(format_str)
    # else:
    #     formatted_vendor_date_time = _ck_vendor_date_time.strftime(format_str)  # Format as a string

    if not selected_date:
        selected_date = timezone.now().date()

    if selected_date:

        # Filter by vehicle code if provided
        vehicle_obj = VehicleData.objects.filter(vehicle_code=get_vehicle_code).only('vehicle_code').first()
        # logger.info(f"Selected vehicle code: {vehicle_obj.vehicle_code}")

        TrackerRawDataExists = TrackerRawData.objects.filter(
            vendor_date_time__date=selected_date,
            vehicle_code_id=vehicle_obj.vehicle_code
        ).exists()
        # logger.info(f"TrackerRawData exists for vehicle {vehicle_obj.vehicle_code}: {TrackerRawDataExists}")

        set_latitude = float(request.POST.get('lat'))
        set_longitude = float(request.POST.get('long'))
        if TrackerRawDataExists:
            ### GET LATEST RECORD
            latest_times = TrackerRawData.objects.filter(
                vendor_date_time__date=selected_date,
                vehicle_code=vehicle_obj.vehicle_code
            ).values('vehicle_code_id').annotate(
                latest_time=Max('vendor_date_time')
            )

            ### FETCH LATEST RECORD INFORMATION
            latest_record = TrackerRawData.objects.filter(
                vehicle_code=vehicle_obj.vehicle_code,
                vendor_date_time=list(latest_times)[0]['latest_time']
            ).values('vehicle_code_id', 'longitude', 'latitude', 'speed', 'device_status',
                     'vendor_date_time').first()

            ### SET PARAMETER
            set_latitude = float(latest_record['latitude'])
            set_longitude = float(latest_record['longitude'])

        extracted_data = {}
        data_to_pushed_to_pitb = []
        short_uuid = uuid.uuid4().hex[:25]
        if submission_type == "Manual":

            get_distance = int(round(float(request.POST.get('distance', 0))))
            get_working_minute = int(round(float(request.POST.get('working_minute', 0))))

            ### UPDATE DISTANCE AND WORKING HOURS IN VEHICLE SCHEDULE
            updated_rows = VehicleScheduleGPRSApi.objects.filter(
                veh_sch_date=selected_date,
                vehicle_code_id=get_vehicle_code
            ).update(
                distance=get_distance,
                working_hours=(get_working_minute / 60),
            )
            success = updated_rows > 0
            # print(f"UPDATE VEHICLE SCHEDULE: {current_date}")
            # print(f"UPDATE VEHICLE SCHEDULE: {success}")

            set_timestamp = selected_date + ' 23:58:23'  # '23:59:00'

            # If selected date is current date, use current time
            # Otherwise use end of day (23:59:59)
            current_date_str = current_date.strftime("%Y-%m-%d")  # string
            if selected_date == current_date_str:

                # Step 1: Add 5 hours
                datetime_plus_5 = current_date_time + timedelta(hours=5)
                # Step 2: Format to string
                datetime_from_5_hour_str = datetime_plus_5.strftime(format_str)

                # Check if request is from a server IP
                server_request = Retrieve_IP_Address(request)
                if server_request == "Local Development":
                    current_time = current_date_time.strftime('%H:%M:%S')
                    set_timestamp = str(selected_date) + " " + str(current_time)
                else:
                    set_timestamp = str(datetime_from_5_hour_str)

            logger.info("Sending Manual Data")
            extracted_data = {
                "vehicle_no": str(vehicle_obj.pitb_code),
                "uuid": str(short_uuid),
                "long": set_longitude,
                "lat": set_latitude,
                "speed": float(request.POST.get('speed')),
                "distance": get_distance,
                "working_hour": get_working_minute,
                "vehicle_status": request.POST.get('vehicle_status'),
                "engine_status": request.POST.get('engine_status'),
                "timestamp": set_timestamp
            }
        else:
            pass

        data_to_pushed_to_pitb.append(extracted_data)
        logger.info(f"Vehicle count {len(data_to_pushed_to_pitb)}")
        logger.info("Payload JSON: %s", json.dumps(data_to_pushed_to_pitb))

        response = None
        if data_to_pushed_to_pitb:
            headers = {
                "authkey": "SGWMC-SaSg-MC4yMTUxODE0Nzk0MzA5MjY0NzAuMzM3Nzk4MzQ0Mjg1",
                "Content-Type": "application/json",
                "Accept": "application/json"
            }
            api_url = "https://elgcd-ms.punjab.gov.pk/api/vtms/post-vtms-bulk-data"
            response = requests.post(
                url=api_url,
                json=data_to_pushed_to_pitb,
                headers=headers,
                timeout=30
            )
            logger.info(f"Response status: {response.status_code}")
            logger.info(f"Response body: {response.text}")

            return JsonResponse({
                "api_response": {
                    "status_code": response.status_code if response else None,
                    "body": response.json() if response and response.status_code in [200, 201] else (
                        response.text if response else None)
                },
                "data_sent": data_to_pushed_to_pitb,
                "count": len(data_to_pushed_to_pitb),
                "status": 200 if response and response.status_code in [200, 201] else (
                    response.status_code if response else None)
            })

        # if get_vehicle_code:
        #     # # Filter by vehicle code if provided
        #     # vehicle_obj = VehicleData.objects.filter(pitb_code=vehicle_code).only('vehicle_code').first()
        #
        #     logger.info(f"Selected vehicle code: {vehicle_obj.vehicle_code}")
        #     TrackerRawDataExists = TrackerRawData.objects.filter(
        #         vendor_date_time__date=selected_date,
        #         vehicle_code_id=vehicle_obj.vehicle_code
        #     ).exists()
        #     logger.info(f"TrackerRawData exists for vehicle {vehicle_obj.vehicle_code}: {TrackerRawDataExists}")
        #     latest_times = TrackerRawData.objects.filter(
        #         vendor_date_time__date=selected_date,
        #         vehicle_code=vehicle_obj.vehicle_code
        #     ).values('vehicle_code_id').annotate(
        #         latest_time=Max('vendor_date_time')
        #     )
        # else:
        #     latest_times = TrackerRawData.objects.filter(
        #         vendor_date_time__date=selected_date
        #     ).values('vehicle_code_id').annotate(
        #         latest_time=Max('vendor_date_time')
        #     )
        #
        # latest_records = []
        # for item in latest_times:
        #     if get_vehicle_code:
        #         record = TrackerRawData.objects.filter(
        #             vehicle_code_id=item['vehicle_code_id'],
        #             vendor_date_time=item['latest_time']
        #         ).values('vehicle_code_id', 'longitude', 'latitude', 'speed', 'device_status',
        #                  'vendor_date_time').first()
        #     else:
        #         record = TrackerRawData.objects.filter(
        #             vendor_date_time=item['latest_time']
        #         ).values('vehicle_code_id', 'longitude', 'latitude', 'speed', 'device_status', 'vendor_date_time',
        #                  'vehicle_code__pitb_code').first()
        #
        #     if record:
        #         latest_records.append(record)
        #
        # data_to_pushed_to_pitb = []
        # for data in latest_records:
        #     set_vehicle_status = "moving"
        #     if data['device_status'] == "ACC Off,Parked":
        #         set_vehicle_status = "waiting"
        #     elif data['device_status'] == "ACC On,Idle":
        #         set_vehicle_status = "idle"
        #
        #     status_match = re.findall(r'\b(Off|On)\b', data.get('device_status', "") or "")
        #     set_engine_status = status_match[0] if status_match else "Off"
        #
        #     vehicle_code_id = data['vehicle_code_id']
        #     longitude = data['longitude']
        #     latitude = data['latitude']
        #     speed = data['speed']
        #
        #     record = TrackerRawData.objects.filter(
        #         vehicle_code_id=vehicle_code_id,
        #         vendor_date_time__date=selected_date
        #     ).order_by('vendor_date_time')
        #     distance = DistanceFinder(record)
        #
        #     print("Length of records : ", len(record))
        #
        #     working_hour = CalculateSingleVehicleWorkingHour_Function(record)
        #     logger.info(f"Working hour for vehicle {get_vehicle_code}: {working_hour}")
        #     logger.info(f"vehicle_no: {get_vehicle_code}")
        #     # short_uuid = uuid.uuid4().hex[:25]
        #     if not get_vehicle_code:
        #         logger.info(f"Vehicle {data['vehicle_code__pitb_code']}")
        #     # Checking timestamp
        #     # Modified timestamp calculation with server IP check
        #     current_timestamp = datetime.now()
        #
        #     formatted_timestamp = current_timestamp.strftime("%Y-%m-%d %H:%M:%S") if str(selected_date) == str(
        #         timezone.now().date()) else str(data['vendor_date_time']).split('+')[0].strip()
        #     if submission_type == 'Auto':
        #         extracted_data = {
        #             "vehicle_no": str(vehicle_obj.pitb_code),
        #             "uuid": str(short_uuid),
        #             "long": float(longitude),
        #             "lat": float(latitude),
        #             "speed": float(speed),
        #             "distance": float(distance),
        #             "working_hour": round(float(working_hour) * 60, 2),
        #             "vehicle_status": set_vehicle_status,
        #             "engine_status": set_engine_status.lower(),
        #             "timestamp": formatted_timestamp
        #
        #         }
        #         # logger.info(f"Data to be pushed for vehicle {vehicle_code_id}")
        #         # data_to_pushed_to_pitb.append(extracted_data)
        #         # logger.info(f"Data to be pushed to PITB: {data_to_pushed_to_pitb}")
        #     else:
        #         pass
        #         # logger.info("Sending Manual Data")
        #         # extracted_data = {
        #         #     "vehicle_no": str(vehicle_obj.pitb_code),
        #         #     "uuid": str(short_uuid),
        #         #     "long": float(request.POST.get('long')),
        #         #     "lat": float(request.POST.get('lat')),
        #         #     "speed": float(request.POST.get('speed')),
        #         #     "distance": float(distance),
        #         #     "working_hour": round(float(working_hour) * 60, 2),
        #         #     "vehicle_status": request.POST.get('vehicle_status'),
        #         #     "engine_status": request.POST.get('engine_status'),
        #         #     "timestamp": request.POST.get('user_timestamp')
        #         # }
        #     data_to_pushed_to_pitb.append(extracted_data)
        #     logger.info(f"Vehicle count {len(data_to_pushed_to_pitb)}")
        #     logger.info("Payload JSON: %s", json.dumps(data_to_pushed_to_pitb))

    else:
        return JsonResponse({
            "api_response": None,
            "data_sent": None,
            "count": 0,
            "status": None
        })


@require_POST
def SaveVehicleTerminalPairing(request):
    try:
        # Get form data
        customer_num = request.POST.get('customer_num')
        vehicle_code = request.POST.get('vehicle_code')
        vehicle_code_missing = request.POST.get('vehicle_pitb_code_missing')
        old_terminal = request.POST.get('old_terminal')
        new_terminal = request.POST.get('new_terminal')
        installation_date = request.POST.get('installation_date')
        testing_time_from = request.POST.get('testing_time_from')
        testing_time_to = request.POST.get('testing_time_to')
        meter_reading = request.POST.get('meter_reading')
        install_type = request.POST.get('install_type')
        remarks = request.POST.get('remarks')
        region = request.POST.get('region')
        technician_name = request.POST.get('technician_name')
        installation_place = request.POST.get('installation_place')

        # Check if customer exists
        customer_obj = Customer.objects.get(customer_no=customer_num)

        # Get vehicle object - handle case where vehicle_code is None (Waiting)
        if vehicle_code and vehicle_code != "None":  # Regular case with existing PITB code
            vehicle_obj = VehicleData.objects.filter(vehicle_code=vehicle_code).first()

        else:  # Waiting case - find by register_no from the dropdown option
            # Extract register_no from the selected option (format: "Waiting - ABC-123")
            option_text = request.POST.get('vehicle_pitb_code_option_text', '')

            # Extract register_no from the option text (format: "Waiting - ABC-123")
            register_no = option_text.split(' - ')[-1].strip() if ' - ' in option_text else ''

            if not register_no:
                return JsonResponse({
                    'success': False,
                    'message': 'Could not determine vehicle registration number'
                })

            vehicle_obj = VehicleData.objects.filter(register_no=register_no).first()
            vehicle_obj.pitb_code = vehicle_code_missing
            vehicle_obj.save()

        if not vehicle_obj:
            return JsonResponse({
                'success': False,
                'message': 'Vehicle not found'
            })

        # Handle different remark cases
        if remarks == 'Installation':
            terminal = VehicleTerminalPairing.objects.filter(new_terminal=new_terminal)
            existing_vehicle_pairing = VehicleTerminalPairing.objects.filter(vehicle_code=vehicle_obj).order_by(
                '-id').first()
            # New installation Remark
            if not new_terminal:
                return JsonResponse({
                    'success': False,
                    'message': 'New terminal is required for installation'
                })
            # Check if the vehicle is already paired with a terminal (for new installation)
            elif existing_vehicle_pairing and existing_vehicle_pairing.new_terminal:
                return JsonResponse({
                    'success': False,
                    'message': f'Vehicle is already paired!'
                })
            # Check Terminal Number is already paired with vehicle
            elif terminal:
                return JsonResponse({
                    'success': False,
                    'message': 'Terminal ID is already paired!'
                })
            new_terminal_obj = TrackerData.objects.get(terminal_code=new_terminal)
            old_terminal_obj = None

        else:
            # Redo/Transfer/Recheck/Reinstall Remarks
            new_terminal_obj = TrackerData.objects.filter(terminal_code=new_terminal).first()
            old_terminal_obj = TrackerData.objects.filter(terminal_code=old_terminal).first() if old_terminal else None

            if not new_terminal_obj:
                return JsonResponse({
                    'success': False,
                    'message': 'New terminal not found'
                })

        # Convert meter reading to float if provided
        meter_reading_float = None
        if meter_reading:
            try:
                meter_reading_float = float(meter_reading)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid meter reading value'
                })

        # Create new VehicleTerminalPairing object
        pairing = VehicleTerminalPairing(
            customer=customer_obj,
            vehicle_code=vehicle_obj,
            old_terminal=old_terminal_obj,
            new_terminal=new_terminal_obj,
            installation_date=installation_date,
            testing_time_from=testing_time_from,
            testing_time_to=testing_time_to,
            meter_reading=meter_reading_float,
            install_type=install_type,
            remarks=remarks,
            region=region,
            technician_name=technician_name,
            installation_place=installation_place,
            created_by=request.user.username if request.user.is_authenticated else 'admin'
        )
        pairing.save()

        return JsonResponse({
            'success': True,
            'message': 'Vehicle terminal pairing saved successfully',
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        })


### ALL VEHICLE DESIGN AND MANAGEMENT (START)
def AllVehicleDataViewset(request):
    template = "VehicleData.html"
    # Get distinct values for dropdowns
    vehicle_type_list = VehicleData.objects.values_list('vehicle_type', flat=True).distinct()
    vehicle_code_list = VehicleData.objects.values_list('vehicle_code', flat=True).distinct()
    pitb_code_list = VehicleData.objects.values_list('pitb_code', flat=True).distinct()
    ownership_status_list = VehicleData.objects.values_list('ownership_status', flat=True).distinct()
    vehicle_data = VehicleData.objects.all()
    owner_code = OwnerData.objects.values_list('owner_code', flat=True).distinct()

    if request.method == "POST":

        if 'update_owner' in request.POST:
            vehicle_id = request.POST.get('vehicle_id')
            try:
                vehicle = VehicleData.objects.get(id=vehicle_id)
                before_data = json.dumps(model_to_dict(vehicle), cls=DjangoJSONEncoder)
                # List all editable fields except audit fields
                editable_fields = [
                    'vehicle_code', 'vendor_code', 'register_no', 'make', 'engine_no', 'chasis_no',
                    'color', 'model', 'cc', 'fuel_type', 'total_mileage', 'vehicle_type', 'status',
                    'pitb_code', 'owner_code', 'ownership_status', 'vtms_status', 'vehicle_use_code'
                ]
                for field in editable_fields:
                    value = request.POST.get(field)
                    if field == 'owner_code':
                        if value:
                            vehicle.owner_id = value
                    elif field == 'vehicle_use_code':
                        if value:
                            try:
                                vehicle_used_for = VehicleUsedFor.objects.get(vehicle_use_code=value)
                                vehicle.vehicle_use_code = vehicle_used_for
                            except VehicleUsedFor.DoesNotExist:
                                pass
                        else:
                            vehicle.vehicle_use_code = None
                    elif value is not None:
                        setattr(vehicle, field, value)
                after_data = json.dumps(model_to_dict(vehicle), cls=DjangoJSONEncoder)
                vehicle.save()
                # Now saving in vehicle data logs
                VehicleDataLog.objects.create(
                    vehicle_code=vehicle,
                    before_data=before_data,
                    after_data=after_data,
                    created_by="admin"
                )
                logger.info("Vehicle data saved in DB and logs")

                return JsonResponse({'success': True, 'message': 'Vehicle data updated successfully'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error updating vehicle: {str(e)}'})
        # Filter part
        vehicle_type = request.POST.get('vehicle_type', '')
        vehicle_code = request.POST.get('vehicle_code', '')
        pitb_code = request.POST.get('pitb_code', '')
        search = request.POST.get('search', '')
        ownership_type = request.POST.get('ownership_type', '')
        owner_code = request.POST.get('owner_code', '')

        filters = Q()
        if vehicle_type:
            filters &= Q(vehicle_type=vehicle_type)
        if vehicle_code:
            filters &= Q(vehicle_code=vehicle_code)
        if pitb_code:
            filters &= Q(pitb_code=pitb_code)
        if ownership_type:
            filters &= Q(ownership_status=ownership_type)
        if owner_code:
            filters &= Q(owner__owner_code=owner_code)
        if search:
            filters &= (
                    Q(vehicle_type__icontains=search) |
                    Q(vehicle_code__icontains=search) |
                    Q(pitb_code__icontains=search) |
                    Q(register_no__icontains=search) |
                    Q(make__icontains=search) |
                    Q(model__icontains=search) |
                    Q(color__icontains=search) |
                    Q(status__icontains=search)
            )
        vehicle_data = VehicleData.objects.filter(filters)
    vehicle_used_for_list = VehicleUsedFor.objects.filter(status='Active').order_by('vehicle_use_name')
    # Get owners separated by type
    self_owners = OwnerData.objects.filter(owner_type='Self').values('owner_code', 'name')
    vendor_owners = OwnerData.objects.filter(owner_type='Vendor').values('owner_code', 'name')
    context = {
        'vehicle_data': vehicle_data,
        'vehicle_type_list': vehicle_type_list,
        'vehicle_code_list': vehicle_code_list,
        'pitb_code_list': pitb_code_list,
        'owner_code_list': owner_code,
        'ownership_status_list': ownership_status_list,
        'vehicle_used_for_list': vehicle_used_for_list,
        'self_owners': self_owners,
        'vendor_owners': vendor_owners,
    }
    return render(request, template, context)


# template = "VehicleData.html"
#
# vehicle_ownership_list = VehicleData.objects.values('ownership_status').annotate(count=Count('id')).order_by(
#     'ownership_status')
# vehicle_type_list = VehicleData.objects.values('vehicle_type').annotate(count=Count('id')).order_by('vehicle_type')
#
# ### VEHICLE DATA DETAIL WITH VEHICLE STATUS AND VEHICLE TYPE
# cmd_vehicle_data = list(
#     VehicleData.objects.values('vehicle_code', 'vehicle_type', 'register_no', 'chasis_no')
#     .annotate(
#         pitb_code=Coalesce('pitb_code', Value('Waiting'), output_field=CharField())
#     )
#     .order_by('register_no')
#     .distinct()
# )
#
# vehicle_code_list = VehicleData.objects.values_list('vehicle_code', flat=True).distinct()
#
# vehicle_data = VehicleData.objects.all()
# owner_code = OwnerData.objects.values_list('owner_code', flat=True).distinct()
#
# if request.method == "POST":
#
#     if 'update_owner' in request.POST:
#         vehicle_id = request.POST.get('vehicle_id')
#         try:
#             vehicle = VehicleData.objects.get(id=vehicle_id)
#             # List all editable fields except audit fields
#             editable_fields = [
#                 'vehicle_code', 'vendor_code', 'register_no', 'make', 'engine_no', 'chasis_no',
#                 'color', 'model', 'cc', 'fuel_type', 'total_mileage', 'vehicle_type', 'status',
#                 'pitb_code', 'owner_code', 'ownership_status', 'vtms_status', 'vehicle_used_for'
#             ]
#             for field in editable_fields:
#                 value = request.POST.get(field)
#                 if field == 'owner_code':
#                     if value:
#                         vehicle.owner_id = value
#
#                 elif value is not None:
#                     setattr(vehicle, field, value)
#             vehicle.save()
#             return JsonResponse({'success': True, 'message': 'Vehicle data updated successfully'})
#         except Exception as e:
#             return JsonResponse({'success': False, 'message': f'Error updating vehicle: {str(e)}'})
#     # Filter part
#     vehicle_type = request.POST.get('vehicle_type', '')
#     vehicle_code = request.POST.get('vehicle_code', '')
#     pitb_code = request.POST.get('pitb_code', '')
#     search = request.POST.get('search', '')
#     ownership_type = request.POST.get('ownership_type', '')
#     owner_code = request.POST.get('owner_code', '')
#
#     filters = Q()
#     if vehicle_type:
#         filters &= Q(vehicle_type=vehicle_type)
#     if vehicle_code:
#         filters &= Q(vehicle_code=vehicle_code)
#     if pitb_code:
#         filters &= Q(pitb_code=pitb_code)
#     if ownership_type:
#         filters &= Q(ownership_status=ownership_type)
#     if owner_code:
#         filters &= Q(owner__owner_code=owner_code)
#     if search:
#         filters &= (
#                 Q(vehicle_type__icontains=search) |
#                 Q(vehicle_code__icontains=search) |
#                 Q(pitb_code__icontains=search) |
#                 Q(register_no__icontains=search) |
#                 Q(make__icontains=search) |
#                 Q(model__icontains=search) |
#                 Q(color__icontains=search) |
#                 Q(status__icontains=search)
#         )
#     vehicle_data = VehicleData.objects.filter(filters)
#
# context = {
#     'vehicle_data': vehicle_data,
#     'vehicle_type_list': vehicle_type_list,
#     'cmd_vehicle_data': cmd_vehicle_data,
#     'owner_code_list': owner_code,
#     'vehicle_ownership_list': vehicle_ownership_list
# }
# return render(request, template, context)


def VehicleDataLogsView(request):
    template = 'VehicleDataLogs.HTML'

    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Base queryset
    data = VehicleDataLog.objects.select_related('vehicle_code').all().order_by('-created_at')

    # Apply date filters if provided
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            data = data.filter(created_at__date__gte=start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            data = data.filter(created_at__date__lte=end_date_obj)
        except ValueError:
            pass

    # Pagination
    paginator = Paginator(data, 25)  # Show 25 logs per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'data': page_obj,
        'page_obj': page_obj,
    }
    return render(request, template, context)


def VehicleTypeManagerView(request):
    template = 'Vehicle/VehicleTypeManager.html'

    # Do not use select_related for non-relational fields (ImageField/TextField)
    vehicle_types_count = VehicleData.objects.values('vehicle_type').annotate(count=Count('id')).order_by(
        'vehicle_type')
    vehicle_types = VehicleType.objects.all().order_by('vehicle_type_name')
    logger.info(vehicle_types_count)
    # Prepare distinct used_for list for filter dropdown

    # Handle AJAX update (expects X-Requested-With header)
    if request.method == "POST" and request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.POST.get(
            'update_vehicle_type'):
        vt_id = request.POST.get('vehicle_type_id')
        try:
            vt = VehicleType.objects.get(id=vt_id)
            # handle uploaded icon file if provided
            if 'vehicle_icon' in request.FILES:
                vt.vehicle_icon = request.FILES['vehicle_icon']

            vt.updated_by = request.user.username if request.user.is_authenticated else vt.updated_by
            vt.updated_at = timezone.now()
            vt.save()
            return JsonResponse({'success': True, 'message': 'Vehicle type updated successfully'})
        except VehicleType.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Vehicle type not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error updating: {str(e)}'})

    # Handle AJAX POST for adding new vehicle types automatically
    if request.method == "POST" and request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.POST.get(
            'add_vehicle_type'):
        try:
            # Get all distinct vehicle types from VehicleData
            existing_vehicle_types = set(VehicleType.objects.values_list('vehicle_type_name', flat=True))
            vehicle_types_from_data = set(VehicleData.objects.values_list('vehicle_type', flat=True).distinct())

            # Find vehicle types that don't exist in VehicleType table
            new_vehicle_types = vehicle_types_from_data - existing_vehicle_types

            if not new_vehicle_types:
                return JsonResponse({
                    'success': False,
                    'message': 'No new vehicle types found to add. All vehicle types from VehicleData already exist in VehicleType table.'
                })

            # Create new vehicle types
            created_count = 0
            for veh_type in new_vehicle_types:
                if veh_type:  # Skip empty values
                    veh_type_count = VehicleType.objects.count()
                    VehicleType.objects.create(
                        vehicle_type_code=f"VH-{veh_type_count + 1:03d}",
                        vehicle_type_name=veh_type,
                        status='Active',
                        created_by=request.user.username if request.user.is_authenticated else "admin"
                    )
                    created_count += 1

            if created_count > 0:
                return JsonResponse({
                    'success': True,
                    'message': f'Successfully added {created_count} new vehicle type(s): {", ".join(new_vehicle_types)}'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'No valid vehicle types found to add.'
                })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error adding vehicle types: {str(e)}'
            })

    # Handle regular POST for adding a new vehicle type (non-AJAX fallback)
    if request.method == "POST" and request.POST.get('add_vehicle_type') and not request.headers.get(
            'x-requested-with'):
        try:
            # Get all distinct vehicle types from VehicleData
            existing_vehicle_types = set(VehicleType.objects.values_list('vehicle_type_name', flat=True))
            vehicle_types_from_data = set(VehicleData.objects.values_list('vehicle_type', flat=True).distinct())

            # Find vehicle types that don't exist in VehicleType table
            new_vehicle_types = vehicle_types_from_data - existing_vehicle_types

            created_count = 0
            for veh_type in new_vehicle_types:
                if veh_type:  # Skip empty values
                    veh_type_count = VehicleType.objects.count()
                    VehicleType.objects.create(
                        vehicle_type_code=f"VH-{veh_type_count + 1:03d}",
                        vehicle_type_name=veh_type,
                        status='Active',
                        created_by=request.user.username if request.user.is_authenticated else "admin"
                    )
                    created_count += 1

            from django.contrib import messages
            if created_count > 0:
                messages.success(request,
                                 f"Successfully added {created_count} new vehicle type(s): {', '.join(new_vehicle_types)}")
            else:
                messages.info(request,
                              "No new vehicle types found to add. All vehicle types from VehicleData already exist.")

            return redirect(request.path)

        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"Error adding vehicle types: {str(e)}")
            return redirect(request.path)

    context = {
        'vehicle_types': vehicle_types,
        'vehicle_types_count': vehicle_types_count
    }
    return render(request, template, context)


def VehicleUsedForManagerView(request):
    """
    View for managing Vehicle Used For records with secure form submission.
    No longer uses query parameters for editing records.
    """
    # message = ""
    template_name = 'VehicleUsedForManager.html'
    current_data_time = datetime.datetime.now()

    if request.method == 'POST':

        get_action_type = request.POST['action_type']
        get_unique_id = request.POST['unique_id']
        get_vehicle_use_name = request.POST['vehicle_use_name']
        get_description = request.POST['description']
        get_cmd_status = request.POST['cmd_status']

        if get_action_type == "NEW":

            # Create new record
            VehicleUsedFor.objects.create(
                vehicle_use_code=f"VU-{VehicleUsedFor.objects.count() + 1:03d}",
                vehicle_use_name=get_vehicle_use_name,
                description=get_description,
                status=get_cmd_status,
                created_at=current_data_time,
                created_by="admin"
            )
            messages.success(request, 'New record created successfully.')


        elif get_action_type == "UPDATE":

            record = get_object_or_404(VehicleUsedFor, id=get_unique_id)
            if record:
                record.vehicle_use_name = get_vehicle_use_name
                record.description = get_description
                record.status = get_cmd_status
                record.updated_at = current_data_time
                record.updated_by = current_data_time
                record.save()
                messages.success(request, 'Record updated successfully.')

        elif get_action_type == "DELETED":

            get_record = get_object_or_404(VehicleUsedFor, id=get_unique_id)

            if VehicleData.objects.filter(vehicle_use_code__id=get_unique_id).exists():
                get_record.status = "Block"
                get_record.updated_at = current_data_time
                get_record.updated_by = current_data_time
                get_record.save()
            else:
                get_record.delete()

            messages.success(request, 'Record deleted successfully.')

    # Get all records for the table
    data = VehicleUsedFor.objects.all().order_by('vehicle_use_name')

    context = {
        'data': data,
        # 'message': message
    }
    return render(request, template_name, context)


### ALL VEHICLE DESIGN AND MANAGEMENT (START)


def OwnerManagmentView(request):
    template_name = 'OwnerManagement.html'
    current_data_time = datetime.datetime.now()

    if request.method == 'POST':
        get_action_type = request.POST.get('action_type')
        get_unique_id = request.POST.get('unique_id')
        get_owner_code = request.POST.get('owner_code')
        get_name = request.POST.get('name')
        get_cnic = request.POST.get('cnic')
        get_phone_number = request.POST.get('phone_number')
        get_address = request.POST.get('address')
        get_owner_type = request.POST.get('owner_type')
        get_status = request.POST.get('status')

        if get_action_type == "NEW":
            # Check if owner_code or cnic already exists
            if OwnerData.objects.filter(owner_code=get_owner_code).exists():
                messages.error(request, f'Owner with code {get_owner_code} already exists.')
                return redirect('owner_management')

            if OwnerData.objects.filter(cnic=get_cnic).exists():
                messages.error(request, f'Owner with CNIC {get_cnic} already exists.')
                return redirect('owner_management')

            # Check if Self type already exists
            if get_owner_type == 'Self' and OwnerData.objects.filter(owner_type='Self').exists():
                messages.error(request, 'Only one owner with type "Self" is allowed.')
                return redirect('owner_management')

            # Generate owner code if not provided
            if not get_owner_code:
                owner_count = OwnerData.objects.count()
                get_owner_code = f"OW-{owner_count + 1:03d}"

            # Create new owner
            OwnerData.objects.create(
                owner_code=get_owner_code,
                name=get_name,
                cnic=get_cnic,
                phone_number=get_phone_number,
                address=get_address,
                owner_type=get_owner_type,
                status=get_status,
                created_at=current_data_time,
                created_by=request.user.username if request.user.is_authenticated else "admin"
            )
            messages.success(request, 'New owner created successfully.')

        elif get_action_type == "UPDATE":
            owner = get_object_or_404(OwnerData, id=get_unique_id)

            # Check if Self type already exists for different owner
            if get_owner_type == 'Self' and OwnerData.objects.filter(owner_type='Self').exclude(
                    id=get_unique_id).exists():
                messages.error(request, 'Only one owner with type "Self" is allowed.')
                return redirect('owner_management')

            # Check if owner_code is being changed and if owner has vehicles
            if owner.owner_code != get_owner_code and hasattr(owner,
                                                              'vehicledata_set') and owner.vehicledata_set.exists():
                messages.error(request, 'Cannot change Owner Code for owners with associated vehicles.')
                return redirect('owner_management')

            # Check if owner_code already exists for different owner
            if OwnerData.objects.filter(owner_code=get_owner_code).exclude(id=get_unique_id).exists():
                messages.error(request, f'Owner with code {get_owner_code} already exists.')
                return redirect('owner_management')

            # Check if cnic already exists for different owner
            if OwnerData.objects.filter(cnic=get_cnic).exclude(id=get_unique_id).exists():
                messages.error(request, f'Owner with CNIC {get_cnic} already exists.')
                return redirect('owner_management')

            # Update owner
            owner.owner_code = get_owner_code
            owner.name = get_name
            owner.cnic = get_cnic
            owner.phone_number = get_phone_number
            owner.address = get_address
            owner.owner_type = get_owner_type
            owner.status = get_status
            owner.updated_at = current_data_time
            owner.updated_by = request.user.username if request.user.is_authenticated else "admin"
            owner.save()
            messages.success(request, 'Owner updated successfully.')

        elif get_action_type == "DELETED":
            owner = get_object_or_404(OwnerData, id=get_unique_id)

            # Soft delete: set status to Block
            owner.status = "Block"
            owner.updated_at = current_data_time
            owner.updated_by = request.user.username if request.user.is_authenticated else "admin"
            owner.save()
            messages.success(request, f'Owner {owner.owner_code} blocked successfully.')

    # Get all owners for the table
    owners = OwnerData.objects.all().order_by('owner_code')

    context = {
        'owners': owners,
    }

    return render(request, template_name, context)


def DriverManagmentView(request):
    template_name = 'DriverManagment.html'
    current_data_time = datetime.datetime.now()

    if request.method == 'POST':
        get_action_type = request.POST.get('action_type')
        get_unique_id = request.POST.get('unique_id')
        get_name = request.POST.get('name')
        get_cnic = request.POST.get('cnic')
        get_license_no = request.POST.get('license_no')
        get_license_expiry = request.POST.get('license_expiry')
        get_phone_number = request.POST.get('phone_number')

        if get_action_type == "NEW":
            # Generate a unique driver code (e.g., DRV-001)
            last_driver = DriverData.objects.order_by('-id').first()
            if last_driver:
                last_id = int(last_driver.driver_code.split('-')[1]) if '-' in last_driver.driver_code else 1
                new_code = f"DRV-{last_id + 1:03d}"
            else:
                new_code = "DRV-001"

            # Create new driver
            DriverData.objects.create(
                driver_code=new_code,
                name=get_name,
                cnic=get_cnic,
                license_no=get_license_no,
                license_expiry=get_license_expiry,
                phone_number=get_phone_number,
                created_at=current_data_time,
                created_by=request.user.username if request.user.is_authenticated else "admin"
            )
            messages.success(request, 'New driver created successfully.')

        elif get_action_type == "UPDATE":
            driver = get_object_or_404(DriverData, id=get_unique_id)
            if driver:
                driver.name = get_name
                driver.cnic = get_cnic
                driver.license_no = get_license_no
                driver.license_expiry = get_license_expiry
                driver.phone_number = get_phone_number
                driver.updated_at = current_data_time
                driver.updated_by = request.user.username if request.user.is_authenticated else "admin"
                driver.save()
                messages.success(request, 'Driver updated successfully.')

        elif get_action_type == "DELETED":
            driver = get_object_or_404(DriverData, id=get_unique_id)

            # Check if driver is assigned to any vehicle (if you have such relationship)
            # For now, we'll just delete the driver directly
            driver.delete()
            messages.success(request, 'Driver deleted successfully.')

    # Get all drivers for the table
    driver_data = DriverData.objects.all().order_by('driver_code')
    today_date = timezone.now().date()

    context = {
        'drivers_data': driver_data,
        'today_date': today_date
    }

    return render(request, template_name, context)


# Driver Vehicle View
def VehicleDriverAssignmentView(request):
    template_name = 'VehicleDriverAssignment.html'
    current_data_time = datetime.datetime.now()

    if request.method == 'POST':
        get_action_type = request.POST.get('action_type')
        get_unique_id = request.POST.get('unique_id')
        get_vehicle_code = request.POST.get('vehicle_code')
        get_driver_code = request.POST.get('driver_code')
        get_start_date = request.POST.get('start_date')
        get_end_date = request.POST.get('end_date') or None

        if get_action_type == "NEW":
            try:
                # Check if this vehicle-driver combination already exists for active assignments
                existing_assignment = VehicleDriver.objects.filter(
                    vehicle_code__vehicle_code=get_vehicle_code,
                    driver_code__driver_code=get_driver_code,
                    end_date__isnull=True  # Active assignments have no end date
                ).exists()

                if existing_assignment:
                    messages.error(request, 'Active assignment already exists for this vehicle-driver combination.')
                    return redirect(request.path)

                vehicle = get_object_or_404(VehicleData, vehicle_code=get_vehicle_code)
                driver = get_object_or_404(DriverData, driver_code=get_driver_code)

                VehicleDriver.objects.create(
                    vehicle_code=vehicle,
                    driver_code=driver,
                    start_date=get_start_date,
                    end_date=get_end_date,
                    updated_by=request.user.username if request.user.is_authenticated else 'admin'
                )
                messages.success(request, 'Vehicle-driver assignment created successfully.')
            except Exception as e:
                messages.error(request, f'Error creating assignment: {str(e)}')

        elif get_action_type == "UPDATE":
            try:
                assignment = get_object_or_404(VehicleDriver, id=get_unique_id)
                vehicle = get_object_or_404(VehicleData, vehicle_code=get_vehicle_code)
                driver = get_object_or_404(DriverData, driver_code=get_driver_code)

                assignment.vehicle_code = vehicle
                assignment.driver_code = driver
                assignment.start_date = get_start_date
                assignment.end_date = get_end_date
                assignment.updated_by = request.user.username if request.user.is_authenticated else 'admin'
                assignment.save()

                messages.success(request, 'Vehicle-driver assignment updated successfully.')
            except Exception as e:
                messages.error(request, f'Error updating assignment: {str(e)}')

        elif get_action_type == "DELETED":
            try:
                assignment = get_object_or_404(VehicleDriver, id=get_unique_id)
                assignment.delete()
                messages.success(request, 'Vehicle-driver assignment deleted successfully.')
            except Exception as e:
                messages.error(request, f'Error deleting assignment: {str(e)}')

        return redirect(request.path)

    # Get all assignments for the table
    today = timezone.now().date()

    vehicle_driver_data = VehicleDriver.objects.select_related('vehicle_code', 'driver_code').annotate(
        driver_phone=F('driver_code__phone_number')
    ).order_by('-updated_at')

    # Updated to use vehicle_code instead of pitb_code
    vehicle_code_list = VehicleData.objects.filter(vehicle_code__isnull=False).exclude(vehicle_code='').values_list(
        'vehicle_code', 'register_no')
    driver_code_list = DriverData.objects.all().values_list('driver_code', 'name')

    context = {
        'vehicle_driver_data': vehicle_driver_data,
        'vehicle_code_list': vehicle_code_list,
        'driver_code_list': driver_code_list,
        'today_date': today,
    }
    return render(request, template_name, context)


# PUsh no response vehicle
def PushPITBNotResponseVehicleData(request):
    vehicles = VehicleData.objects.annotate(
        vehicle_code=F('vehicle_code'),
        delay_hours=ExpressionWrapper(
            (Now() - F('vendor_date_time')) / timedelta(hours=1),
            output_field=FloatField()
        ),
        vendor_date=TruncDate(Cast(F('live_monitor__vendor_date_time'), DateTimeField()))
    )

    # Vehicles with delay more than 20 minutes (0.333 hours)
    delay_vehicles = vehicles.filter(delay_hours__gte=0.333).values_list('vehicle_code', flat=True)
    logger.info(f"Delayed vehicles more than 20 min {delay_vehicles}")
    selected_date = timezone.now().date()
    latest_times = TrackerRawData.objects.filter(
        vendor_date_time__date=selected_date,
        vehicle_code__in=delay_vehicles
    ).values('vehicle_code_id').annotate(
        latest_time=Max('vendor_date_time')
    )
    latest_records = []
    for item in latest_times:
        record = TrackerRawData.objects.filter(
            vehicle_code_id=item['vehicle_code_id'],
            vendor_date_time=item['latest_time']
        ).values('vehicle_code_id', 'longitude', 'latitude', 'speed', 'device_status', 'vendor_date_time').first()
        if record:
            latest_records.append(record)

    data_to_pushed_to_pitb = []
    for data in latest_records:
        set_vehicle_status = "moving"
        if data['device_status'] == "ACC Off,Parked":
            set_vehicle_status = "waiting"
        elif data['device_status'] == "ACC On,Idle":
            set_vehicle_status = "idle"

        status_match = re.findall(r'\b(Off|On)\b', data.get('device_status', "") or "")
        set_engine_status = status_match[0] if status_match else "Off"
        vehicle_code_id = data['vehicle_code_id']
        longitude = data['longitude']
        latitude = data['latitude']

        # Getting distance and working hour
        Distance_Working_hr = VehicleScheduleGPRSApi.objects.filter(vehicle_code_id=data['vehicle_code_id'],
                                                                    veh_sch_date=selected_date).values('distance',
                                                                                                       'working_hours').first()
        distance = Distance_Working_hr['distance']
        working_hr = Distance_Working_hr['working_hours']
        # Getting the pitb code of vehicle code
        vehicle = VehicleData.objects.filter(vehicle_code=vehicle_code_id).first()
        short_uuid = uuid.uuid4().hex[:25]
        current_timestamp = datetime.now()
        formatted_timestamp = current_timestamp.strftime("%Y-%m-%d %H:%M:%S")
        extracted_data = {
            "vehicle_no": str(vehicle.pitb_code),
            "uuid": str(short_uuid),
            "long": float(longitude),
            "lat": float(latitude),
            "speed": 0,
            "distance": float(distance),
            "working_hour": round(float(working_hr) * 60, 2),
            "vehicle_status": set_vehicle_status,
            "engine_status": set_engine_status.lower(),
            "timestamp": formatted_timestamp

        }
        data_to_pushed_to_pitb.append(extracted_data)
        logger.info(f"Vehicle count {len(data_to_pushed_to_pitb)}")
        logger.info("Payload JSON: %s", json.dumps(data_to_pushed_to_pitb))
    response = None
    if data_to_pushed_to_pitb:
        headers = {
            "authkey": "SGWMC-SaSg-MC4yMTUxODE0Nzk0MzA5MjY0NzAuMzM3Nzk4MzQ0Mjg1",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        api_url = "https://elgcd-ms.punjab.gov.pk/api/vtms/post-vtms-bulk-data"
        response = requests.post(
            url=api_url,
            json=data_to_pushed_to_pitb,
            headers=headers,
            timeout=30
        )
        logger.info(f"Response status: {response.status_code}")
        logger.info(f"Response body: {response.text}")

        return JsonResponse({
            "api_response": {
                "status_code": response.status_code if response else None,
                "body": response.json() if response and response.status_code in [200, 201] else (
                    response.text if response else None)
            },
            "data_sent": data_to_pushed_to_pitb,
            "count": len(data_to_pushed_to_pitb),
            "status": 200 if response and response.status_code in [200, 201] else (
                response.status_code if response else None)
        })
    else:
        return JsonResponse({
            "api_response": None,
            "data_sent": None,
            "count": 0,
            "status": None
        })


def PostPITBVehicleIcon(request):
    template = 'PostPITBVehicleIconPage.html'

    vehicle_types = VehicleType.objects.all().only('vehicle_type_name')

    context = {
        'vehicle_types': vehicle_types,
    }
    return render(request, template, context)


# Single Vehicle Info
def SingleVehicleInfo(request, register_no):
    template_name = "SingleVehicleInfo.html"
    vehicle = VehicleData.objects.filter(register_no=register_no).first()
    logger.info(f"Current vehicle type {vehicle.vehicle_type}")
    vehicle_type = VehicleType.objects.filter(vehicle_type_name=vehicle.vehicle_type).first()

    # Access the foreign key correctly
    vehicle_used_for = vehicle.vehicle_use_code.vehicle_use_name if vehicle.vehicle_use_code else None
    logger.info(f"Vehicle used for: {vehicle_used_for}")

    vehicle_icon = vehicle_type.vehicle_icon if vehicle_type else None

    context = {
        "vehicle": vehicle,
        "vehicle_icon": vehicle_icon,
        "vehicle_used_for": vehicle_used_for
    }
    return render(request, template_name, context)


def TrackerHistoryView(request):
    template_name = 'TrackerHistoryReport.html'

    get_vehicle_type = request.POST.get('cmd_vehicle_type')
    get_vehicle_code = request.POST.get('cmd_vehicle_list')
    get_status = request.POST.get('status')
    get_selected_date = request.POST.get('get_selected_date')
    get_from_time = request.POST.get('get_from_time')
    get_to_time = request.POST.get('get_to_time')

    print(
        f"Vehicle Code : {get_vehicle_code}, Date : {get_selected_date}, From time : {get_from_time}, To time: {get_to_time}")

    # Initialize filters dictionary for ORM queries
    filters = {}
    if get_vehicle_type:
        filters['vehicle_code__vehicle_type'] = get_vehicle_type
    if get_vehicle_code:
        filters['vehicle_code__vehicle_code'] = get_vehicle_code
    if get_selected_date:
        # Filter by date range (start and end of the same day)
        selected_date = datetime.datetime.strptime(get_selected_date, "%Y-%m-%d").date()
        filters['vendor_date_time__date'] = get_selected_date

    # Call sync function if date + time range is provided
    if get_from_time and get_to_time:
        sync_records = SyncTrackerGPRSVehicleData_By_Vendor_Function(request, get_vehicle_code, get_selected_date,
                                                                     get_from_time,
                                                                     get_to_time)
        # Add success message
        if sync_records:  # Assuming sync_records indicates success
            messages.success(request, f'Sync completed successfully! {len(sync_records)} records processed.')
        else:
            messages.warning(request, 'No new records found to sync.')

    # Apply filters if any; otherwise return empty queryset
    if filters:
        # Query the filtered data and order by vehicle code and datetime
        queryset = TrackerRawData.objects.select_related('vehicle_code').filter(**filters).order_by('vendor_date_time')
    else:
        queryset = TrackerRawData.objects.none()

    # Initialize accumulators
    total_distance = 0.0
    total_working_duration = timedelta(0)
    working_points_duration = timedelta(0)

    # Tracking previous state for comparisons
    current_geom = None
    prev_geom = None
    prev_vendor_date_time = None
    prev_status = None
    prev_vendor_date_time = None

    tracker_history_json = []  # Final output list

    # if queryset response exist then loop iteration start otherwise not
    if queryset:
        # Iterate through each tracker data record
        for data in queryset:
            # vehicle_code = data.vehicle_code.vehicle_code
            current_vendor_date_time = data.vendor_date_time
            acc_status = data.acc_status
            status = data.g_status
            speed = data.speed
            mileage = data.mileage_cur_value
            vt_latitude = data.latitude
            vt_longitude = data.longitude
            location = data.location
            ext_bat_voltage = data.ext_bat_voltage
            int_bat_voltage = data.int_bat_voltage
            current_geom = data.geom

            # Handle first record of the day → check for missing data since start of day
            if prev_vendor_date_time is None:
                start_of_day = timezone.make_aware(
                    datetime.datetime.combine(current_vendor_date_time.date(), time.min)
                    # 00:00:00 timezone.get_current_timezone()
                )
                start_time = start_of_day

                time_diff = current_vendor_date_time - start_time  # 05:02:00 - 00:00:00 = 5 Hours (offline)

                # If gap > 30 minutes → log missing data
                if time_diff.total_seconds() > 30 * 60:
                    tracker_history_json.append({
                        "vehicle_code": get_vehicle_code,
                        "prev_vendor_date_time": start_time.strftime("%Y-%m-%d %H:%M:%S %p"),
                        "current_vendor_date_time": current_vendor_date_time.strftime("%Y-%m-%d %H:%M:%S %p"),
                        "status": "MISSING DATA",
                        "filter_status": "Offline",

                        "get_from_time": start_time.strftime("%H:%M:%S"),
                        "get_to_time":
                            current_vendor_date_time.strftime("%H:%M:%S"),
                    })

            # Compute time difference from previous record
            if prev_vendor_date_time:
                time_diff = current_vendor_date_time - prev_vendor_date_time

            current_geom = current_geom

            ## Distance calculation between consecutive points  (START)
            if prev_geom and current_geom and prev_status == 'Moving':
                # geodesic needs (lat, lon)
                distance = geodesic(
                    (prev_geom.y, prev_geom.x),
                    (current_geom.y, current_geom.x)
                ).meters
                total_distance += distance

            ## Distance calculation between consecutive points (END)

            ## Working hours calculation (when status is Moving) (Start)
            if prev_status is not None and prev_status == 'Moving':
                total_working_duration += time_diff
                working_points_duration = time_diff

            working_duration = format_duration_hours_minutes(working_points_duration)
            ## Working hours calculation (when status is Moving) (END)

            ## Calculate Distance in KM and Round to 1 Decimals
            distance_km = (float(total_distance) / 1000) if total_distance > 0 else 0.0
            round_distance_km = "{:.1f}".format(distance_km)

            # Detect data gaps between consecutive records (>30 min)
            if prev_vendor_date_time:
                # If gap > 30 minutes → log missing data
                if time_diff.total_seconds() > 30 * 60:
                    tracker_history_json.append({
                        "vehicle_code": get_vehicle_code,
                        "prev_vendor_date_time": prev_vendor_date_time.strftime("%Y-%m-%d %H:%M:%S %p"),
                        "current_vendor_date_time": current_vendor_date_time.strftime("%Y-%m-%d %H:%M:%S %p"),
                        "status": "MISSING DATA",
                        "filter_status": "Offline",

                        "get_from_time": prev_vendor_date_time.strftime("%H:%M:%S"),
                        "get_to_time":
                            current_vendor_date_time.strftime("%H:%M:%S"),
                    })

            # If the previous status was Moving, assign working time to that row instead of current
            if prev_status == 'Moving' and tracker_history_json:
                tracker_history_json[-1]["working_hour"] = working_duration
                tracker_history_json[-1]["distance"] = round_distance_km

            # Append current record with computed details
            tracker_history_json.append({
                "vehicle_code": get_vehicle_code,
                "vendor_date_time": current_vendor_date_time.strftime("%Y-%m-%d %H:%M:%S %p"),
                "acc_status": acc_status,
                "status": status,
                "speed": speed,
                "mileage": mileage,
                "lat": vt_latitude,
                "lon": vt_longitude,
                "location": location,
                "int_bat_voltage": int_bat_voltage,
                "ext_bat_voltage": ext_bat_voltage,
                "filter_status": "Working",

                "distance": 0.0,
                "working_hour": 0.0,
                # "status": 'N/A',  # Placeholder for status calculation
            })

            if get_status == 'Working' or not get_status:
                # Update working hours & Distance in DB if vehicle was moving
                total_seconds = working_points_duration.total_seconds()
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                total_working_hour = float(f"{hours}.{minutes:02d}")

                TrackerRawData.objects.filter(id=data.id).update(
                    distance=distance_km,
                    working_hours=total_working_hour if prev_status is not None and prev_status == 'Moving' else 0.0
                )

            # Update "previous" values for next iteration
            working_points_duration = timedelta(0)
            prev_geom = current_geom
            prev_vendor_date_time = current_vendor_date_time
            prev_status = status
        else:
            # Convert Selected date (str) into date object for condition below
            if isinstance(get_selected_date, str):
                selected_date = datetime.datetime.strptime(get_selected_date, "%Y-%m-%d").date()
            else:
                selected_date = get_selected_date

            # After loop → check for missing data until end of the day
            if prev_vendor_date_time is not None and selected_date != localdate():
                end_of_day = timezone.make_aware(
                    datetime.datetime.combine(prev_vendor_date_time.date(), time.max)
                    # 23:59:59 timezone.get_current_timezone()
                )

                end_time = end_of_day

                time_diff = end_time - prev_vendor_date_time  # 23:59:59 - 17:59:59 = 18 Hours (offline)

                # If gap > 30 minutes → log missing data
                if time_diff.total_seconds() > 30 * 60:
                    tracker_history_json.append({
                        "vehicle_code": get_vehicle_code,
                        "prev_vendor_date_time": prev_vendor_date_time.strftime("%Y-%m-%d %I:%M:%S %p"),
                        "current_vendor_date_time": end_time.strftime("%Y-%m-%d %I:%M:%S %p"),
                        "status": "MISSING DATA",
                        "filter_status": "Offline",

                        "get_from_time": prev_vendor_date_time.strftime("%H:%M:%S"),
                        "get_to_time":
                            end_time.strftime("%H:%M:%S"),
                    })

    # Final summary metrics
    working_hours = format_duration_hours_minutes(total_working_duration)
    distance_km = (float(total_distance) / 1000) if total_distance > 0 else 0.0
    round_distance_km = "{:.1f}".format(distance_km)

    ##  Threshold calculations (business rules) (START)
    working_duration_hours = int(total_working_duration.total_seconds() // 3600)
    working_duration_minutes = int((total_working_duration.total_seconds() % 3600) // 60)
    vsgs_working_hour = float(f"{working_duration_hours}.{working_duration_minutes:02d}")

    set_threshold = "No"  # default value

    if get_vehicle_type is None or get_vehicle_type == '':  # If vehicle type not selected, fetch from vehicle code
        get_vehicle_type = VehicleData.objects.filter(vehicle_code=get_vehicle_code).values_list('vehicle_type',
                                                                                                 flat=True).first()

    # Fetch threshold settings from DB for this vehicle type
    vehicle_threshold_record = VehicleThreshold.objects.filter(vehicle_type=get_vehicle_type).first()

    if vehicle_threshold_record:
        vtr_distance = vehicle_threshold_record.distance or 0
        vtr_working_hour = vehicle_threshold_record.working_hours or 0
        vtr_ignition_status = vehicle_threshold_record.ignition_status or 'No'
    else:
        vtr_distance = 0
        vtr_working_hour = 0
        vtr_ignition_status = 'No'

    if set_threshold == "No" or set_threshold is None:
        ## WHICH VEHICLE HAVING DISTANCE AND WORKING HOUR
        if vtr_ignition_status == "No" and (float(total_distance) > vtr_distance) and (
                float(vsgs_working_hour) >= vtr_working_hour):
            set_threshold = "Yes"

        ## VEHICLE ONLY WORKING HOUR
        elif vtr_ignition_status == "Yes" and (float(vsgs_working_hour) > vtr_working_hour):
            set_threshold = "Yes"
    ## Calculate Threshold (END)

    # Default selected date → today if not provided (on initially page load or no date selected in frontend)
    if not get_selected_date:
        get_selected_date = localdate().strftime('%Y-%m-%d')

    vehicle_type_list = VehicleData.objects.filter(status="Active").values('vehicle_type').annotate(
        count=Count('id')).order_by('vehicle_type')

    ### VEHICLE DATA DETAIL WITH VEHICLE STATUS AND VEHICLE TYPE
    cmd_vehicle_data = list(
        # VehicleData.objects.filter(**filters_vehicle_data)
        VehicleData.objects.values('vehicle_code', 'vehicle_type', 'register_no', 'chasis_no')
        .annotate(
            pitb_code=Coalesce('pitb_code', Value('Waiting'), output_field=CharField())
        )
        .order_by('register_no')
        .distinct()
    )

    # Filter the JSON/Table using get_status (Working/Offline)
    if get_status:
        tracker_history_json = [
            row for row in tracker_history_json
            if row["filter_status"] == get_status
        ]

    # Context data passed to template
    context = {
        'total_distance': round_distance_km,
        'working_hours': working_hours,
        'threshold': set_threshold,

        'selected_vehicle_type': get_vehicle_type,
        'selected_vehicle_code': get_vehicle_code,
        'selected_status': get_status,
        'get_selected_date': get_selected_date,

        'tracker_history_json': tracker_history_json,
        'vehicle_type_list': vehicle_type_list,

        'cmd_vehicle_data': cmd_vehicle_data,
    }
    # Render the response
    return render(request, template_name, context)
